from fnmatch import fnmatch
from itertools import count
from lib2to3.refactor import RefactoringTool
from msilib.schema import ListBox
import os #시스템 접근
import re #정규식
import csv
import glob
import datetime

import atexit #프로그램 종료시 사용될 문구
import subprocess #메모장 실행
import shutil
from tkinter import font
from turtle import undo #파일 지우기

import openpyxl #엑셀
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side


from fractions import Fraction #유리수 값 사용

import tkinter
import tkinter.ttk as ttk
import tkinter.messagebox
from tkinter import *
from tkinter import filedialog
#from tkinter.tix import *
from TkinterDnD2 import DND_FILES, TkinterDnD

import win32com.client as win32  # 한/글 열 수 있는 모듈
import win32gui  # 창 숨기기 위한 모듈
import webbrowser #웹 열기

treeColumn_header = ['번호', '개수', '요리명', '자격증 분류', '시간대']
treeItem_List = []


restoreNum = 0 #refine된 재료 개수 카운트(출력 전용)
restoreNumList = [] #refine된 재료 개수 카운트 리스트(출력 전용)

DateSetName = [] #날짜 종류

#============================================================
#[프로그램] 종료시 실행 -Done
#============================================================
def exit_Function():
    global USERNAMEDIR_VAR
    closefile = open("./res/sys/systemp.txt",'w',encoding='utf-8')
    print(DARKMODE_VAR.get(),file=closefile)
    print(USERNAMEDIR_VAR,file=closefile)
    print(StandardCheckbox_Var.get(),file= closefile)
    print(NoteCheckBox_Var.get(),file= closefile)
    print(RemoveTimeCheckbox_Var.get(),file= closefile)
    print(RemoveDateCheckbox_Var.get(),file= closefile)
    print(RemoveReduplicationCheckbox_Var.get(),file= closefile)
    print(TimeKind_Var.get(),file= closefile)
    print(ToolTipCheckbox_Var.get(),file= closefile)
    print(DATE_TOTAL_IMPORT_COMBO_INIT,file=closefile)
    print(DATE_TOTAL_EXPORT_COMBO_INIT,file=closefile)
    print(Program_Save_Dir_Var.get(),file=closefile)
    print(Program_Result_Save_Dir_Var.get(),file=closefile)
    print(TotalDate_Result_Save_Dir_Var.get(),file=closefile)
    print(RESULT_COMBO_VAR, file=closefile)
    print(Program_UseFile_Dir_Var.get(),file=closefile)
    closefile.close()

#============================================================
#[프로그램]툴팁
#============================================================
class CreateToolTip(object):
    '''
    create a tooltip for a given widget
    '''
    def __init__(self, widget, text='widget info'):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.close)


    def enter(self, event=None):
        if ToolTipCheckbox_Var.get() == 1:
            x = y = 0
            x, y, cx, cy = self.widget.bbox("insert")
            x += self.widget.winfo_rootx() + 30
            y += self.widget.winfo_rooty() + 30
            # creates a toplevel window
            self.tw = Toplevel(self.widget)
            # Leaves only the label and removes the app window
            self.tw.wm_overrideredirect(True)
            self.tw.wm_geometry("+%d+%d" % (x, y))
            label = Label(self.tw, text=self.text, justify='left',
                        background='lightyellow', relief='solid', borderwidth=1,
                        font=("times", "8", "normal"))
            label.pack(ipadx=1)
        else:
            pass
    def close(self, event=None):
        if ToolTipCheckbox_Var.get() == 1:
            if self.tw:
                self.tw.destroy()

        else:
            pass
#============================================================
#[프로그램][메인 - 메뉴바 - 파일]파일 변환창 실행중 취소 -Done
#============================================================
def SubWindow_Convert_InfoWindow_Close_Cancle():
    infowindow.destroy()
    tkinter.messagebox.showwarning("확인", "취소 되었습니다.")

#============================================================
#[메인 - 메뉴바 - 파일]파일 변환 hwp -> txt
#============================================================
def BTN_ConvertHWP():
    pLog.append_log("버튼 동작 실행: ", "파일 변환")
    root_dir = filedialog.askdirectory(initialdir=Program_UseFile_Dir_Var.get())
    if not root_dir:
        return
    SAVE_DIR = os.getcwd() + '\\TempFileList\\'
    
    #if root_dir[-4:] == ".hwp":
    hwp = win32.gencache.EnsureDispatch('HWPFrame.HwpObject')  # 한/글 열기
    hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글')  # 해당 윈도우의 핸들값 찾기

    win32gui.ShowWindow(hwnd, 0)#창 백그라운드에서 실행
    hwp.RegisterModule('FilePathCheckDLL', 'FilePathCheckerModule')

    LoadingfileText = StringVar()

    global infowindow
    infowindow = Toplevel()
    infowindow.title("변환 정보")
    infowindow.geometry("260x140+400+400")
    infowindow.resizable(False, False) #창 사이즈 변경 불가능
    infowindow.wm_attributes("-topmost", 1) #창 맨앞으로

    Progress_Info_Text = Label(infowindow, text="파일 변환 진행중")
    Progress_Info_Text.pack()

    Progress_Info_Text1 = Label(infowindow, textvariable=LoadingfileText)
    Progress_Info_Text1.pack()
    Progress_Info_Text1.place(x=80,y=45)

    OkBtn = Button(infowindow, text="확인", command=infowindow.destroy)
    OkBtn.pack()
    OkBtn.place(x=80,y=85)

    CancleBtn = Button(infowindow, text="취소", command=SubWindow_Convert_InfoWindow_Close_Cancle)
    CancleBtn.pack()
    CancleBtn.place(x=130,y=85)

    curProg = DoubleVar()
    progNum = 0
    progressbar= ttk.Progressbar(infowindow, maximum=100, variable=curProg)
    progressbar.pack()

    file_count = sum(len(files1) for _, _, files1 in os.walk(root_dir))
    
    for (root_path, dirs, files) in os.walk(root_dir):
        print("# root : " + root_path)

        if len(dirs) > 0:
            for dir_name in dirs:
                print("dir: " + dir_name)
                
        if len(files) > 0:
            for file_name in files:
                if file_name[-3:] == "hwp" or file_name[-3:] == "HWP" or file_name[-3:] == "Hwp":
                    #print("file: " + file_name)
                    LoadingfileText.set(file_name)
                    hwp.Open(os.path.join(root_path, file_name))  # 한/글로 열기
                    #hwp에 ,을 없애줘야 split(",")을 사용할수 있음
                    hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                    option=hwp.HParameterSet.HFindReplace
                    option.FindString = ","
                    option.ReplaceString = ""
                    option.IgnoreMessage = 1
                    hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)

                    hwp.HAction.GetDefault("AllReplace", hwp.HParameterSet.HFindReplace.HSet)
                    option=hwp.HParameterSet.HFindReplace
                    option.FindString = "^n"
                    option.ReplaceString = "."
                    option.IgnoreMessage = 1
                    hwp.HAction.Execute("AllReplace", hwp.HParameterSet.HFindReplace.HSet)

                    TempFileName = file_name[:-4]
                    hwp.SaveAs(os.path.join(SAVE_DIR, TempFileName + ".txt"),"TEXT")

                    curProg.set(progNum / file_count * 100)
                    progressbar.update()
                    progNum += 1
                else:
                    pass
        if progNum / file_count * 100 > 99:
            Progress_Info_Text.config(text="파일 변환 완료")

    win32gui.ShowWindow(hwnd, 5)  # 다시 숨겼던 한/글 창을 보여주고,
    hwp.XHwpDocuments.Close(isDirty=False)  # 열려있는 문서가 있다면 닫아줘(저장할지 물어보지 말고)
    hwp.Quit()  # 한/글 종료


#============================================================
#[요일별 종합] 날짜 세팅
#============================================================
Total_treeNumCount = 1
DateSetCount = 0
def Total_DaySetting():
    pLog.append_log("버튼 동작 실행: ", "날짜 세팅")
    os.startfile('.\\res\\sys\\sys_date.txt')

#============================================================
#[요일별 종합 - 버튼] 결과 확인
#============================================================
def Total_Result():
    pLog.append_log("버튼 동작 실행: ", "결과 출력")
    global DateSetName
    global Total_listbox
    global Total_list_Combo_Import
    global Total_list_Combo_Export
    
    TEMP_DIR = os.getcwd() + '\\TempFileList\\'
    pLog.append_log("TEMP_DIR", TEMP_DIR)
    #임시 파일 지우기
    for delname in glob.glob(TEMP_DIR+"_Total*.txt"):
        os.remove(delname)
    #전체적인 맥락
    #1. 요일 종합 목록에서 파일을 불러오는 부분[빈 값 추가, 확장자 확인]
    #2. 파일 경로를 따라 각 파일에 맞게 정제해서 불러옴listText
    #3. 파일을 내보내기 위해 필요한 각종 정보들을 만듬
    #4. 실질적인 내보내기
    if Total_list_Combo_Import.get() == "한글(.hwp)":
        if DateSetCount > Total_listbox.size():
            for x in range(0,DateSetCount-Total_listbox.size()):
                Total_listbox.insert(END,"Empty")

        TempTotalList = []
        for x in range(0,Total_listbox.size()):
            if Total_listbox.get(x) == "Empty":
                TempTotalList.append("")
            else:
                TempTotalList.append(Program_Result_Save_Dir_Var.get()+"\\"+Total_listbox.get(x))
        pLog.append_log("TempTotalList: ", TempTotalList)
        for x in range(0, len(TempTotalList)):
            if not TempTotalList[x]:
                pass
            else:
                if TempTotalList[x][-4:] != ".hwp":
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
        print(TempTotalList)
        
        if len(TempTotalList)>0:
            #Total_list_Combo['values']=("한글(.hwp)", "엑셀(.xlsx)", "텍스트(.txt)")
            convertFile = []
            hwp = win32.gencache.EnsureDispatch('HWPFrame.HwpObject')  # 한/글 열기
            hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글')  # 해당 윈도우의 핸들값 찾기

            win32gui.ShowWindow(hwnd, 0)#창 백그라운드에서 실행
            hwp.RegisterModule('FilePathCheckDLL', 'FilePathCheckerModule')
            txtFileCount = 1
            for TempTotal in TempTotalList:
                file_name = TempTotal[TempTotal.rfind("\\")+1:]
                FolderPath = TempTotal[:TempTotal.find(file_name)]

                hwp.Open(os.path.join(FolderPath, file_name))
                TempFileName = file_name[:-4]

                hwp.SaveAs(os.path.join(TEMP_DIR, "_Total"+str(txtFileCount)+"_"+TempFileName + ".txt"),"TEXT")
                convertFile.append(TEMP_DIR + "_Total"+str(txtFileCount)+"_" + TempFileName + ".txt")#저장된 임시파일을 불러오기위해
                txtFileCount+=1

            win32gui.ShowWindow(hwnd, 5)
            hwp.XHwpDocuments.Close(isDirty=False)
            hwp.Quit()  # 한/글 종료
                
            #변환했던 임시파일 문자열로 읽어들이기
            listText = []
            for x in range(0,len(convertFile)):
                with open(convertFile[x]) as f:
                    lines = f.readlines()

                lines = [line.rstrip('\n') for line in lines]#엔터 제거
                strText = str(lines)

                strText = strText.replace("\'비고\',", "\'**비고**\',")
                orignalText = strText[strText.find("**비고**")+8:]
                orignalText=orignalText.replace("'","")
                Local_listText=orignalText.split(",")
                listText.append(Local_listText)

            LocalTimeList = [] #각 파일별 [오전,오후,저녁] 재료 항목의 개수를 모을 리스트
            #시간값이 포함된 결과물인지 확인용(사용자가 등록한 항목 개수 이상이면 시간값 포함)
            for x in range(0,len(listText)):#불러온 파일들
                incount = 0 #단순 카운트용 변수
                
                timeList = []#오전,오후,저녁별 재료 개수 담아둘 변수
                lenListNum = 0
                #재료 길이 구하는 반복문
                for y in listText[x]: #[번호,이름,규격,단위,개수,비고]
                    if incount % 6 == 0:#번호만 추출
                        onlyNumber = str(y.replace(" ",""))
                        if onlyNumber.isdigit() == True:
                            lenListNum+=1
                        else:
                            lenListNum+=0
                    incount+=1

                incount = 0
                allList = [] #전부 박아둘 리스트
                #파일의 시간 분리 하는 반복문
                for y in listText[x]: #[번호,이름,규격,단위,개수,비고]
                    if incount % 6 == 0:#번호만 접근
                        onlyNumber = str(y.replace(" ",""))
                        if not onlyNumber:
                            allList.append(0)
                            pastNum = 1
                        else:
                            allList.append(int(onlyNumber))
                    incount +=1

                leftList = [] #리스트의 마지막을 제외한 재료 번호 중 끝값들
                pastNum = 1
                justCount = 0
                for x in range(0,len(allList)):

                    if allList[x] == pastNum:
                        if justCount == 0:
                            pass
                        else:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(allList[x])
                    #기준값이 이전값 보다 크면 정상
                    if allList[x] > pastNum:
                        pastNum += 1
                    #기준값이 이전값 보다 작으며 비정상 = 시간대가 달라진다.
                    if allList[x] < pastNum:
                        if justCount == 0:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(0)
                            continue
                        
                        leftList.append(pastNum)
                        pastNum = allList[x]
                        if allList[x] == 0:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(0)
                        
                    justCount += 1
                timeList.append(leftList)

                rightList = allList[-1]
                leftList.append(rightList)

                LocalTimeList.append(leftList)
            print(LocalTimeList)

            #0을 1로 변경
            for x in LocalTimeList:
                for y in range(0,len(x)):
                    if x[y] == 0:
                        x[y] = 1
            
            #각 파일별 시간대 개수
            maxSizeOfLocalTimeList = 0
            for x in LocalTimeList:
                if len(x) > maxSizeOfLocalTimeList:
                    maxSizeOfLocalTimeList = len(x)
            print(maxSizeOfLocalTimeList)
            
            #최대 시간대 개수 로 통일
            for x in range(0, len(LocalTimeList)):
                if len(LocalTimeList[x]) <maxSizeOfLocalTimeList:
                    for y in range(0,maxSizeOfLocalTimeList-len(LocalTimeList[x])):
                        LocalTimeList[x].append(1)
            
            #각 파일별 시간대별 최대값 [오전 최댓값, 오후 최댓값, 저녁 최댓값...]
            MaxLocalTime = []        
            for x in range(0,maxSizeOfLocalTimeList):
                MaxLocalTime.append(max(t[x] for t in LocalTimeList))

            if Total_list_Combo_Export.get() == "한글(.hwp)":
                filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".hwp", filetypes=[("Hwp files", "*.hwp")])
                hwp = win32.gencache.EnsureDispatch('HWPFrame.HwpObject')  # 한/글 열기
                hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글')  # 해당 윈도우의 핸들값 찾기

                #win32gui.ShowWindow(hwnd,1)#창 백그라운드에서 실행
                hwp.RegisterModule('FilePathCheckDLL', 'FilePathCheckerModule')
                hwp.XHwpWindows.Item(0).Visible = True  # 숨김해제

                #여백 세팅
                hwp.HAction.GetDefault("ModifySection", hwp.HParameterSet.HSecDef.HSet)
                hwp.HParameterSet.HSecDef.PageDef.LeftMargin = hwp.MiliToHwpUnit(20.0)
                hwp.HParameterSet.HSecDef.PageDef.TopMargin = hwp.MiliToHwpUnit(15.0)
                hwp.HParameterSet.HSecDef.PageDef.RightMargin = hwp.MiliToHwpUnit(20.0)
                hwp.HParameterSet.HSecDef.PageDef.BottomMargin = hwp.MiliToHwpUnit(15.0)
                hwp.HParameterSet.HSecDef.PageDef.HeaderLen = hwp.MiliToHwpUnit(10.0)
                hwp.HParameterSet.HSecDef.PageDef.FooterLen = hwp.MiliToHwpUnit(10.0)
                hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyClass", 24)
                hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyTo", 2)
                hwp.HAction.Execute("ModifySection", hwp.HParameterSet.HSecDef.HSet)
                
                hwp.HAction.Run("ParagraphShapeAlignCenter")
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = "발주서 종합"
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("SelectAll")
                hwp_fontSetting(hwp,"HY헤드라인M",22,1)

                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("BreakPara")
                #열 - 가로, 행 - 세로
                hwp.HAction.GetDefault("TableCreate", hwp.HParameterSet.HTableCreation.HSet)  # 표 생성 시작
                hwp.HParameterSet.HTableCreation.Rows = maxSizeOfLocalTimeList+1  # 행 갯수
                hwp.HParameterSet.HTableCreation.Cols = DateSetCount  # 열 갯수
                hwp.HParameterSet.HTableCreation.WidthType = 2  # 너비 지정(0:단에맞춤, 1:문단에맞춤, 2:임의값)
                hwp.HParameterSet.HTableCreation.HeightType = 0  # 높이 지정(0:자동, 1:임의값)
                #hwp.HParameterSet.HTableCreation.WidthValue = hwp.MiliToHwpUnit(148.0)  # 표 너비
                #hwp.HParameterSet.HTableCreation.HeightValue = hwp.MiliToHwpUnit(150)  # 표 높이
                hwp.HParameterSet.HTableCreation.CreateItemArray("ColWidth", DateSetCount)  # 열 5개 생성
                for x in range(0,DateSetCount):
                    hwp.HParameterSet.HTableCreation.ColWidth.SetItem(x, hwp.MiliToHwpUnit(30.0))  # 1열(가로 크기)

                hwp.HParameterSet.HTableCreation.CreateItemArray("RowHeight", maxSizeOfLocalTimeList+1)  # 행 5개 생성
                for x in range(0,maxSizeOfLocalTimeList+1):
                    if x == 0:
                        hwp.HParameterSet.HTableCreation.RowHeight.SetItem(x, hwp.MiliToHwpUnit(15.0))  # 1행
                    else:
                        hwp.HParameterSet.HTableCreation.RowHeight.SetItem(x, hwp.MiliToHwpUnit(50.0))  # 2행

                hwp.HParameterSet.HTableCreation.TableProperties.TreatAsChar = 1  # 글자처럼 취급
                hwp.HParameterSet.HTableCreation.TableProperties.Width = hwp.MiliToHwpUnit(148)  # 표 너비
                hwp.HAction.Execute("TableCreate", hwp.HParameterSet.HTableCreation.HSet)  # 위 코드 실행

                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                for x in range(0,DateSetCount-1): 
                    hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("ParagraphShapeAlignCenter")

                hwp_fontSetting(hwp,"HY헤드라인M",16,1)

                hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
                hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")

                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("TableCellBlockExtend")


                for y in range(0, DateSetCount):
                    for x in range(0,maxSizeOfLocalTimeList):
                        hwp.HAction.Run("TableLowerCell")

                    hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                    hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
                    hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
                    hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
                    hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
                    hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                    hwp.HAction.Run("TableRightCell")

                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("Cancel")

                
                for x in range(0,len(DateSetName)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = DateSetName[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                    hwp.HAction.Run("TableRightCell")
                
                for y in range(0,len(listText)):#파일 개수
                    groupCount = 0#카운트용 변수
                    subgroupCount = 0#재료 개수
                    separateNum = 0

                    for x in listText[y]:#일별로 파일에 있는 항목들

                        if groupCount % 6 == 1: #재료명
                            hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                            hwp.HAction.Run("CharShapeBold")
                            hwp.HParameterSet.HInsertText.Text = x
                            hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        if groupCount % 6 == 3: #단위
                            hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                            hwp.HAction.Run("CharShapeBold")
                            hwp.HParameterSet.HInsertText.Text = x
                            hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        if groupCount % 6 == 4: #개수
                            hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                            hwp.HParameterSet.HInsertText.Text = x
                            hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                            hwp.HAction.Run("BreakPara")
                        if groupCount % 6 == 5: #비고
                            subgroupCount+=1
                                            
                        if subgroupCount >= LocalTimeList[y][separateNum]:
                            hwp.HAction.Run("TableLowerCell")
                            separateNum+=1
                            subgroupCount= 0

                        groupCount += 1
                    
                    hwp.Run("MoveDocBegin")
                    hwp.Run("SelectCtrlFront")
                    hwp.HAction.Run("ShapeObjTableSelCell")

                    for z in range(0,y+1):
                        hwp.HAction.Run("TableRightCell")
                    
                    hwp.HAction.Run("TableLowerCell")
                    hwp.HAction.Run("Cancel")

                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")
                try:
                    hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
                    hwp.HParameterSet.HShapeObject.TreatAsChar = 0#글자취급x
                    hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
                except:
                    print("서버")
                    #아마도 범위 초과 설정 해서 그런듯
                hwp.MovePos(3)
                hwp.SaveAs(filename)

            elif Total_list_Combo_Export.get()=="엑셀(.xlsx)":
                filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                if not filename:
                    return
                write_wb = openpyxl.Workbook()
                write_ws = write_wb.active
                write_ws['A1'].font = Font(size=20,bold=True)
                write_ws['A1'] = '발주서 종합'
                write_ws['A1'].border = Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))

                write_ws.append([""])


                for col in range(1,(DateSetCount*3)+1):
                    if col % 3 == 1:
                        write_ws.column_dimensions[get_column_letter(col)].width = 10
                    else:
                        write_ws.column_dimensions[get_column_letter(col)].width = 5

                refineDateSetName = []
                for x in DateSetName:
                    refineDateSetName.append(x)
                    refineDateSetName.append("")
                    refineDateSetName.append("")
                
                write_ws.append(refineDateSetName)


                #엑셀 전용으로 값을 조정
                excelTemplistText = []
                for y in range(0,len(listText)):#파일 개수
                    groupCount = 0#카운트용 변수

                    separateNum = 0
                    SubexcelTemplistText = []
                    emptyCellCount = 0
                    prevSeparategroupCount = 0
                    for x in listText[y]:#일별로 파일에 있는 항목들
                        separategroupCount = groupCount/6
                        if groupCount % 6 == 1: #재료명
                            SubexcelTemplistText.append(x.lstrip())
                        if groupCount % 6 == 3: #단위
                            SubexcelTemplistText.append(x.lstrip())
                        if groupCount % 6 == 4: #개수
                            SubexcelTemplistText.append(x.lstrip())
                        
                        if separategroupCount-prevSeparategroupCount == LocalTimeList[y][emptyCellCount]:
                            if LocalTimeList[y][emptyCellCount] < MaxLocalTime[emptyCellCount]:
                                for i in range(0,MaxLocalTime[emptyCellCount]-LocalTimeList[y][emptyCellCount]):
                                    SubexcelTemplistText.append("")
                                    SubexcelTemplistText.append("")
                                    SubexcelTemplistText.append("")
                            prevSeparategroupCount += separategroupCount
                            emptyCellCount +=1
                        groupCount += 1
                    excelTemplistText.append(SubexcelTemplistText)

                tempCol = 1
                for j in range(0,len(excelTemplistText)):
                    localCount = 0
                    rowJ = 4
                    colJ = tempCol

                    for i in excelTemplistText[j]:
                        write_ws.cell(column=colJ,row=rowJ,value=i)
                        if localCount % 3 == 2:
                            colJ = tempCol
                            rowJ +=1
                            if rowJ != 4:
                                colJ-=1
                        localCount +=1
                        colJ+=1
                    
                    tempCol += 3

                write_ws.merge_cells("A1:"+get_column_letter(DateSetCount*3)+"2")#TODO:5를 SetCount로 변경
                mergeCount = 1
                for x in range(1,(DateSetCount*3)+1):
                    write_ws.merge_cells(get_column_letter(mergeCount)+"3:"+get_column_letter(mergeCount+2)+"3")#TODO:5를 SetCount로 변경
                    mergeCount+=3

                write_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

                for x in range(1,(DateSetCount*3)+1):
                    write_ws[get_column_letter(x)+"3"].font = Font(size=15,bold=True)
                    write_ws[get_column_letter(x)+"3"].alignment = Alignment(horizontal='center', vertical='center')

                startChr = 65
                for x in range(0,DateSetCount):
                    set_border(write_ws, chr(startChr)+str(3)+":"+chr(startChr+2)+str(3),"medium")
                    startChr += 3

                startChr = 65
                for y in range(0, DateSetCount):
                    startNum = 4
                    for x in range(0,len(MaxLocalTime)):
                        set_border(write_ws, chr(startChr)+str(startNum)+":"+chr(startChr+2)+str(startNum+MaxLocalTime[x]),"medium")
                        startNum += MaxLocalTime[x]
                    startChr += 3
                write_wb.save(filename)

                os.startfile(filename)

            elif Total_list_Combo_Export.get()=="텍스트(.txt)":
                filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".txt", filetypes=[("TXT files", "*.txt")])
                if not filename:
                    return
                data = open(filename, 'w', encoding="UTF8")

                print("발주서 종합\n",file = data)
                
                txtTemplistText = []
                for y in range(0,len(listText)):#파일 개수
                    
                    groupCount = 0#카운트용 변수
                    SubTxtTemplistText = []

                    for x in listText[y]:#일별로 파일에 있는 항목들

                        if groupCount % 6 == 1: #재료명
                            SubTxtTemplistText.append(x.lstrip())
                        if groupCount % 6 == 3: #단위
                            SubTxtTemplistText.append(x.lstrip())
                        if groupCount % 6 == 4: #개수
                            SubTxtTemplistText.append(x.lstrip())

                        groupCount += 1

                    txtTemplistText.append(SubTxtTemplistText)

                print("====================",file=data)
                for x in range(0,DateSetCount):
                    print(DateSetName[x],file=data)
                    print("====================",file=data)
                    for startNum in range(0,len(txtTemplistText[x]),3):
                        print(txtTemplistText[x][startNum],txtTemplistText[x][startNum+1],txtTemplistText[x][startNum+2], file=data)
                    print("====================\n",file=data)

                os.startfile(filename)

    elif Total_list_Combo_Import.get() == "엑셀(.xlsx)":
        if DateSetCount > Total_listbox.size():
            for x in range(0,DateSetCount-Total_listbox.size()):
                Total_listbox.insert(END,"Empty")

        TempTotalList = []
        for x in range(0,Total_listbox.size()):
            if Total_listbox.get(x) == "Empty":
                TempTotalList.append("")
            else:
                TempTotalList.append(Program_Result_Save_Dir_Var.get()+"\\"+Total_listbox.get(x))
        pLog.append_log("TempTotalList: ", TempTotalList)
        for x in range(0, len(TempTotalList)):
            if not TempTotalList[x]:
                pass
            else:
                if TempTotalList[x][-5:] != ".xlsx" :
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
        print(TempTotalList)
        print(len(TempTotalList))
        
        if len(TempTotalList)>0:
            listText = []
            for x in range(0,len(TempTotalList)):
                subListText = []
                if not TempTotalList[x]:
                    subListText.append("")
                    listText.append(subListText)
                    continue
                else:
                    wb1 = openpyxl.load_workbook(TempTotalList[x])
                    sheet = wb1['Sheet'] 

                    Sheet_row = sheet.max_row
                    for data in sheet['A13':'F'+str(Sheet_row)]: 
                        for cell in data: 
                            if cell.value == None:
                                subListText.append("")
                            else:
                                subListText.append(str(cell.value))
                wb1.close()
                listText.append(subListText)
            print(listText)
            LocalTimeList = [] #각 파일별 [오전,오후,저녁] 재료 항목의 개수를 모을 리스트
            #시간값이 포함된 결과물인지 확인용(사용자가 등록한 항목 개수 이상이면 시간값 포함)
            for x in range(0,len(listText)):#불러온 파일들
                incount = 0 #단순 카운트용 변수
                
                timeList = []#오전,오후,저녁별 재료 개수 담아둘 변수
                lenListNum = 0
                for y in listText[x]: #[번호,이름,규격,단위,개수,비고]
                    if incount % 6 == 0:#번호만 추출
                        onlyNumber = str(y.replace(" ",""))
                        if onlyNumber.isdigit() == True:
                            lenListNum+=1
                        else:
                            lenListNum+=0
                    incount+=1
                print(lenListNum)#총 재료의 길이 구하는거

                incount = 0
                allList = [] #전부 박아둘 리스트
                for y in listText[x]: #[번호,이름,규격,단위,개수,비고]
                    if incount % 6 == 0:#번호만 접근
                        onlyNumber = str(y.replace(" ",""))
                        
                        if not onlyNumber:
                            allList.append(0)
                            pastNum = 1
                        else:
                            allList.append(int(onlyNumber))
                    incount +=1
                print("AllList: "+str(allList))

                leftList = [] #리스트의 마지막을 제외한 재료 번호 중 끝값들
                pastNum = 1
                justCount = 0
                for x in range(0,len(allList)):

                    if allList[x] == pastNum:
                        if justCount == 0:
                            pass
                        else:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(allList[x])
                    #기준값이 이전값 보다 크면 정상
                    if allList[x] > pastNum:
                        pastNum += 1
                    #기준값이 이전값 보다 작으며 비정상 = 시간대가 달라진다.
                    if allList[x] < pastNum:
                        if justCount == 0:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(0)
                            continue
                        
                        leftList.append(pastNum)
                        pastNum = allList[x]
                        if allList[x] == 0:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(0)
                        
                    justCount += 1
                timeList.append(leftList)

                rightList = allList[-1]
                leftList.append(rightList)

                LocalTimeList.append(leftList)
            print(LocalTimeList)

            #0을 1로 변경
            for x in LocalTimeList:
                for y in range(0,len(x)):
                    if x[y] == 0:
                        x[y] = 1
            
            #각 파일별 시간대 개수
            maxSizeOfLocalTimeList = 0
            for x in LocalTimeList:
                if len(x) > maxSizeOfLocalTimeList:
                    maxSizeOfLocalTimeList = len(x)
            print(maxSizeOfLocalTimeList)
            
            #최대 시간대 개수 로 통일
            for x in range(0, len(LocalTimeList)):
                if len(LocalTimeList[x]) <maxSizeOfLocalTimeList:
                    for y in range(0,maxSizeOfLocalTimeList-len(LocalTimeList[x])):
                        LocalTimeList[x].append(1)
            print(LocalTimeList)
            #각 파일별 시간대별 최대값 [오전 최댓값, 오후 최댓값, 저녁 최댓값...]
            MaxLocalTime = []        
            for x in range(0,maxSizeOfLocalTimeList):
                MaxLocalTime.append(max(t[x] for t in LocalTimeList))

        if Total_list_Combo_Export.get() == "한글(.hwp)":
            filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".hwp", filetypes=[("Hwp files", "*.hwp")])
            hwp = win32.gencache.EnsureDispatch('HWPFrame.HwpObject')  # 한/글 열기
            hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글')  # 해당 윈도우의 핸들값 찾기

            #win32gui.ShowWindow(hwnd,1)#창 백그라운드에서 실행
            hwp.RegisterModule('FilePathCheckDLL', 'FilePathCheckerModule')
            hwp.XHwpWindows.Item(0).Visible = True  # 숨김해제

            #여백 세팅
            hwp.HAction.GetDefault("ModifySection", hwp.HParameterSet.HSecDef.HSet)
            hwp.HParameterSet.HSecDef.PageDef.LeftMargin = hwp.MiliToHwpUnit(20.0)
            hwp.HParameterSet.HSecDef.PageDef.TopMargin = hwp.MiliToHwpUnit(15.0)
            hwp.HParameterSet.HSecDef.PageDef.RightMargin = hwp.MiliToHwpUnit(20.0)
            hwp.HParameterSet.HSecDef.PageDef.BottomMargin = hwp.MiliToHwpUnit(15.0)
            hwp.HParameterSet.HSecDef.PageDef.HeaderLen = hwp.MiliToHwpUnit(10.0)
            hwp.HParameterSet.HSecDef.PageDef.FooterLen = hwp.MiliToHwpUnit(10.0)
            hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyClass", 24)
            hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyTo", 2)
            hwp.HAction.Execute("ModifySection", hwp.HParameterSet.HSecDef.HSet)
            
            hwp.HAction.Run("ParagraphShapeAlignCenter")
            hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
            hwp.HParameterSet.HInsertText.Text = "발주서 종합"
            hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
            hwp.HAction.Run("SelectAll")
            hwp_fontSetting(hwp,"HY헤드라인M",22,1)

            hwp.HAction.Run("MoveRight")
            hwp.HAction.Run("BreakPara")
            #열 - 가로, 행 - 세로
            hwp.HAction.GetDefault("TableCreate", hwp.HParameterSet.HTableCreation.HSet)  # 표 생성 시작
            hwp.HParameterSet.HTableCreation.Rows = maxSizeOfLocalTimeList+1  # 행 갯수
            hwp.HParameterSet.HTableCreation.Cols = DateSetCount  # 열 갯수
            hwp.HParameterSet.HTableCreation.WidthType = 2  # 너비 지정(0:단에맞춤, 1:문단에맞춤, 2:임의값)
            hwp.HParameterSet.HTableCreation.HeightType = 0  # 높이 지정(0:자동, 1:임의값)
            #hwp.HParameterSet.HTableCreation.WidthValue = hwp.MiliToHwpUnit(148.0)  # 표 너비
            #hwp.HParameterSet.HTableCreation.HeightValue = hwp.MiliToHwpUnit(150)  # 표 높이
            hwp.HParameterSet.HTableCreation.CreateItemArray("ColWidth", DateSetCount)  # 열 5개 생성
            for x in range(0,DateSetCount):
                hwp.HParameterSet.HTableCreation.ColWidth.SetItem(x, hwp.MiliToHwpUnit(30.0))  # 1열(가로 크기)

            hwp.HParameterSet.HTableCreation.CreateItemArray("RowHeight", maxSizeOfLocalTimeList+1)  # 행 5개 생성
            for x in range(0,maxSizeOfLocalTimeList+1):
                if x == 0:
                    hwp.HParameterSet.HTableCreation.RowHeight.SetItem(x, hwp.MiliToHwpUnit(15.0))  # 1행
                else:
                    hwp.HParameterSet.HTableCreation.RowHeight.SetItem(x, hwp.MiliToHwpUnit(50.0))  # 2행

            hwp.HParameterSet.HTableCreation.TableProperties.TreatAsChar = 1  # 글자처럼 취급
            hwp.HParameterSet.HTableCreation.TableProperties.Width = hwp.MiliToHwpUnit(148)  # 표 너비
            hwp.HAction.Execute("TableCreate", hwp.HParameterSet.HTableCreation.HSet)  # 위 코드 실행

            hwp.HAction.Run("TableCellBlock")
            hwp.HAction.Run("TableCellBlockExtend")
            for x in range(0,DateSetCount-1): 
                hwp.HAction.Run("TableRightCell")
            hwp.HAction.Run("ParagraphShapeAlignCenter")

            hwp_fontSetting(hwp,"HY헤드라인M",16,1)

            hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
            hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
            hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
            hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
            hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
            hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

            hwp.Run("MoveDocBegin")
            hwp.Run("SelectCtrlFront")

            hwp.HAction.Run("ShapeObjTableSelCell")
            hwp.HAction.Run("TableCellBlockExtend")


            for y in range(0, DateSetCount):
                for x in range(0,maxSizeOfLocalTimeList):
                    hwp.HAction.Run("TableLowerCell")

                hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
                hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                hwp.HAction.Run("TableRightCell")

            hwp.Run("MoveDocBegin")
            hwp.Run("SelectCtrlFront")
            hwp.HAction.Run("ShapeObjTableSelCell")
            hwp.HAction.Run("Cancel")

            
            for x in range(0,len(DateSetName)):
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = DateSetName[x]
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                hwp.HAction.Run("TableRightCell")
            
            for y in range(0,len(listText)):#파일 개수
                groupCount = 0#카운트용 변수
                subgroupCount = 0#재료 개수
                separateNum = 0

                for x in listText[y]:#일별로 파일에 있는 항목들

                    if groupCount % 6 == 1: #재료명
                        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HAction.Run("CharShapeBold")
                        hwp.HParameterSet.HInsertText.Text = x + " "
                        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if groupCount % 6 == 3: #단위
                        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HAction.Run("CharShapeBold")
                        hwp.HParameterSet.HInsertText.Text = x + " "
                        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if groupCount % 6 == 4: #개수
                        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HParameterSet.HInsertText.Text = x + " "
                        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HAction.Run("BreakPara")
                    if groupCount % 6 == 5: #비고
                        subgroupCount+=1
                                        
                    if subgroupCount >= LocalTimeList[y][separateNum]:
                        hwp.HAction.Run("TableLowerCell")
                        separateNum+=1
                        subgroupCount= 0

                    groupCount += 1
                
                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")

                for z in range(0,y+1):
                    hwp.HAction.Run("TableRightCell")
                
                hwp.HAction.Run("TableLowerCell")
                hwp.HAction.Run("Cancel")

            hwp.Run("MoveDocBegin")
            hwp.Run("SelectCtrlFront")
            try:
                hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
                hwp.HParameterSet.HShapeObject.TreatAsChar = 0#글자취급x
                hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
            except:
                print("서버")
                #아마도 범위 초과 설정 해서 그런듯
            hwp.MovePos(3)
            hwp.SaveAs(filename)

        elif Total_list_Combo_Export.get()=="엑셀(.xlsx)":
            filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not filename:
                return
            write_wb = openpyxl.Workbook()
            write_ws = write_wb.active
            write_ws['A1'].font = Font(size=20,bold=True)
            write_ws['A1'] = '발주서 종합'
            write_ws['A1'].border = Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))

            write_ws.append([""])


            for col in range(1,(DateSetCount*3)+1):
                if col % 3 == 1:
                    write_ws.column_dimensions[get_column_letter(col)].width = 10
                else:
                    write_ws.column_dimensions[get_column_letter(col)].width = 5

            refineDateSetName = []
            for x in DateSetName:
                refineDateSetName.append(x)
                refineDateSetName.append("")
                refineDateSetName.append("")
            
            write_ws.append(refineDateSetName)


            #엑셀 전용으로 값을 조정
            excelTemplistText = []
            for y in range(0,len(listText)):#파일 개수
                groupCount = 0#카운트용 변수

                separateNum = 0
                SubexcelTemplistText = []
                emptyCellCount = 0
                prevSeparategroupCount = 0
                for x in listText[y]:#일별로 파일에 있는 항목들
                    separategroupCount = groupCount/6
                    if groupCount % 6 == 1: #재료명
                        SubexcelTemplistText.append(x.lstrip())
                    if groupCount % 6 == 3: #단위
                        SubexcelTemplistText.append(x.lstrip())
                    if groupCount % 6 == 4: #개수
                        SubexcelTemplistText.append(x.lstrip())
                    
                    if separategroupCount-prevSeparategroupCount == LocalTimeList[y][emptyCellCount]:
                        if LocalTimeList[y][emptyCellCount] < MaxLocalTime[emptyCellCount]:
                            for i in range(0,MaxLocalTime[emptyCellCount]-LocalTimeList[y][emptyCellCount]):
                                SubexcelTemplistText.append("")
                                SubexcelTemplistText.append("")
                                SubexcelTemplistText.append("")
                        prevSeparategroupCount += separategroupCount
                        emptyCellCount +=1
                    groupCount += 1
                excelTemplistText.append(SubexcelTemplistText)

            tempCol = 1
            for j in range(0,len(excelTemplistText)):
                localCount = 0
                rowJ = 4
                colJ = tempCol

                for i in excelTemplistText[j]:
                    write_ws.cell(column=colJ,row=rowJ,value=i)
                    if localCount % 3 == 2:
                        colJ = tempCol
                        rowJ +=1
                        if rowJ != 4:
                            colJ-=1
                    localCount +=1
                    colJ+=1
                
                tempCol += 3

            write_ws.merge_cells("A1:"+get_column_letter(DateSetCount*3)+"2")#TODO:5를 SetCount로 변경
            mergeCount = 1
            for x in range(1,(DateSetCount*3)+1):
                write_ws.merge_cells(get_column_letter(mergeCount)+"3:"+get_column_letter(mergeCount+2)+"3")#TODO:5를 SetCount로 변경
                mergeCount+=3

            write_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            for x in range(1,(DateSetCount*3)+1):
                write_ws[get_column_letter(x)+"3"].font = Font(size=15,bold=True)
                write_ws[get_column_letter(x)+"3"].alignment = Alignment(horizontal='center', vertical='center')

            startChr = 65
            for x in range(0,DateSetCount):
                set_border(write_ws, chr(startChr)+str(3)+":"+chr(startChr+2)+str(3),"medium")
                startChr += 3

            startChr = 65
            for y in range(0, DateSetCount):
                startNum = 4
                for x in range(0,len(MaxLocalTime)):
                    set_border(write_ws, chr(startChr)+str(startNum)+":"+chr(startChr+2)+str(startNum+MaxLocalTime[x]),"medium")
                    startNum += MaxLocalTime[x]
                startChr += 3
            write_wb.save(filename)

            os.startfile(filename)

        elif Total_list_Combo_Export.get()=="텍스트(.txt)":
            filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".txt", filetypes=[("TXT files", "*.txt")])
            if not filename:
                return
            data = open(filename, 'w', encoding="UTF8")

            print("발주서 종합\n",file = data)
            
            txtTemplistText = []
            for y in range(0,len(listText)):#파일 개수
                
                groupCount = 0#카운트용 변수
                SubTxtTemplistText = []

                for x in listText[y]:#일별로 파일에 있는 항목들

                    if groupCount % 6 == 1: #재료명
                        SubTxtTemplistText.append(x.lstrip())
                    if groupCount % 6 == 3: #단위
                        SubTxtTemplistText.append(x.lstrip())
                    if groupCount % 6 == 4: #개수
                        SubTxtTemplistText.append(x.lstrip())

                    groupCount += 1

                txtTemplistText.append(SubTxtTemplistText)

            print("====================",file=data)
            for x in range(0,DateSetCount):
                print(DateSetName[x],file=data)
                print("====================",file=data)
                for startNum in range(0,len(txtTemplistText[x]),3):
                    print(txtTemplistText[x][startNum],txtTemplistText[x][startNum+1],txtTemplistText[x][startNum+2], file=data)
                print("====================\n",file=data)

            os.startfile(filename)

    elif Total_list_Combo_Import.get() == "텍스트(.txt)":
        if DateSetCount > Total_listbox.size():
            for x in range(0,DateSetCount-Total_listbox.size()):
                Total_listbox.insert(END,"Empty")

        TempTotalList = []
        for x in range(0,Total_listbox.size()):
            if Total_listbox.get(x) == "Empty":
                TempTotalList.append("")
            else:
                TempTotalList.append(Program_Result_Save_Dir_Var.get()+"\\"+Total_listbox.get(x))
        pLog.append_log("TempTotalList: ", TempTotalList)
        for x in range(0, len(TempTotalList)):
            if not TempTotalList[x]:
                pass
            else:
                if TempTotalList[x][-4:] != ".txt" :
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
        print(TempTotalList)
        print(len(TempTotalList))
        
        if len(TempTotalList)>0:
            listText = []
            for x in range(0,len(TempTotalList)):
                subListText = []
                if not TempTotalList[x]:
                    subListText.append("")
                    listText.append(subListText)
                    continue
                openfile = open(TempTotalList[x],'r',encoding="utf-8")
                readtext = openfile.read()
                readsplit_text = readtext.split("\n")
                openfile.close()

                #removeT = str(readsplit_text).replace("[]",",,,,,")
                removeT = str(readsplit_text).replace("[","")
                removeT = removeT.replace("]","")
                removeT = removeT.replace("'","")
                removeT = removeT.replace("\"","")
                removeT = removeT[removeT.find("번호, 재료명, 규격, 단위, 수량, 비고")+26:]
                subListText = removeT.split(",")
                subListText1 = []

                for x in range(0,len(subListText)):
                    subListText1.append(subListText[x].lstrip(" "))
                listText.append(subListText1)

            print(listText)

            LocalTimeList = [] #각 파일별 [오전,오후,저녁] 재료 항목의 개수를 모을 리스트
            #시간값이 포함된 결과물인지 확인용(사용자가 등록한 항목 개수 이상이면 시간값 포함)
            for x in range(0,len(listText)):#불러온 파일들
                incount = 0 #단순 카운트용 변수
                
                timeList = []#오전,오후,저녁별 재료 개수 담아둘 변수
                lenListNum = 0
                for y in listText[x]: #[번호,이름,규격,단위,개수,비고]
                    if incount % 6 == 0:#번호만 추출
                        onlyNumber = str(y.replace(" ",""))
                        if onlyNumber.isdigit() == True:
                            lenListNum+=1
                        else:
                            lenListNum+=0
                    incount+=1
                print(lenListNum)#총 재료의 길이 구하는거

                incount = 0
                allList = [] #전부 박아둘 리스트
                for y in listText[x]: #[번호,이름,규격,단위,개수,비고]
                    if incount % 6 == 0:#번호만 접근
                        onlyNumber = str(y.replace(" ",""))
                        
                        if not onlyNumber:
                            allList.append(0)
                            pastNum = 1
                        else:
                            allList.append(int(onlyNumber))
                    incount +=1
                print("AllList: "+str(allList))

                leftList = [] #리스트의 마지막을 제외한 재료 번호 중 끝값들
                pastNum = 1
                justCount = 0
                for x in range(0,len(allList)):

                    if allList[x] == pastNum:
                        if justCount == 0:
                            pass
                        else:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(allList[x])
                    #기준값이 이전값 보다 크면 정상
                    if allList[x] > pastNum:
                        pastNum += 1
                    #기준값이 이전값 보다 작으며 비정상 = 시간대가 달라진다.
                    if allList[x] < pastNum:
                        if justCount == 0:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(0)
                            continue
                        
                        leftList.append(pastNum)
                        pastNum = allList[x]
                        if allList[x] == 0:
                            if justCount == len(allList)-1:
                                continue
                            leftList.append(0)
                        
                    justCount += 1
                timeList.append(leftList)

                rightList = allList[-1]
                leftList.append(rightList)

                LocalTimeList.append(leftList)
            print(LocalTimeList)

            #0을 1로 변경
            for x in LocalTimeList:
                for y in range(0,len(x)):
                    if x[y] == 0:
                        x[y] = 1
            
            #각 파일별 시간대 개수
            maxSizeOfLocalTimeList = 0
            for x in LocalTimeList:
                if len(x) > maxSizeOfLocalTimeList:
                    maxSizeOfLocalTimeList = len(x)
            print(maxSizeOfLocalTimeList)
            
            #최대 시간대 개수 로 통일
            for x in range(0, len(LocalTimeList)):
                if len(LocalTimeList[x]) <maxSizeOfLocalTimeList:
                    for y in range(0,maxSizeOfLocalTimeList-len(LocalTimeList[x])):
                        LocalTimeList[x].append(1)

            print(LocalTimeList)
            #각 파일별 시간대별 최대값 [오전 최댓값, 오후 최댓값, 저녁 최댓값...]
            MaxLocalTime = []        
            for x in range(0,maxSizeOfLocalTimeList):
                MaxLocalTime.append(max(t[x] for t in LocalTimeList))

        if Total_list_Combo_Export.get() == "한글(.hwp)":
            filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".hwp", filetypes=[("Hwp files", "*.hwp")])
            hwp = win32.gencache.EnsureDispatch('HWPFrame.HwpObject')  # 한/글 열기
            hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글')  # 해당 윈도우의 핸들값 찾기

            #win32gui.ShowWindow(hwnd,1)#창 백그라운드에서 실행
            hwp.RegisterModule('FilePathCheckDLL', 'FilePathCheckerModule')
            hwp.XHwpWindows.Item(0).Visible = True  # 숨김해제

            #여백 세팅
            hwp.HAction.GetDefault("ModifySection", hwp.HParameterSet.HSecDef.HSet)
            hwp.HParameterSet.HSecDef.PageDef.LeftMargin = hwp.MiliToHwpUnit(20.0)
            hwp.HParameterSet.HSecDef.PageDef.TopMargin = hwp.MiliToHwpUnit(15.0)
            hwp.HParameterSet.HSecDef.PageDef.RightMargin = hwp.MiliToHwpUnit(20.0)
            hwp.HParameterSet.HSecDef.PageDef.BottomMargin = hwp.MiliToHwpUnit(15.0)
            hwp.HParameterSet.HSecDef.PageDef.HeaderLen = hwp.MiliToHwpUnit(10.0)
            hwp.HParameterSet.HSecDef.PageDef.FooterLen = hwp.MiliToHwpUnit(10.0)
            hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyClass", 24)
            hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyTo", 2)
            hwp.HAction.Execute("ModifySection", hwp.HParameterSet.HSecDef.HSet)
            
            hwp.HAction.Run("ParagraphShapeAlignCenter")
            hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
            hwp.HParameterSet.HInsertText.Text = "발주서 종합"
            hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
            hwp.HAction.Run("SelectAll")
            hwp_fontSetting(hwp,"HY헤드라인M",22,1)

            hwp.HAction.Run("MoveRight")
            hwp.HAction.Run("BreakPara")
            #열 - 가로, 행 - 세로
            hwp.HAction.GetDefault("TableCreate", hwp.HParameterSet.HTableCreation.HSet)  # 표 생성 시작
            hwp.HParameterSet.HTableCreation.Rows = maxSizeOfLocalTimeList+1  # 행 갯수
            hwp.HParameterSet.HTableCreation.Cols = DateSetCount  # 열 갯수
            hwp.HParameterSet.HTableCreation.WidthType = 2  # 너비 지정(0:단에맞춤, 1:문단에맞춤, 2:임의값)
            hwp.HParameterSet.HTableCreation.HeightType = 0  # 높이 지정(0:자동, 1:임의값)
            #hwp.HParameterSet.HTableCreation.WidthValue = hwp.MiliToHwpUnit(148.0)  # 표 너비
            #hwp.HParameterSet.HTableCreation.HeightValue = hwp.MiliToHwpUnit(150)  # 표 높이
            hwp.HParameterSet.HTableCreation.CreateItemArray("ColWidth", DateSetCount)  # 열 5개 생성
            for x in range(0,DateSetCount):
                hwp.HParameterSet.HTableCreation.ColWidth.SetItem(x, hwp.MiliToHwpUnit(30.0))  # 1열(가로 크기)

            hwp.HParameterSet.HTableCreation.CreateItemArray("RowHeight", maxSizeOfLocalTimeList+1)  # 행 5개 생성
            for x in range(0,maxSizeOfLocalTimeList+1):
                if x == 0:
                    hwp.HParameterSet.HTableCreation.RowHeight.SetItem(x, hwp.MiliToHwpUnit(15.0))  # 1행
                else:
                    hwp.HParameterSet.HTableCreation.RowHeight.SetItem(x, hwp.MiliToHwpUnit(50.0))  # 2행

            hwp.HParameterSet.HTableCreation.TableProperties.TreatAsChar = 1  # 글자처럼 취급
            hwp.HParameterSet.HTableCreation.TableProperties.Width = hwp.MiliToHwpUnit(148)  # 표 너비
            hwp.HAction.Execute("TableCreate", hwp.HParameterSet.HTableCreation.HSet)  # 위 코드 실행

            hwp.HAction.Run("TableCellBlock")
            hwp.HAction.Run("TableCellBlockExtend")
            for x in range(0,DateSetCount-1): 
                hwp.HAction.Run("TableRightCell")
            hwp.HAction.Run("ParagraphShapeAlignCenter")

            hwp_fontSetting(hwp,"HY헤드라인M",16,1)

            hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
            hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
            hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
            hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
            hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
            hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

            hwp.Run("MoveDocBegin")
            hwp.Run("SelectCtrlFront")

            hwp.HAction.Run("ShapeObjTableSelCell")
            hwp.HAction.Run("TableCellBlockExtend")


            for y in range(0, DateSetCount):
                for x in range(0,maxSizeOfLocalTimeList):
                    hwp.HAction.Run("TableLowerCell")

                hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
                hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                hwp.HAction.Run("TableRightCell")

            hwp.Run("MoveDocBegin")
            hwp.Run("SelectCtrlFront")
            hwp.HAction.Run("ShapeObjTableSelCell")
            hwp.HAction.Run("Cancel")

            
            for x in range(0,len(DateSetName)):
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = DateSetName[x]
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                hwp.HAction.Run("TableRightCell")
            
            for y in range(0,len(listText)):#파일 개수
                groupCount = 0#카운트용 변수
                subgroupCount = 0#재료 개수
                separateNum = 0

                for x in listText[y]:#일별로 파일에 있는 항목들

                    if groupCount % 6 == 1: #재료명
                        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HAction.Run("CharShapeBold")
                        hwp.HParameterSet.HInsertText.Text = x + " "
                        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if groupCount % 6 == 3: #단위
                        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HAction.Run("CharShapeBold")
                        hwp.HParameterSet.HInsertText.Text = x + " "
                        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if groupCount % 6 == 4: #개수
                        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HParameterSet.HInsertText.Text = x + " "
                        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HAction.Run("BreakPara")
                    if groupCount % 6 == 5: #비고
                        subgroupCount+=1
                                        
                    if subgroupCount >= LocalTimeList[y][separateNum]:
                        hwp.HAction.Run("TableLowerCell")
                        separateNum+=1
                        subgroupCount= 0

                    groupCount += 1
                
                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")

                for z in range(0,y+1):
                    hwp.HAction.Run("TableRightCell")
                
                hwp.HAction.Run("TableLowerCell")
                hwp.HAction.Run("Cancel")

            hwp.Run("MoveDocBegin")
            hwp.Run("SelectCtrlFront")
            try:
                hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
                hwp.HParameterSet.HShapeObject.TreatAsChar = 0#글자취급x
                hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
            except:
                print("서버")
                #아마도 범위 초과 설정 해서 그런듯
            hwp.MovePos(3)
            hwp.SaveAs(filename)

        elif Total_list_Combo_Export.get()=="엑셀(.xlsx)":
            filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if not filename:
                return
            write_wb = openpyxl.Workbook()
            write_ws = write_wb.active
            write_ws['A1'].font = Font(size=20,bold=True)
            write_ws['A1'] = '발주서 종합'
            write_ws['A1'].border = Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))

            write_ws.append([""])


            for col in range(1,(DateSetCount*3)+1):
                if col % 3 == 1:
                    write_ws.column_dimensions[get_column_letter(col)].width = 10
                else:
                    write_ws.column_dimensions[get_column_letter(col)].width = 5

            refineDateSetName = []
            for x in DateSetName:
                refineDateSetName.append(x)
                refineDateSetName.append("")
                refineDateSetName.append("")
            
            write_ws.append(refineDateSetName)


            #엑셀 전용으로 값을 조정
            excelTemplistText = []
            for y in range(0,len(listText)):#파일 개수
                groupCount = 0#카운트용 변수

                separateNum = 0
                SubexcelTemplistText = []
                emptyCellCount = 0
                prevSeparategroupCount = 0
                for x in listText[y]:#일별로 파일에 있는 항목들
                    separategroupCount = groupCount/6
                    if groupCount % 6 == 1: #재료명
                        SubexcelTemplistText.append(x.lstrip())
                    if groupCount % 6 == 3: #단위
                        SubexcelTemplistText.append(x.lstrip())
                    if groupCount % 6 == 4: #개수
                        SubexcelTemplistText.append(x.lstrip())
                    
                    if separategroupCount-prevSeparategroupCount == LocalTimeList[y][emptyCellCount]:
                        if LocalTimeList[y][emptyCellCount] < MaxLocalTime[emptyCellCount]:
                            for i in range(0,MaxLocalTime[emptyCellCount]-LocalTimeList[y][emptyCellCount]):
                                SubexcelTemplistText.append("")
                                SubexcelTemplistText.append("")
                                SubexcelTemplistText.append("")
                        prevSeparategroupCount += separategroupCount
                        emptyCellCount +=1
                    groupCount += 1
                excelTemplistText.append(SubexcelTemplistText)

            tempCol = 1
            for j in range(0,len(excelTemplistText)):
                localCount = 0
                rowJ = 4
                colJ = tempCol

                for i in excelTemplistText[j]:
                    write_ws.cell(column=colJ,row=rowJ,value=i)
                    if localCount % 3 == 2:
                        colJ = tempCol
                        rowJ +=1
                        if rowJ != 4:
                            colJ-=1
                    localCount +=1
                    colJ+=1
                
                tempCol += 3

            write_ws.merge_cells("A1:"+get_column_letter(DateSetCount*3)+"2")#TODO:5를 SetCount로 변경
            mergeCount = 1
            for x in range(1,(DateSetCount*3)+1):
                write_ws.merge_cells(get_column_letter(mergeCount)+"3:"+get_column_letter(mergeCount+2)+"3")#TODO:5를 SetCount로 변경
                mergeCount+=3

            write_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            for x in range(1,(DateSetCount*3)+1):
                write_ws[get_column_letter(x)+"3"].font = Font(size=15,bold=True)
                write_ws[get_column_letter(x)+"3"].alignment = Alignment(horizontal='center', vertical='center')

            startChr = 65
            for x in range(0,DateSetCount):
                set_border(write_ws, chr(startChr)+str(3)+":"+chr(startChr+2)+str(3),"medium")
                startChr += 3

            startChr = 65
            for y in range(0, DateSetCount):
                startNum = 4
                for x in range(0,len(MaxLocalTime)):
                    set_border(write_ws, chr(startChr)+str(startNum)+":"+chr(startChr+2)+str(startNum+MaxLocalTime[x]),"medium")
                    startNum += MaxLocalTime[x]
                startChr += 3
            write_wb.save(filename)

            os.startfile(filename)

        elif Total_list_Combo_Export.get()=="텍스트(.txt)":
            filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=TotalDate_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".txt", filetypes=[("TXT files", "*.txt")])
            if not filename:
                return
            data = open(filename, 'w', encoding="UTF8")

            print("발주서 종합\n",file = data)
            
            txtTemplistText = []
            for y in range(0,len(listText)):#파일 개수
                
                groupCount = 0#카운트용 변수
                SubTxtTemplistText = []

                for x in listText[y]:#일별로 파일에 있는 항목들

                    if groupCount % 6 == 1: #재료명
                        SubTxtTemplistText.append(x.lstrip())
                    if groupCount % 6 == 3: #단위
                        SubTxtTemplistText.append(x.lstrip())
                    if groupCount % 6 == 4: #개수
                        SubTxtTemplistText.append(x.lstrip())

                    groupCount += 1

                txtTemplistText.append(SubTxtTemplistText)

            print("====================",file=data)
            for x in range(0,DateSetCount):
                print(DateSetName[x],file=data)
                print("====================",file=data)
                for startNum in range(0,len(txtTemplistText[x]),3):
                    print(txtTemplistText[x][startNum],txtTemplistText[x][startNum+1],txtTemplistText[x][startNum+2], file=data)
                print("====================\n",file=data)

            os.startfile(filename)
   
#============================================================
#[요일별 종합 - 버튼] 빈 값 추가
#============================================================
def Total_Add_EmptyValue():
    global Total_treeNumCount
    global Total_listbox
    global DateSetCount

    if Total_treeNumCount <= DateSetCount:
        Total_listbox.insert(Total_treeNumCount,"Empty")
        Total_treeNumCount+=1

#============================================================
#[요일별 종합 - 버튼] 항목 추가
#============================================================
def Total_day_Add():
    if Total_list_Combo_Import.get() == "한글(.hwp)":
        original_path = filedialog.askopenfilename(initialdir=Program_Result_Save_Dir_Var.get(), multiple=True, title="Select file",
                                          filetypes=(("HWP files", "*.hwp"), ("all files", "*.*")))
    elif Total_list_Combo_Import.get()=="엑셀(.xlsx)":
        original_path = filedialog.askopenfilename(initialdir=Program_Result_Save_Dir_Var.get(), multiple=True, title="Select file",
                                          filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")))
    elif Total_list_Combo_Import.get()=="텍스트(.txt)":
        original_path = filedialog.askopenfilename(initialdir=Program_Result_Save_Dir_Var.get(), multiple=True, title="Select file",
                                          filetypes=(("TXT files", "*.txt"), ("all files", "*.*")))
    
    if not original_path:
        return           

    global Total_treeNumCount
    global Total_listbox
    global DateSetCount
    if Total_treeNumCount <= DateSetCount:

        for x in original_path: #선택한 경로에서 요리명, 자격증 종류 가져옴
            strx=str(x)
            fileName = os.path.basename(strx)
            fileName1 = fileName[-4:]

            if Total_list_Combo_Import.get() == "한글(.hwp)":
                if fileName1 != ".hwp":
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
            elif Total_list_Combo_Import.get()=="엑셀(.xlsx)":
                if fileName1 != "xlsx":
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
            elif Total_list_Combo_Import.get()=="텍스트(.txt)":
                if fileName1 != ".txt":
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
            C_fileName = re.sub(r"[^a-zA-Z|가-힣|/※~()]","",fileName1) 

            fileDir = os.path.dirname(strx)        
            licenseType = fileDir[fileDir.rfind("/")+2:fileDir.rfind("/")+4]

            if Total_treeNumCount <= DateSetCount:
                Total_listbox.insert(Total_treeNumCount,fileName)
                Total_treeNumCount+=1

#============================================================
#[요일별 종합 - 프로그램] 항목 추가 - DND
#============================================================
def Total_day_Add_DnD(event):

    global Total_treeNumCount
    global Total_listbox
    global DateSetCount

    event.data = event.data.replace("{","")
    sub_original_path = event.data.split("}")

    original_path = []
    for x in sub_original_path:
        if x == "":
            break
        else: 
            if x[0] == " " or x[0] == "":
                original_path.append(x[1:])
            else:
                original_path.append(x)

    if Total_treeNumCount <= DateSetCount:
        
        for x in original_path: #선택한 경로에서 요리명, 자격증 종류 가져옴
            strx=str(x)
            fileName = os.path.basename(strx)
            fileName1 = fileName[-4:]

            if Total_list_Combo_Import.get() == "한글(.hwp)":
                if fileName1 != ".hwp":
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
            elif Total_list_Combo_Import.get()=="엑셀(.xlsx)":
                if fileName1 != "xlsx":
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
            elif Total_list_Combo_Import.get()=="텍스트(.txt)":
                if fileName1 != ".txt":
                    tkinter.messagebox.showwarning("경고", "불러들이는 파일의 확장자가 올바르지 않습니다.")
                    return
            C_fileName = re.sub(r"[^a-zA-Z|가-힣|/※~()]","",fileName1) 

            fileDir = os.path.dirname(strx)        
            licenseType = fileDir[fileDir.rfind("/")+2:fileDir.rfind("/")+4]

            if Total_treeNumCount <= DateSetCount:
                Total_listbox.insert(Total_treeNumCount,fileName)
                Total_treeNumCount+=1

#============================================================
#[요일별 종합 - 버튼] 항목 삭제
#============================================================
def Total_day_Del():
    global Total_treeNumCount
    selected_item= Total_listbox.curselection()
    for item in selected_item[::-1]:
        Total_listbox.delete(item)
    Total_treeNumCount -=1

#============================================================
#[요일별 종합 - 버튼] 전체 삭제
#============================================================
def Total_day_Del_All():
    global Total_treeNumCount
    Total_listbox.delete(0,END)
    Total_treeNumCount = 1

#============================================================
#[요일별 종합 - 버튼] 위로 올리기
#============================================================
def Total_List_MoveUp():
    try:
        idxs = Total_listbox.curselection()
        if not idxs:
            return
        for pos in idxs:
            if pos==0:
                continue
            text=Total_listbox.get(pos)
            Total_listbox.delete(pos)
            Total_listbox.insert(pos-1, text)

        Total_listbox.selection_set(pos-1)
    except:
        pass

#============================================================
#[요일별 종합 - 버튼] 아래로 내리기
#============================================================
def Total_List_MoveDown():
    global DateSetCount
    try:
        idxs = Total_listbox.curselection()
        if not idxs:
            return
        for pos in idxs:

            text=Total_listbox.get(pos)
            Total_listbox.delete(pos)
            Total_listbox.insert(pos+1, text)
            Total_listbox.selection_set(pos + 1)
    except:
        pass
#============================================================
#[메인 - 메뉴바 - 요일별 종합] 요일별 종합 - 종료시
#============================================================
def Close_Total_days_Window():
    global date_root
    global isOnTotalWindow
    global DATE_TOTAL_IMPORT_COMBO_INIT
    global DATE_TOTAL_EXPORT_COMBO_INIT
    
    DATE_TOTAL_IMPORT_COMBO_INIT = Total_list_Combo_Import.current()
    DATE_TOTAL_EXPORT_COMBO_INIT = Total_list_Combo_Export.current()
    date_root.destroy()
    isOnTotalWindow = 0
#============================================================
#[메인 - 메뉴바 - 요일별 종합] 요일별 종합
#============================================================
isOnTotalWindow = 0
def BTN_Total_days():
    global isOnTotalWindow
    global Total_treeNumCount
    global Total_listbox
    global DateSetCount
    global DateSetName
    global Total_list_Combo_Import
    global Total_list_Combo_Export
    global date_root
    
    if isOnTotalWindow == 0:
        isOnTotalWindow = 1
        date_root = tkinter.Tk()
        date_root.title("요일별 종합")
        date_root.geometry("600x270")
        date_root.lift()
        date_root.resizable(False, False) #창 사이즈 변경 불가능
        date_root.attributes('-topmost', True)
        date_root.attributes('-topmost', False)
        
        Total_treeNumCount = 1
        if os.path.isfile('./res/sys/sys_date.txt') == True: #시스템- 날짜 파일이 존재하면
            openfile = open('./res/sys/sys_date.txt','r',encoding="utf-8")
            readtext = openfile.read()
            readsplit_text = readtext.split("\n")
            openfile.close()
            DateSetName = readsplit_text
            DateSetCount = len(readsplit_text)
            if DateSetCount == 0 :
                tkinter.messagebox.showwarning("요일 등록", "등록된 요일이 없습니다.\n설정해 주세요.")
        else:
            tkinter.messagebox.showwarning("요일 등록", "요일 설정 파일이 없습니다.\n등록해주세요.")

        Total_list_Date_Setting_Button = tkinter.Button(date_root, text='요일 설정', command=Total_DaySetting, overrelief="solid", width=8)
        Total_list_Date_Setting_Button.place(x=60,y=10)
    
        Import_Total_list_Label = Label(date_root,text="【불러올 값】",font=("Arial",11))
        Import_Total_list_Label.place(x=130,y=10)
        
        #불러들일 파일 형식
        Total_list_Combo_Import = ttk.Combobox(date_root,width=9)
        Total_list_Combo_Import['values']=("한글(.hwp)", "엑셀(.xlsx)", "텍스트(.txt)")
        Total_list_Combo_Import.current(DATE_TOTAL_IMPORT_COMBO_INIT)
        Total_list_Combo_Import.place(x=230,y=10,height=26)

        Import_Total_list_Label = Label(date_root,text="【내보낼 값】",font=("Arial",11))
        Import_Total_list_Label.place(x=320,y=10)
        #내보낼 형식
        Total_list_Combo_Export = ttk.Combobox(date_root,width=9)
        Total_list_Combo_Export['values']=("한글(.hwp)", "엑셀(.xlsx)", "텍스트(.txt)")
        Total_list_Combo_Export.current(DATE_TOTAL_EXPORT_COMBO_INIT)
        Total_list_Combo_Export.place(x=420,y=10,height=26)

        Date_Name_info_Frame = Frame(date_root,width=50,height=225, relief="solid", bd=1, bg="lightgray")
        Date_Name_info_Frame.pack()
        Date_Name_info_Frame.place(x=5,y=40)
        
        for x in range(0,len(readsplit_text)):
            date_Label = Label(date_root,text=readsplit_text[x]+"\n",background="lightgray")
            date_Label.place(x=10,y=42+(x*16))
        
        Total_listbox = Listbox(date_root,selectmode=SINGLE,background="#ffe0d6")
        Total_listbox.place(x=60,y=40,width=450,height=225)

        # Total_listbox.drop_target_register(DND_FILES)
        # Total_listbox.dnd_bind("<<Drop>>",Total_day_Add_DnD)

        Total_list_Button_EmptyValue = tkinter.Button(date_root, text='빈 값 추가', command=Total_Add_EmptyValue, overrelief="solid", width=8)
        Total_list_Button_EmptyValue.place(x=520,y=10)

        Total_list_Button_Add = tkinter.Button(date_root, text='항목 추가', command=Total_day_Add, overrelief="solid", width=8)
        Total_list_Button_Add.place(x=520,y=40)

        Total_list_Button_Delete = tkinter.Button(date_root, text='항목 삭제', command=Total_day_Del, overrelief="solid", width=8)
        Total_list_Button_Delete.place(x=520,y=70)

        Total_list_Button_All_Delete = tkinter.Button(date_root, text='전체 삭제', command=Total_day_Del_All, overrelief="solid", width=8)
        Total_list_Button_All_Delete.place(x=520,y=100)

        

        Total_list_Up_Button = tkinter.Button(date_root, text='▲', command=Total_List_MoveUp, overrelief="solid", width=8)
        Total_list_Up_Button.place(x=520,y=160)

        Total_list_Down_Button = tkinter.Button(date_root, text='▼', command=Total_List_MoveDown, overrelief="solid", width=8)
        Total_list_Down_Button.place(x=520,y=190)

        Total_list_TotalResult_Button = tkinter.Button(date_root, text='결과 확인', command=Total_Result, overrelief="solid", width=8)
        Total_list_TotalResult_Button.place(x=520,y=235)
        date_root.protocol('WM_DELETE_WINDOW', Close_Total_days_Window)
        date_root.mainloop()
    else:
        tkinter.messagebox.showwarning("오류", "이미 창이 실행중입니다.")
#============================================================
#[메인 - 메뉴바 - 파일] 작업 목록 불러오기
#============================================================
def BTN_LoadList_txt():
    global treeNumCount
    if treeNumCount >= 2:
        if tkinter.messagebox.askokcancel("목록 불러오기", "지금 하던 작업이 삭제됩니다.\n정말 불러 오시겠습니까?"):
            
            treeNumCount = 1
            for y in listbox.tree.get_children():
                listbox.tree.delete(y)

            filePath = filedialog.askopenfilename(initialdir=Program_Save_Dir_Var.get(),title="파일 열기",defaultextension=".txt",filetypes=[('txt file','*.txt')])
            if not filePath:
                return
            loadlistfile = open(filePath,"r",encoding="utf-8")
            
            readtext = loadlistfile.read()
            readtext1 = re.sub(r"[^a-zA-Z0-9|가-힣|,|\"|\n]","",readtext) 
            readsplit_text = readtext1.split("\n")
            
            #print(readsplit_text)
            loadlistfile.close()
            
            tempList = []
            treeNewCount = 1
            for y in listbox.tree.get_children():
                tempList =  listbox.tree.item(y)["values"]
                listbox.tree.delete(y)
                listbox.tree.insert('',index= treeNewCount,iid=treeNewCount, values=(treeNewCount,tempList[1],tempList[2],tempList[3],tempList[4]))
                treeNewCount += 1

            treeNumCount = treeNewCount
            tempList =[]
            for x in readsplit_text:
                if x != "":
                    tempList=x.split(",")
                    
                    listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,tempList[1],tempList[2],tempList[3],tempList[4]))
                    treeNumCount+=1
            #listboxCount = treeNumCount
    else:
        
        treeNumCount = 1
        for y in listbox.tree.get_children():
            listbox.tree.delete(y)

        filePath = filedialog.askopenfilename(initialdir= Program_Save_Dir_Var.get(),title="파일 열기",defaultextension=".txt",filetypes=[('txt file','*.txt')])
        if not filePath:
            return
        loadlistfile = open(filePath,"r",encoding="utf-8")
        
        readtext = loadlistfile.read()
        readtext1 = re.sub(r"[^a-zA-Z0-9|가-힣|,|\"|\n]","",readtext) 
        readsplit_text = readtext1.split("\n")
        
        #print(readsplit_text)
        loadlistfile.close()
        
        tempList = []
        treeNewCount = 1
        for y in listbox.tree.get_children():
            tempList =  listbox.tree.item(y)["values"]
            listbox.tree.delete(y)
            listbox.tree.insert('',index= treeNewCount,iid=treeNewCount, values=(treeNewCount,tempList[1],tempList[2],tempList[3],tempList[4]))
            treeNewCount += 1

        treeNumCount = treeNewCount
        tempList =[]
        for x in readsplit_text:
            if x != "":
                tempList=x.split(",")
                
                listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,tempList[1],tempList[2],tempList[3],tempList[4]))
                treeNumCount+=1
        #listboxCount = treeNumCount
#============================================================
#[메인 - 메뉴바 - 파일] 작업 목록 불러오기 [단축키] Ctrl + o
#============================================================
def BTN_LoadList_txt_Key(event):
    global treeNumCount
    if treeNumCount >= 2:
        if tkinter.messagebox.askokcancel("목록 불러오기", "지금 하던 작업이 삭제됩니다.\n정말 불러 오시겠습니까?"):
            
            treeNumCount = 1
            for y in listbox.tree.get_children():
                listbox.tree.delete(y)

            filePath = filedialog.askopenfilename(initialdir=Program_Save_Dir_Var.get(),title="파일 열기",defaultextension=".txt",filetypes=[('txt file','*.txt')])
            if not filePath:
                return
            loadlistfile = open(filePath,"r",encoding="utf-8")
            readtext = loadlistfile.read()
            readtext1 = re.sub(r"[^a-zA-Z0-9|가-힣|,|\"|\n]","",readtext) 
            readsplit_text = readtext1.split("\n")
            #print(readsplit_text)
            loadlistfile.close()

            tempList = []
            treeNewCount = 1
            for y in listbox.tree.get_children():
                tempList =  listbox.tree.item(y)["values"]
                listbox.tree.delete(y)
                listbox.tree.insert('',index= treeNewCount,iid=treeNewCount, values=(treeNewCount,tempList[1],tempList[2],tempList[3],tempList[4]))
                treeNewCount += 1

            treeNumCount = treeNewCount
            tempList =[]
            for x in readsplit_text:
                if x != "":
                    tempList=x.split(",")
                    
                    listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,tempList[1],tempList[2],tempList[3],tempList[4]))
                    treeNumCount+=1
            #listboxCount = treeNumCount
    else:
        treeNumCount = 1
        for y in listbox.tree.get_children():
            listbox.tree.delete(y)

        filePath = filedialog.askopenfilename(initialdir=Program_Save_Dir_Var.get(),title="파일 열기",defaultextension=".txt",filetypes=[('txt file','*.txt')])
        if not filePath:
            return
        loadlistfile = open(filePath,"r",encoding="utf-8")
        readtext = loadlistfile.read()
        readtext1 = re.sub(r"[^a-zA-Z0-9|가-힣|,|\"|\n]","",readtext) 
        readsplit_text = readtext1.split("\n")
        #print(readsplit_text)
        loadlistfile.close()

        tempList = []
        treeNewCount = 1
        for y in listbox.tree.get_children():
            tempList =  listbox.tree.item(y)["values"]
            listbox.tree.delete(y)
            listbox.tree.insert('',index= treeNewCount,iid=treeNewCount, values=(treeNewCount,tempList[1],tempList[2],tempList[3],tempList[4]))
            treeNewCount += 1

        treeNumCount = treeNewCount
        tempList =[]
        for x in readsplit_text:
            if x != "":
                tempList=x.split(",")
                
                listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,tempList[1],tempList[2],tempList[3],tempList[4]))
                treeNumCount+=1
        #listboxCount = treeNumCount
#============================================================
#[메인 - 메뉴바 - 파일] 작업 목록 저장하기
#============================================================
def BTN_SaveList_txt():
    filePath = filedialog.asksaveasfilename(initialdir=Program_Save_Dir_Var.get(),title="파일 저장",defaultextension=".txt",filetypes=[('txt file','*.txt')])
    if not filePath:
        return
    savelistfile = open(filePath,"w",encoding="utf-8")
    for y in listbox.tree.get_children():
        print(listbox.tree.item(y)["values"],file = savelistfile)
    savelistfile.close()

    tkinter.messagebox.showinfo("알림", "완료 되었습니다.")

#============================================================
#[메인 - 메뉴바 - 파일] 작업 목록 저장하기 [단축키] Ctrl + S
#============================================================
def BTN_SaveList_txt_Key(event):
    filePath = filedialog.asksaveasfilename(initialdir=Program_Save_Dir_Var.get(),title="파일 저장",defaultextension=".txt",filetypes=[('txt file','*.txt')])
    if not filePath:
        return
    savelistfile = open(filePath,"w",encoding="utf-8")
    for y in listbox.tree.get_children():
        print(listbox.tree.item(y)["values"],file = savelistfile)
    savelistfile.close()

    tkinter.messagebox.showinfo("알림", "완료 되었습니다.")

#============================================================
#[사용자 등록 - 메뉴바 - 파일] 새 파일
#============================================================
def MenuBTN_NewFile():
    CompanyNameEntry_Var.set("")
    CompanyAddressEntry_Var.set("")
    CompanyCEONameEntry_Var.set("")
    CompanyCEOTelEntry_Var.set("")
    CompanyMailEntry_Var.set("")
    DeliveryAddressEntry_Var.set("")
    PurchasingManagerNameEntry_Var.set("")
    PurchasingManagerTelEntry_Var.set("")

#============================================================
#[사용자 등록 - 메뉴바 - 파일] 유저 파일 열기
#============================================================
def MenuBTN_OpenFile():
    filePath = filedialog.askopenfilename(initialdir="./res/user",title="파일 열기",defaultextension=".csv",filetypes=[('csv file','*.csv')])
    if not filePath:
        return
    #print(filePath)
    file = open(filePath,'r',encoding='utf-8')
    rdr = csv.reader(file)

    lstr = []
    for line in rdr:
        lstr.append("".join(line)) 
          
    CompanyNameEntry_Var.set(lstr[0])
    CompanyAddressEntry_Var.set(lstr[1])
    CompanyCEONameEntry_Var.set(lstr[2])
    CompanyCEOTelEntry_Var.set(lstr[3])
    CompanyMailEntry_Var.set(lstr[4])
    DeliveryAddressEntry_Var.set(lstr[5])
    PurchasingManagerNameEntry_Var.set(lstr[6])
    PurchasingManagerTelEntry_Var.set(lstr[7])
    file.close()

#============================================================
#[사용자 등록 - 메뉴바 - 파일] 유저 파일 저장
#============================================================
def MenuBTN_SaveFile():
    filePath = filedialog.asksaveasfilename(initialdir="./res/user",title="파일 저장",defaultextension=".csv",filetypes=[('csv file','*.csv')])
    if not filePath:
        return
    menufile = open(filePath,'w',encoding='utf-8',newline='\n')
    wr = csv.writer(menufile)
    wr.writerow(CompanyNameEntry_Var.get())
    wr.writerow(CompanyAddressEntry_Var.get())
    wr.writerow(CompanyCEONameEntry_Var.get())
    wr.writerow(CompanyCEOTelEntry_Var.get())
    wr.writerow(CompanyMailEntry_Var.get())
    wr.writerow(DeliveryAddressEntry_Var.get())
    wr.writerow(PurchasingManagerNameEntry_Var.get())
    wr.writerow(PurchasingManagerTelEntry_Var.get())
    menufile.close()

#============================================================
#[사용자 등록 - 메뉴바 - 설정] 사용할 유저 설정
#============================================================
def MenuBTN_UserSetting():
    pLog.append_log("사용자 등록창 :" ,"사용할 유저 설정 실행")
    global USERNAMEDIR_VAR
    filePath = filedialog.askopenfilename(initialdir="./res/user",title="파일 열기",defaultextension=".csv",filetypes=[('csv file','*.csv')])
    #print(filePath)
    if not filePath:
        return
    strx=str(filePath)
    fileName = os.path.basename(strx)
    fileName1 = fileName[:-4]
    C_fileName = re.sub(r"[^a-zA-Z|가-힣|/※~()]","",fileName1) 
    C_fileName = C_fileName + " 님"
    NowSettingFile.set(C_fileName)
    USERNAMEDIR_VAR = strx

def BTN_Open_UserFolder():
    pLog.append_log("유저 폴더 실행: ",os.getcwd()+"\\res\\user")
    os.startfile(os.getcwd()+"\\res\\user")
#============================================================
#[메인 - 메뉴바 - 설정] 제외시킬 재료 설정
#[메인 - 우측패널 - 수량 입력 - 버튼] 제외시킬 재료
#============================================================
def MenuBTN_OpenBan():
    pLog.append_log("벤 파일 실행: ",os.getcwd()+'\\res\\banList\\BanListFile.txt')
    os.startfile(os.getcwd()+'\\res\\banList\\BanListFile.txt')
    

#============================================================
#[메인- 메뉴바 - 설정]사용자 설정
#============================================================
def BTN_Regist():
    pLog.append_log("버튼 클릭 실행:","메뉴바 - 사용자 설정")
    Regist_window=Toplevel()
    Regist_window.title("사용자 등록")
    Regist_window.geometry("500x250+250+300")
    Regist_window.config(background="gainsboro")
    Regist_window.resizable(False, False) #창 사이즈 변경 불가능
    Regist_window.lift()    
    Regist_window.attributes('-topmost', True)
    Regist_window.attributes('-topmost', False)
    

    menubar = Menu(Regist_window)
    menu1 = Menu(menubar,tearoff=0)
    menu1.add_command(label="새 파일",command=MenuBTN_NewFile)
    menu1.add_command(label="유저 파일 열기",command=MenuBTN_OpenFile)
    menu1.add_command(label="유저 파일 저장",command=MenuBTN_SaveFile)
    menu1.add_separator()
    menu1.add_command(label="창 닫기", command=Regist_window.destroy)
    menubar.add_cascade(label="파일",menu=menu1)

    menu2 = Menu(menubar,tearoff=0)
    menu2.add_command(label="사용할 유저 설정",command= MenuBTN_UserSetting)
    menu2.add_command(label="유저 목록 폴더 열기",command= BTN_Open_UserFolder)
    
    menubar.add_cascade(label="설정",menu=menu2)

    CompanyNameText = Label(Regist_window,text="【업체명】",background="gainsboro",justify="left")
    CompanyNameText.pack()
    CompanyNameText.place(x= 0, y= 20)

    CompanyNameEntry = Entry(Regist_window, width=13, textvariable=CompanyNameEntry_Var)
    CompanyNameEntry.place(x=130,y=20)

    CompanyAddressText = Label(Regist_window,text="【업체 주소】",background="gainsboro",justify="left")
    CompanyAddressText.pack()
    CompanyAddressText.place(x= 0, y= 60)

    CompanyAddressEntry = Entry(Regist_window, width=13, textvariable= CompanyAddressEntry_Var)
    CompanyAddressEntry.place(x=130,y=60)
    
    CompanyCEONameText = Label(Regist_window,text="【대표명】",background="gainsboro",justify="left")
    CompanyCEONameText.pack()
    CompanyCEONameText.place(x= 0, y= 100)

    CompanyCEONameEntry = Entry(Regist_window, width=13, textvariable= CompanyCEONameEntry_Var)
    CompanyCEONameEntry.place(x=130,y=100)

    CompanyCEOTelText = Label(Regist_window,text="【대표 전화번호】\n  (업체 번호)",background="gainsboro",justify="left")
    CompanyCEOTelText.pack()
    CompanyCEOTelText.place(x= 0, y= 140)

    CompanyCEOTelEntry = Entry(Regist_window, width=13, textvariable= CompanyCEOTelEntry_Var)
    CompanyCEOTelEntry.place(x=130,y= 140)

    CompanyMailText = Label(Regist_window,text="【대표 메일】",background="gainsboro",justify="left")
    CompanyMailText.pack()
    CompanyMailText.place(x= 0, y= 190)

    CompanyMailEntry = Entry(Regist_window, width=13, textvariable=CompanyMailEntry_Var)
    CompanyMailEntry.place(x=130,y= 190)
#납품장소, 구매담당자, 전화번호
    vert_Frame_Sep = ttk.Separator(Regist_window, orient="vertical")	
    vert_Frame_Sep.place(relx=0.5, rely=0, relwidth=0, relheight=1)

    DeliveryAddressText = Label(Regist_window,text="【납품 장소】",background="gainsboro",justify="left")
    DeliveryAddressText.pack()
    DeliveryAddressText.place(x= 255, y= 20)

    DeliveryAddressEntry = Entry(Regist_window, width=13, textvariable=DeliveryAddressEntry_Var)
    DeliveryAddressEntry.place(x=380,y= 20)

    PurchasingManagerNameText = Label(Regist_window,text="【구매 담당자】",background="gainsboro",justify="left")
    PurchasingManagerNameText.pack()
    PurchasingManagerNameText.place(x= 255, y= 60)

    PurchasingManagerNameEntry = Entry(Regist_window, width=13, textvariable=PurchasingManagerNameEntry_Var)
    PurchasingManagerNameEntry.place(x=380,y= 60)

    PurchasingManagerTelText = Label(Regist_window,text="【담당자 전화】",background="gainsboro",justify="left")
    PurchasingManagerTelText.pack()
    PurchasingManagerTelText.place(x= 255, y= 100)

    PurchasingManagerTelEntry = Entry(Regist_window, width=13,textvariable=PurchasingManagerTelEntry_Var)
    PurchasingManagerTelEntry.place(x=380,y= 100)


    Name_Frame = Frame(Regist_window,width=250,height=90, relief="solid", bd=1,background="lightblue")
    Name_Frame.pack()
    Name_Frame.place(x=250,y=140)

    SettingFileNameText = Label(Regist_window,text="【현재 설정된 파일】",background="lightblue",justify="left")
    SettingFileNameText.pack()
    SettingFileNameText.place(x= 255, y= 150)


    SettingFileNameText1 = Label(Regist_window,textvariable=NowSettingFile,background="lightblue",justify="left")
    SettingFileNameText1.pack()
    SettingFileNameText1.place(x= 265, y= 180)

    
    Regist_window.config(menu=menubar)

#============================================================
#[메인- 탑패널] 이름 클릭
#============================================================
def BTN_Regist_click(event):
    pLog.append_log("이름 클릭:","창 실행")
    Regist_window=Toplevel()
    Regist_window.title("사용자 등록")
    Regist_window.geometry("500x250+250+300")
    Regist_window.config(background="gainsboro")
    Regist_window.resizable(False, False) #창 사이즈 변경 불가능
    Regist_window.lift()    
    Regist_window.attributes('-topmost', True)
    Regist_window.attributes('-topmost', False)

    menubar = Menu(Regist_window)
    menu1 = Menu(menubar,tearoff=0)
    menu1.add_command(label="새 파일",command=MenuBTN_NewFile)
    menu1.add_command(label="유저 파일 열기",command=MenuBTN_OpenFile)
    menu1.add_command(label="유저 파일 저장",command=MenuBTN_SaveFile)
    menu1.add_separator()
    menu1.add_command(label="창 닫기", command=Regist_window.destroy)
    menubar.add_cascade(label="파일",menu=menu1)

    menu2 = Menu(menubar,tearoff=0)
    menu2.add_command(label="사용할 유저 설정",command= MenuBTN_UserSetting)
    menu2.add_command(label="유저 목록 폴더 열기",command= BTN_Open_UserFolder)
    menubar.add_cascade(label="설정",menu=menu2)

    

    CompanyNameText = Label(Regist_window,text="【업체명】",background="gainsboro",justify="left")
    CompanyNameText.pack()
    CompanyNameText.place(x= 0, y= 20)

    CompanyNameEntry = Entry(Regist_window, width=13, textvariable=CompanyNameEntry_Var)
    CompanyNameEntry.place(x=130,y=20)

    CompanyAddressText = Label(Regist_window,text="【업체 주소】",background="gainsboro",justify="left")
    CompanyAddressText.pack()
    CompanyAddressText.place(x= 0, y= 60)

    CompanyAddressEntry = Entry(Regist_window, width=13, textvariable= CompanyAddressEntry_Var)
    CompanyAddressEntry.place(x=130,y=60)
    
    CompanyCEONameText = Label(Regist_window,text="【대표명】",background="gainsboro",justify="left")
    CompanyCEONameText.pack()
    CompanyCEONameText.place(x= 0, y= 100)

    CompanyCEONameEntry = Entry(Regist_window, width=13, textvariable= CompanyCEONameEntry_Var)
    CompanyCEONameEntry.place(x=130,y=100)

    CompanyCEOTelText = Label(Regist_window,text="【대표 전화번호】\n  (업체 번호)",background="gainsboro",justify="left")
    CompanyCEOTelText.pack()
    CompanyCEOTelText.place(x= 0, y= 140)

    CompanyCEOTelEntry = Entry(Regist_window, width=13, textvariable= CompanyCEOTelEntry_Var)
    CompanyCEOTelEntry.place(x=130,y= 140)

    CompanyMailText = Label(Regist_window,text="【대표 메일】",background="gainsboro",justify="left")
    CompanyMailText.pack()
    CompanyMailText.place(x= 0, y= 190)

    CompanyMailEntry = Entry(Regist_window, width=13, textvariable=CompanyMailEntry_Var)
    CompanyMailEntry.place(x=130,y= 190)
#납품장소, 구매담당자, 전화번호
    vert_Frame_Sep = ttk.Separator(Regist_window, orient="vertical")	
    vert_Frame_Sep.place(relx=0.5, rely=0, relwidth=0, relheight=1)

    DeliveryAddressText = Label(Regist_window,text="【납품 장소】",background="gainsboro",justify="left")
    DeliveryAddressText.pack()
    DeliveryAddressText.place(x= 255, y= 20)

    DeliveryAddressEntry = Entry(Regist_window, width=13, textvariable=DeliveryAddressEntry_Var)
    DeliveryAddressEntry.place(x=380,y= 20)

    PurchasingManagerNameText = Label(Regist_window,text="【구매 담당자】",background="gainsboro",justify="left")
    PurchasingManagerNameText.pack()
    PurchasingManagerNameText.place(x= 255, y= 60)

    PurchasingManagerNameEntry = Entry(Regist_window, width=13, textvariable=PurchasingManagerNameEntry_Var)
    PurchasingManagerNameEntry.place(x=380,y= 60)

    PurchasingManagerTelText = Label(Regist_window,text="【담당자 전화】",background="gainsboro",justify="left")
    PurchasingManagerTelText.pack()
    PurchasingManagerTelText.place(x= 255, y= 100)

    PurchasingManagerTelEntry = Entry(Regist_window, width=13,textvariable=PurchasingManagerTelEntry_Var)
    PurchasingManagerTelEntry.place(x=380,y= 100)


    Name_Frame = Frame(Regist_window,width=250,height=90, relief="solid", bd=1,background="lightblue")
    Name_Frame.pack()
    Name_Frame.place(x=250,y=140)

    SettingFileNameText = Label(Regist_window,text="【현재 설정된 파일】",background="lightblue",justify="left")
    SettingFileNameText.pack()
    SettingFileNameText.place(x= 255, y= 150)

    SettingFileNameText1 = Label(Regist_window,textvariable=NowSettingFile,background="lightblue",justify="left")
    SettingFileNameText1.pack()
    SettingFileNameText1.place(x= 265, y= 180)


    Regist_window.config(menu=menubar)

#============================================================
#[메인 - 메뉴바 - 설정] 프로그램 초기화
#============================================================
def BTN_SettingReset():
    pLog.append_log("버튼 동작 실행","프로그램 초기화")
    global USERNAMEDIR_VAR
    if tkinter.messagebox.askokcancel("리셋", "정말 초기화 하시겠습니까?"):
        userfile = open("./res/sys/systemp.txt",'w')
        print(0,file=userfile)#0
        print("",file=userfile)#1
        print(0,file=userfile)#2
        print(0,file=userfile)#3
        print(0,file=userfile)#4
        print(0,file=userfile)#5
        print(0,file=userfile)#6
        print("오전,오후,저녁",file=userfile)#7
        print(0,file=userfile)#8
        print(0,file=userfile)#9
        print(0,file=userfile)#10
        print(os.getcwd()+"\\Result\\작업 목록 데이터",file=userfile)#11
        print(os.getcwd()+"\\Result\\최종 결과",file=userfile)#12
        print(os.getcwd()+"\\Result\\요일 종합",file=userfile)#13
        print(0,file=userfile)#14
        print(os.getcwd()+"\\FolderList",file=userfile)#15
        userfile.close()
        NowSettingFile.set("")
        USERNAMEDIR_VAR = ""
        
        if os.path.exists("./TempFileList"):
            for file in os.scandir("./TempFileList"):
                os.remove(file.path)
        else:
            pass
                #tkinter.messagebox.showwarning("확인", "경로를 찾을수 없습니다.")


def BTN_settingTime():
    pLog.append_log("버튼 동작 실행:", "시간값 설정")
    global SettingProgram_window
    global DateTimeKind_Entry

    TimeKind_Var.set(DateTimeKind_Entry.get())
    
    tempKindTime=TimeKind_Var.get().split(",")
    for x in range(0,len(tempKindTime)):
        tempKindTime[x]=tempKindTime[x].replace(" ","")
    Right_Time_Combo['values']=tempKindTime

    pLog.append_log("tempKindTime 값: ", tempKindTime)
    pLog.append_log("TimeKind_Var 값:", TimeKind_Var.get())

def BTN_DirSetting():
    pLog.append_log("버튼 동작 실행:", "경로 설정")
    global Program_UseFile_Dir_Entry
    global Program_Save_Dir_Entry
    global Program_Result_Save_Dir_Entry
    global TotalDate_Result_Save_Dir_Entry
    
    Program_UseFile_Dir_Var.set(Program_UseFile_Dir_Entry.get())
    Program_Save_Dir_Var.set(Program_Save_Dir_Entry.get())
    Program_Result_Save_Dir_Var.set(Program_Result_Save_Dir_Entry.get())
    TotalDate_Result_Save_Dir_Var.set(TotalDate_Result_Save_Dir_Entry.get())

    pLog.append_log("Program_UseFile_Dir_Var 설정값:", Program_UseFile_Dir_Var.get())
    pLog.append_log("Program_Save_Dir_Var 설정값:", Program_Save_Dir_Var.get())
    pLog.append_log("Program_Result_Save_Dir_Var 설정값:", Program_Result_Save_Dir_Var.get())
    pLog.append_log("TotalDate_Result_Save_Dir_Var 설정값:", TotalDate_Result_Save_Dir_Var.get())
#============================================================
#[메인 - 메뉴바 - 설정] 프로그램 설정
#============================================================
def BTN_SettingProgram():
    pLog.append_log("버튼 동작 실행:", "프로그램 설정 창 실행")
    global SettingProgram_window
    global DateTimeKind_Entry
    global Program_Save_Dir_Entry
    global Program_Result_Save_Dir_Entry
    global TotalDate_Result_Save_Dir_Entry
    global Program_UseFile_Dir_Entry

    SettingProgram_window = Toplevel()
    SettingProgram_window.title("프로그램 설정")
    SettingProgram_window.geometry("500x380+850+300")
    SettingProgram_window.config(background="darkgray")
    SettingProgram_window.resizable(False, False) #창 사이즈 변경 불가능
    SettingProgram_window.lift()    
    SettingProgram_window.attributes('-topmost', True)
    SettingProgram_window.attributes('-topmost', False)

    notebook=tkinter.ttk.Notebook(SettingProgram_window, width=480, height=345)
    notebook.pack()

    SettingProgram_window_Frame1=tkinter.Frame(SettingProgram_window)
    notebook.add(SettingProgram_window_Frame1, text="프로그램 설정")

    DarkMode_CheckBox=Checkbutton(SettingProgram_window_Frame1,text="다크 모드",variable=DARKMODE_VAR,background="lightblue")
    DarkMode_CheckBox.place(x=15,y=20)

    DarkMode_Button = tkinter.Button(SettingProgram_window_Frame1, text='적용', command=BTN_DarkMode, overrelief="solid")
    DarkMode_Button.place(x=125,y=20)

    ToolTip_CheckBox=Checkbutton(SettingProgram_window_Frame1,text="툴팁 사용",variable=ToolTipCheckbox_Var,background="lightblue")
    ToolTip_CheckBox.place(x=15,y=60)

    TimeKind_Text = Label(SettingProgram_window_Frame1, text="시간 적용값", background="lightblue")
    TimeKind_Text.place(x=15,y=100)

    DateTimeKind_Entry = Entry(SettingProgram_window_Frame1, width=20, textvariable=TimeKind_Var)
    DateTimeKind_Entry.place(x=15,y=120)

    DateTimeKind_Button = tkinter.Button(SettingProgram_window_Frame1, text='적용', command=BTN_settingTime, overrelief="solid")
    DateTimeKind_Button.place(x=185,y=120)

    tempKindTime=TimeKind_Var.get().split(",")
    for x in range(0,len(tempKindTime)):
        tempKindTime[x]=tempKindTime[x].replace(" ","")


    SettingProgram_window_Frame2=tkinter.Frame(SettingProgram_window)
    notebook.add(SettingProgram_window_Frame2, text="경로 설정")

    Program_UseFile_Dir_Text = Label(SettingProgram_window_Frame2, text="사용할 폴더 경로", background="lightblue")
    Program_UseFile_Dir_Text.place(x=15,y=160)

    Program_UseFile_Dir_Entry = Entry(SettingProgram_window_Frame2, width=35, textvariable=Program_UseFile_Dir_Var)
    Program_UseFile_Dir_Entry.place(x=15,y=180)
    
    
    Program_Save_Dir_Text = Label(SettingProgram_window_Frame2, text="작업목록 저장 경로", background="lightblue")
    Program_Save_Dir_Text.place(x=15,y=210)

    Program_Save_Dir_Entry = Entry(SettingProgram_window_Frame2, width=35, textvariable=Program_Save_Dir_Var)
    Program_Save_Dir_Entry.place(x=15,y=230)

    Program_Result_Save_Dir_Text = Label(SettingProgram_window_Frame2, text="결과 저장 경로", background="lightblue")
    Program_Result_Save_Dir_Text.place(x=15,y=260)

    Program_Result_Save_Dir_Entry = Entry(SettingProgram_window_Frame2, width=35, textvariable=Program_Result_Save_Dir_Var)
    Program_Result_Save_Dir_Entry.place(x=15,y=280)

    TotalDate_Result_Save_Dir_Text = Label(SettingProgram_window_Frame2, text="요일별 결과 저장 경로", background="lightblue")
    TotalDate_Result_Save_Dir_Text.place(x=15,y=310)

    TotalDate_Result_Save_Dir_Entry = Entry(SettingProgram_window_Frame2, width=35, textvariable=TotalDate_Result_Save_Dir_Var)
    TotalDate_Result_Save_Dir_Entry.place(x=15,y=330)

    Save_Dir_Button = tkinter.Button(SettingProgram_window_Frame2, text='경로\n적용', command=BTN_DirSetting, overrelief="solid",height=7)
    Save_Dir_Button.place(x=305,y=180)


    SettingProgram_window_Frame3=tkinter.Frame(SettingProgram_window)
    notebook.add(SettingProgram_window_Frame3, text="로그")

    label3=tkinter.Label(SettingProgram_window_Frame3, text="페이지4의 내용")
    label3.pack()

    # frame4=tkinter.Frame(SettingProgram_window)
    # notebook.insert(2, frame4, text="페이지3")

    # label4=tkinter.Label(frame4, text="페이지3의 내용")
    # label4.pack()
#============================================================
#GUI_LOG
#============================================================
class GUIT():
    def __init__(self):
        self.tkhandler = Tk()
        self.tkhandler.geometry('800x760')
        self.tkhandler.title('Program Log')

        ###################공간띄우기##########################
        self.label_title = Label(self.tkhandler, text='')
        self.label_title.grid(row=0, column=0, sticky="w")
        ######################################################
 
        # 텍스트박스에 스크롤 연결
        self.scroll = Scrollbar(self.tkhandler, orient='vertical')
        self.lbox = Listbox(self.tkhandler, yscrollcommand=self.scroll.set, width=116,height=40,background="#161618",fg="white")
        self.scroll.config(command=self.lbox.yview)
        self.lbox.grid(row=0, column=0, columnspan=5, sticky="s")
 

    def append_log(self, comment, msg):
        global now
        self.now = str(datetime.datetime.now())[0:-7]
        cutLength = 100 
        Tempmsg = [msg[i:i+cutLength] for i in range(0, len(msg), cutLength)]
        if len(msg)+len(comment) > 100 or len(msg) > 100 or len(comment)>100:
            self.lbox.insert(END, "[{}] {}".format(self.now, comment))
            for x in range(0,len(Tempmsg)):
                self.lbox.insert(END, "{}".format(Tempmsg[x]))
        else:
            self.lbox.insert(END, "[{}] {} {}".format(self.now,comment, msg))
        self.lbox.update()
        self.lbox.see(END)
    def run(self):
        self.tkhandler.mainloop()

#============================================================
#[메인 - 메뉴바 - 도움말] 상세 사용법
#[간단 사용법] 상세 설명
#============================================================
def BTN_Net():
    pLog.append_log("버튼 동작 실행:", "메뉴바 - 상세 사용법")
    url = os.getcwd()+"/res/how_to_use/how_to_use.html"
    webbrowser.open(url)

#============================================================
#상세 사용법 [단축키] F1
#============================================================
def BTN_Net_Key(event):
    pLog.append_log("버튼 동작 실행:", "메뉴바 - 상세 사용법[단축키]")
    url = os.getcwd()+"/res/how_to_use/how_to_use.html"
    webbrowser.open(url)

#============================================================
#다크 모드  [] - Done
#============================================================
def BTN_DarkMode():
    pLog.append_log("버튼 동작 실행:", "다크모드 동작: {}".format(DARKMODE_VAR.get()))
    #0 = 일반 모드, 1 = 다크모드, 2 = 사용자 지정 모드


    if DARKMODE_VAR.get() == 1:
        #topFrame.config(bg="black")

        treeStyle = ttk.Style(root)
        treeStyle.theme_use("clam")
        treeStyle.configure("Treeview", background="#131314", 
                    fieldbackground="#131314", foreground="white")
        Right_InputDate_Text
        topFrame.config(background="#22242b")

        R_frame.config(background="#292b33")
        Right_List_Text.config(background="#292b33",foreground="white")
        Right_InputNum_Text.config(background="#292b33",foreground="white")
        Right_InputDate_Text.config(background="#292b33",foreground="white")
        Connect_Text.config(background="#292b33",foreground="white")
        Right_InputTime_Text.config(background="#292b33",foreground="white")
        Right_Uniqueness_Text.config(background="#292b33",foreground="white")
        Right_Export_Text.config(background="#292b33",foreground="white")
        Right_InputDate_Text.config(background="#292b33",foreground="white")
        Right_InputDate_infoText.config(background="#292b33",foreground="#5f6273")
        Right_Date_CheckBox.config(background="#292b33",foreground="white",selectcolor = "black")
        Right_Time_CheckBox.config(background="#292b33",foreground="white",selectcolor = "black")
        Right_CheckBox.config(background="#292b33",foreground="white",selectcolor = "black")
        Right_CheckBox1.config(background="#292b33",foreground="white",selectcolor = "black")
        Right_CheckBox2.config(background="#292b33",foreground="white",selectcolor = "black")

        ComboStyle = ttk.Style(root)
        ComboStyle.configure("TCombobox", background="#a1a5b6", 
                    fieldbackground="#a1a5b6", foreground="black")

        Top_Search_Entry.config(background="#494d5c",fg="white")
        Right_InputNum_Entry.config(background="#494d5c",fg="white")
        Right_InputDate1_Entry.config(background="#494d5c",fg="white")

        Right_InputDate1_Entry1.config(background="#494d5c",fg="white")
        Right_InputDate2_Entry.config(background="#494d5c",fg="white")
        Right_InputDate2_Entry1.config(background="#494d5c",fg="white")
        Right_Uniqueness_Entry.config(background="#494d5c",fg="white")

        Top_Search_Button.config(background="#2e313d", foreground="white")
        Right_AddItem_Button.config(background="#2e313d", foreground="white")
        Right_Remove_Button.config(background="#2e313d", foreground="white")
        Right_ALLRemove_Button.config(background="#2e313d", foreground="white")
        Right_Open_Dir_Button.config(background="#2e313d", foreground="white")
        Right_BanList_Button.config(background="#2e313d", foreground="white")
        Right_InputNum_Button.config(background="#2e313d", foreground="white")
        Right_InputNum_Button.config(background="#2e313d", foreground="white")
        Right_InputDate_Button.config(background="#2e313d", foreground="white")
        Right_InputTimeNum_Button.config(background="#2e313d", foreground="white")
        Right_Uniqueness_Button.config(background="#2e313d", foreground="white")
        Right_Empty_Export_Button.config(background="#2e313d", foreground="white")
        Right_Result_Export_Button.config(background="#2e313d", foreground="white")
        Right_TotalResult_Export_Button.config(background="#2e313d", foreground="white")

    elif DARKMODE_VAR.get() == 0:
        #topFrame.config(bg="darkgray")

        treeStyle = ttk.Style(root)
        treeStyle.theme_use("clam")
        treeStyle.configure("Treeview", background="white", 
                    fieldbackground="white", foreground="black")

        topFrame.config(background="darkgray")
        R_frame.config(background="lightblue")

        Right_List_Text.config(background="lightblue",foreground="black")
        Right_InputNum_Text.config(background="lightblue",foreground="black")
        Right_InputDate_Text.config(background="lightblue",foreground="black")
        Connect_Text.config(background="lightblue",foreground="black")
        Right_InputTime_Text.config(background="lightblue",foreground="black")
        Right_Uniqueness_Text.config(background="lightblue",foreground="black")
        Right_Export_Text.config(background="lightblue",foreground="black")
        Right_InputDate_Text.config(background="lightblue",foreground="black")
        Right_InputDate_infoText.config(background="lightblue",foreground="black")
        Right_Date_CheckBox.config(background="lightblue",foreground="black",selectcolor = "white")
        Right_Time_CheckBox.config(background="lightblue",foreground="black",selectcolor = "white")
        Right_CheckBox.config(background="lightblue",foreground="black",selectcolor = "white")
        Right_CheckBox1.config(background="lightblue",foreground="black",selectcolor = "white")
        Right_CheckBox2.config(background="lightblue",foreground="black",selectcolor = "white")

        ComboStyle = ttk.Style(root)
        ComboStyle.configure("TCombobox", background="white", 
                    fieldbackground="white", foreground="black")
        

        Top_Search_Entry.config(background="white",fg="black")
        Right_InputNum_Entry.config(background="white",fg="black")
        Right_InputDate1_Entry.config(background="white",fg="black")

        Right_InputDate1_Entry1.config(background="white",fg="black")
        Right_InputDate2_Entry.config(background="white",fg="black")
        Right_InputDate2_Entry1.config(background="white",fg="black")
        Right_Uniqueness_Entry.config(background="white",fg="black")

        Top_Search_Button.config(background="#e6e9f3", foreground="black")
        Right_AddItem_Button.config(background="#e6e9f3", foreground="black")
        Right_Remove_Button.config(background="#e6e9f3", foreground="black")
        Right_ALLRemove_Button.config(background="#e6e9f3", foreground="black")
        Right_Open_Dir_Button.config(background="#e6e9f3", foreground="black")

        Right_BanList_Button.config(background="#e6e9f3", foreground="black")
        Right_InputNum_Button.config(background="#e6e9f3", foreground="black")
        Right_InputNum_Button.config(background="#e6e9f3", foreground="black")
        Right_InputDate_Button.config(background="#e6e9f3", foreground="black")
        Right_InputTimeNum_Button.config(background="#e6e9f3", foreground="black")
        Right_Uniqueness_Button.config(background="#e6e9f3", foreground="black")
        Right_Empty_Export_Button.config(background="#e6e9f3", foreground="black")
        Right_Result_Export_Button.config(background="#e6e9f3", foreground="black")
        Right_TotalResult_Export_Button.config(background="#e6e9f3", foreground="black")

#============================================================
#[메인 - 메뉴바 - 도움말] 간단 사용법
#============================================================
def BTN_HOW_TO_USE():
    pLog.append_log("버튼 동작 실행:", "메뉴바 - 간단 사용법")
    howtouse_window = Toplevel()
    howtouse_window.title("사용법")
    howtouse_window.geometry("585x430+850+300")
    howtouse_window.config(background="darkgray")
    howtouse_window.resizable(False, False) #창 사이즈 변경 불가능
    howtouse_window.lift()    
    howtouse_window.attributes('-topmost', True)
    howtouse_window.attributes('-topmost', False)

    Frame1=Frame(howtouse_window,width=540,height=360, relief="solid", bd=1, background="white")
    Frame1.place(x=10,y=10)
    
    commentText = "※간단 사용 설명서 입니다. (사용자 등록을 완료하고 읽어주시기 바랍니다)\n"\
        "사진을 포함한 설명을 원하시면 상세 설명 버튼을 눌러주세요.\n\n" \
        "1. 메뉴 바(파일) - 파일 변환을 누릅니다.\n" \
        "1-1. 등록할 요리 파일(.hwp)이 있는 최상위 폴더를 선택합니다. \n\n" \
        "2. 우측 패널【항목】에 [항목 추가] 버튼을 눌러 원하는 항목을 추가합니다. \n" \
        "3. 추가가 완료 되었으면 요리를 선택하고【수량 입력】에 수량을 입력합니다. \n\n" \
        "4.【날짜 입력】에 해당하는 날짜를 입력해줍니다. \n" \
        "(1번째 칸은 발주일자, 2번째 칸은 납기 일자 입니다.) \n" \
        "5.【시간대 입력】에 항목별로 시간대를 설정 할수 있습니다.\n\n" \
        "6. 만약 발주 할때 특이사항이 있으면 【특이사항】에 적어주세요. \n" \
        "7. 모든 설정을 완료하였다면 【결과 내보내기】에서 결과를 내보냅니다.\n" \
        "----------------------------------------------------------------------\n" \
        "8. 항목 존재 유무를 확인하려면 상단 패널 에서 기준값 으로 검색합니다. \n\n" \
        "9. 메뉴 바 - 목록 저장하기를 통해 현재 작업중인 항목을 저장할 수 있습니다. \n" \
        "9-1. 메뉴 바 - 목록 불러오기를 통해 저장했던 항목을 불러올 수 있습니다. \n" \
        "※결과 내보내기에 재료의 수량이 집계됩니다.\n" \
        "우측 패널【결과 내보내기】요일 종합 버튼을 눌러 파일을 종합 할 수 있습니다.\n"\


    explanationText = Label(Frame1,text=commentText,background="white",justify="left")
    explanationText.pack()

    NetBtn = Button(howtouse_window, text="상세 설명", command= BTN_Net)
    NetBtn.pack()
    NetBtn.place(x=10,y=395)

    CloseBtn = Button(howtouse_window, text="닫기", command= howtouse_window.destroy,width=6)
    CloseBtn.pack()
    CloseBtn.place(x=510,y=395)
    
#============================================================
#[우측 패널 - [항목] - 버튼] 항목 추가
#============================================================
treeNumCount = 1
def BTN_AddItem():
    pLog.append_log("버튼 동작 실행:", "항목 - 추가")
    #item_t = filedialog.askdirectory()
    original_path = filedialog.askopenfilename(initialdir=Program_UseFile_Dir_Var.get(), multiple=True, title="Select file",
                                          filetypes=(("HWP files", "*.hwp"), ("all files", "*.*")))
    if not original_path:
        return
    global treeNumCount
    for x in range(0, len(original_path)):
        pLog.append_log("original_path: ", original_path[x])

    if len(listbox.tree.get_children())== 0: #리스트에 아무것도 없으면
        pass#넘기기
    else: #리스트에 뭔가 있으면
        tempList = []
        treeNumCount = 1
        #현재 트리에 저장된 값 temp에 옮겨놓기 -> iid, index 재정렬
        for y in listbox.tree.get_children():
            tempList =  listbox.tree.item(y)["values"]
            listbox.tree.delete(y)
            listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,tempList[1],tempList[2],tempList[3],tempList[4]))
            treeNumCount += 1
            
    for x in original_path: #선택한 경로에서 요리명, 자격증 종류 가져옴
        strx=str(x)
        fileName = os.path.basename(strx)
        fileName1 = fileName[:-4]
        C_fileName = re.sub(r"[^a-zA-Z|가-힣|/※~()]","",fileName1) 

        fileDir = os.path.dirname(strx)        
        licenseType = fileDir[fileDir.rfind("/")+2:fileDir.rfind("/")+4]

        if RemoveTimeCheckbox_Var.get() == 0:
            listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,1,C_fileName,licenseType,""))
            treeNumCount+=1
        else:
            islistboxin_listbox = 0
            
            for y in listbox.tree.get_children(): #트리에 있는 값들 중에서 
                if C_fileName == listbox.tree.item(y)["values"][2]: #파일명이랑 리스트에 파일명이랑 같으면
                    islistboxin_listbox = 1

            if islistboxin_listbox == 0: #중복 없음
                listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,1,C_fileName,licenseType,""))
                treeNumCount+=1
            else:
                tkinter.messagebox.showwarning("추가 오류", "이미 등록된 값이 있습니다.\n다시 선택해주세요.")
                break
            
#============================================================
#[프로그램][좌측 패널] 항목 추가 [드래그 앤 드롭]
#============================================================     
def drag_n_drop_AddItem(event):
    pLog.append_log("버튼 동작 실행:", "항목 - 추가 [드래그 앤 드롭]")
    global treeNumCount
    pLog.append_log("드래그앤 드롭 추가",event.data)
    print(event.data[0][0])
    if event.data[0][0] =="{" or event.data[0][0] ==" {" or event.data[0][0] =="{ " or event.data[0][0] ==" { ":
        event.data = event.data.replace("{","")
        sub_original_path = event.data.split("}")
    else:
        sub_original_path = event.data.split(".hwp")
        for x in range(0,len(sub_original_path)):
            if sub_original_path[x] == " " or sub_original_path[x] == "":
                pass
            else:
                sub_original_path[x] = sub_original_path[x]+".hwp"
    print("=====")
    print(sub_original_path)
    
    pLog.append_log("분리 값:",sub_original_path)
    # print(type(sub_original_path))
    # print(sub_original_path)
    original_path = []
    for x in sub_original_path:
        if x == "":
            break
        else: 
            if x[0] == " " or x[0] == "":
                original_path.append(x[1:])
            else:
                original_path.append(x)
    pLog.append_log("origin_Path:",original_path)        
    if len(listbox.tree.get_children())== 0: #리스트에 아무것도 없으면
        pass#넘기기
    else: #리스트에 뭔가 있으면
        tempList = []
        treeNumCount = 1
        #현재 트리에 저장된 값 temp에 옮겨놓기 -> iid, index 재정렬
        for y in listbox.tree.get_children():
            tempList =  listbox.tree.item(y)["values"]
            listbox.tree.delete(y)
            listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,tempList[1],tempList[2],tempList[3],tempList[4]))
            treeNumCount += 1

    for x in original_path: #선택한 경로에서 요리명, 자격증 종류 가져옴
        strx=str(x)
        fileName = os.path.basename(strx)
        fileName1 = fileName[:-4]
        C_fileName = re.sub(r"[^a-zA-Z|가-힣|/※~()]","",fileName1) 

        fileDir = os.path.dirname(strx)        
        licenseType = fileDir[fileDir.rfind("/")+2:fileDir.rfind("/")+4]
        
        if RemoveTimeCheckbox_Var.get() == 0:
            listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,1,C_fileName,licenseType,""))
            treeNumCount+=1
        else:
            islistboxin_listbox = 0
            
            for y in listbox.tree.get_children(): #트리에 있는 값들 중에서 
                if C_fileName == listbox.tree.item(y)["values"][2]: #파일명이랑 리스트에 파일명이랑 같으면
                    islistboxin_listbox = 1

            if islistboxin_listbox == 0: #중복 없음
                listbox.tree.insert('',index= treeNumCount,iid=treeNumCount, values=(treeNumCount,1,C_fileName,licenseType,""))
                treeNumCount+=1
            else:
                tkinter.messagebox.showwarning("추가 오류", "이미 등록된 값이 있습니다.\n다시 선택해주세요.")
                break

#============================================================
#[우측 패널 - [항목] - 버튼] 항목 삭제
#============================================================      
def BTN_RemoveItem():
    pLog.append_log("버튼 동작 실행:", "항목 - 항목 삭제")
    global treeNumCount
    selected_item = listbox.tree.selection() ## get selected item

    for x in selected_item:
        listbox.tree.delete(x)
        treeNumCount -= 1
    
    tempList = []
    treeNewCount = 1
    #현재 트리에 저장된 값 temp에 옮겨놓기 -> iid, index 재정렬
    for y in listbox.tree.get_children():
        tempList =  listbox.tree.item(y)["values"]
        listbox.tree.delete(y)
        listbox.tree.insert('',index= treeNewCount,iid=treeNewCount, values=(treeNewCount,tempList[1],tempList[2],tempList[3],tempList[4]))
        treeNewCount += 1
    treeNumCount = treeNewCount

#============================================================
#[우측 패널 - [항목] - 버튼] 전체 삭제
#============================================================
def BTN_ALL_RemoveItem():
    pLog.append_log("버튼 동작 실행:", "항목 - 전체삭제")
    global treeNumCount
    treeNumCount = 1
    for y in listbox.tree.get_children():
        listbox.tree.delete(y)

#============================================================
#[메인 - 메뉴바 - 파일] 새 파일
#============================================================   
def BTN_ALL_RemoveItem_RootMenu():
    pLog.append_log("버튼 동작 실행:", "메뉴바 - 새 파일")
    global treeNumCount
    if treeNumCount >= 2:
        if tkinter.messagebox.askokcancel("새 파일", "현재 진행하던 작업은 삭제됩니다.\n새 파일을 만드시겠습니까?"):
            treeNumCount = 1
            for y in listbox.tree.get_children():
                listbox.tree.delete(y)
    else:
        treeNumCount = 1
        for y in listbox.tree.get_children():
            listbox.tree.delete(y)
#============================================================
#[메인 - 메뉴바 - 파일] 새 파일 [단축키] Ctrl + N
#============================================================   
def BTN_ALL_RemoveItem_RootMenu_Key(event):
    pLog.append_log("버튼 동작 실행:", "메뉴바 - 새 파일[단축키]")
    global treeNumCount
    if treeNumCount >= 2:
        if tkinter.messagebox.askokcancel("새 파일", "현재 진행하던 작업은 삭제됩니다.\n새 파일을 만드시겠습니까?"):
            treeNumCount = 1
            for y in listbox.tree.get_children():
                listbox.tree.delete(y)
    else:
        treeNumCount = 1
        for y in listbox.tree.get_children():
            listbox.tree.delete(y)
#============================================================
#[메인 - 항목 - 버튼] 폴더열기 
#============================================================  
def BTN_Start_AddItem():
    os.startfile(Program_UseFile_Dir_Var.get())
#============================================================
#[프로그램][수량 입력]수량 입력창 활성화 될때 
#============================================================
def focus_InputNum(event):
    Right_InputNum_Entry.delete(0,END)

#============================================================
#[메인 - 우측패널 - 수량 입력] 수량 추가 [엔터키]
#============================================================
def BTN_InputNum_Return(event):
    
    if Right_InputNum_Entry.get() != "":
        if Right_InputNum_Entry.get().isdigit() == True:
            for x in listbox.tree.selection():
                listbox.tree.item(x, text="",values=(listbox.tree.item(x).get('values')[0],Right_InputNum_Entry.get(),listbox.tree.item(x).get('values')[2],listbox.tree.item(x).get('values')[3],listbox.tree.item(x).get('values')[4]))
        else:
            tkinter.messagebox.showwarning("입력 오류", "잘못된 입력값 입니다.\n유효한 숫자를 입력해주세요.")

#============================================================
#[메인 - 우측패널 - 수량 입력 - 버튼] 수량 추가
#============================================================
def BTN_InputNum():
    
    if Right_InputNum_Entry.get() != "":
        if Right_InputNum_Entry.get().isdigit() == True:
            for x in listbox.tree.selection():
                listbox.tree.item(x, text="",values=(listbox.tree.item(x).get('values')[0],Right_InputNum_Entry.get(),listbox.tree.item(x).get('values')[2],listbox.tree.item(x).get('values')[3],listbox.tree.item(x).get('values')[4]))
        else:
            tkinter.messagebox.showwarning("입력 오류", "잘못된 입력값 입니다.\n유효한 숫자를 입력해주세요.")

#============================================================
#[메인 - 우측패널 - 날짜 입력] 날짜 입력
#============================================================
def BTN_InputDate():
    try:
        Right_InputDate1_Entry_Week.set(days[datetime.date(int(Right_InputDate1_Entry_Date.get().split(".")[0]),int(Right_InputDate1_Entry_Date.get().split(".")[1]),int(Right_InputDate1_Entry_Date.get().split(".")[2])).weekday()])
        Right_InputDate2_Entry_Week.set(days[datetime.date(int(Right_InputDate2_Entry_Date.get().split(".")[0]),int(Right_InputDate2_Entry_Date.get().split(".")[1]),int(Right_InputDate2_Entry_Date.get().split(".")[2])).weekday()])
        if int(Right_InputDate2_Entry_Date.get().split(".")[2]) < int(Right_InputDate1_Entry_Date.get().split(".")[2]):
            tkinter.messagebox.showwarning("입력 오류", "잘못된 입력값 입니다.\n기준점이 되는 날짜보다 작을수 없습니다.")
            Right_InputDate2_Entry_Week.set(days[datetime.date(int(Right_InputDate1_Entry_Date.get().split(".")[0]),int(Right_InputDate1_Entry_Date.get().split(".")[1]),int(Right_InputDate1_Entry_Date.get().split(".")[2])).weekday()])
            Right_InputDate2_Entry_Date.set(Right_InputDate1_Entry_Date.get())
        elif int(Right_InputDate2_Entry_Date.get().split(".")[2]) >= int(Right_InputDate1_Entry_Date.get().split(".")[2]):
            tkinter.messagebox.showinfo("확인", "정상적인 입력값 입니다.")
    except:
        tkinter.messagebox.showerror("잘못된 날짜 형식", "잘못된 날짜 형식 입니다.\n다시 입력 해주세요.")

#============================================================
#[메인 - 우측패널 - 날짜 입력 - 버튼] 입력 확인
#============================================================
def BTN_InputDate_Check():
    try:
        Right_InputDate1_Entry_Week.set(days[datetime.date(int(Right_InputDate1_Entry_Date.get().split(".")[0]),int(Right_InputDate1_Entry_Date.get().split(".")[1]),int(Right_InputDate1_Entry_Date.get().split(".")[2])).weekday()])
        Right_InputDate2_Entry_Week.set(days[datetime.date(int(Right_InputDate2_Entry_Date.get().split(".")[0]),int(Right_InputDate2_Entry_Date.get().split(".")[1]),int(Right_InputDate2_Entry_Date.get().split(".")[2])).weekday()])
        if int(Right_InputDate2_Entry_Date.get().split(".")[2]) < int(Right_InputDate1_Entry_Date.get().split(".")[2]):
            tkinter.messagebox.showwarning("입력 오류", "잘못된 입력값 입니다.\n기준점이 되는 날짜보다 작을수 없습니다.")
            Right_InputDate2_Entry_Week.set(days[datetime.date(int(Right_InputDate1_Entry_Date.get().split(".")[0]),int(Right_InputDate1_Entry_Date.get().split(".")[1]),int(Right_InputDate1_Entry_Date.get().split(".")[2])).weekday()])
            Right_InputDate2_Entry_Date.set(Right_InputDate1_Entry_Date.get())
        elif int(Right_InputDate2_Entry_Date.get().split(".")[2]) >= int(Right_InputDate1_Entry_Date.get().split(".")[2]):
            pass
    except:
        tkinter.messagebox.showerror("잘못된 날짜 형식", "잘못된 날짜 형식 입니다.\n다시 입력 해주세요.")

#============================================================
#[메인 - 우측패널 - 시간대 입력 - 버튼] 확인
#============================================================
def BTN_InputTimeNum():
    
    if Right_Time_Combo.get() != "":
            for x in listbox.tree.selection():
                listbox.tree.item(x, text="",values=(listbox.tree.item(x).get('values')[0],listbox.tree.item(x).get('values')[1],listbox.tree.item(x).get('values')[2],listbox.tree.item(x).get('values')[3],Right_Time_Combo.get()))
    else:
        tkinter.messagebox.showwarning("입력 오류", "잘못된 입력값 입니다.\n유효한 시간값을 입력해주세요.")


#============================================================
#[메인 - 상단패널 - 버튼] 검색 [엔터]
#============================================================
def BTN_SearchItem_Return(event):
    isCount = False  
    for child in listbox.tree.get_children():
        if Top_Combo.get() == "번호":
            if int(listbox.tree.item(child)["values"][0]) == int(Top_Search_Entry.get()):
                isCount =True 
                break
            else:
                isCount = False
                continue
        elif Top_Combo.get() == "수량":
            if int(listbox.tree.item(child)["values"][1]) == int(Top_Search_Entry.get()):
                isCount =True
                break
            else:
                isCount = False
                continue
        elif Top_Combo.get() == "이름":
            if str(listbox.tree.item(child)["values"][2].replace(" ","")) == str(Top_Search_Entry.get().replace(" ","")):
                isCount =True
                break
            else:
                isCount = False
                continue

    if isCount ==True:
        listbox.tree.selection_set(listbox.tree.item(child)["values"][0])
        tkinter.messagebox.showinfo("확인", "검색하신 항목이 목록에 존재합니다.")
    else:
        tkinter.messagebox.showwarning("확인", "검색하신 항목이 목록에 없습니다.")

#============================================================
#[메인 - 상단패널 - 버튼] 검색
#============================================================
def BTN_SearchItem():
    isCount = False  
    for child in listbox.tree.get_children():
        if Top_Combo.get() == "번호":
            if int(listbox.tree.item(child)["values"][0]) == int(Top_Search_Entry.get()):
                isCount =True
                break
            else:
                isCount = False
                continue
        elif Top_Combo.get() == "수량":
            if int(listbox.tree.item(child)["values"][1]) == int(Top_Search_Entry.get()):
                isCount =True
                break
            else:
                isCount = False
                continue
        elif Top_Combo.get() == "이름":
            if str(listbox.tree.item(child)["values"][2].replace(" ","")) == str(Top_Search_Entry.get().replace(" ","")):
                isCount =True
                break
            else:
                isCount = False
                continue

    if isCount ==True:
        listbox.tree.selection_set(listbox.tree.item(child)["values"][0])
        tkinter.messagebox.showinfo("확인", "검색하신 항목이 목록에 존재합니다.")
    else:
        tkinter.messagebox.showwarning("확인", "검색하신 항목이 목록에 없습니다.")

#============================================================
#[메인 - 우측패널 - 특이사항 - 버튼] 입력 확인
#============================================================
def BTN_UniquenessText():
    tkinter.messagebox.showinfo("입력 확인", Right_Uniqueness_Entry.get())


#============================================================
#[프로그램]내보낼 데이터 정제
#============================================================
refineList = []
TempRefineList = []
refineListItemName = []
def refine_table():

    global TempRefineList
    global refineList
    global refineListItemName

    TempRefineList =[]
    refineList = []
    path=[]

    #"C:/Users/nsn04/OneDrive/바탕 화면/integrative/TempFileList/2.도미머리맑은국.txt"
    #"C:/Users/nsn04/OneDrive/바탕 화면/integrative/TempFileList/1.오징어냉채.txt"
    for c in range(1,treeNumCount): #트리에 카운트 만큼 반복
        for infileName in os.listdir("./TempFileList/"):
            reinFileName=infileName.replace(" ","")
            if reinFileName[reinFileName.find(".")+1:-4] == listbox.tree.item(c)["values"][2]:

                refineListItemName.append(infileName[infileName.find(".")+1:-4])
                path.append("./TempFileList/"+infileName)
            
    treelistCount = 0
    #31= 0~30
    for q in range(0,treeNumCount-1):
        with open(path[treelistCount]) as file:
            lines = file.readlines()
        lines = [line.rstrip('\n') for line in lines]#엔터 제거

        strText = str(lines)
        orignalText = strText[strText.find("비      고")+10:strText.find("※국가기술자격")]
        orignalText=orignalText.replace("'","")
        orignalText=orignalText.replace("\\u3000","")
        listText=orignalText.split(",")

        #요리 내 재료 개수 카운트
        count = 0 #그냥 1씩 증가만 하는 카운트용 변수
        listCount = 0 #요리에 재료의 개수
        for x in listText:
            if count%6==0 :
                listCount += 1
                if x == "" or x == " ": #리스트에 1번째 값이 없으면 중지
                    break
            count+=1

        listboxVar = listbox.tree.item(int(treelistCount+1))["values"][1]
        #print(listbox.tree.item(int(treelistCount+1))["values"][2])
        count=0
        for i in range(listCount-1):
            nlist = []
            for j in range(6): #항목은 6개씩 끊어서
                if count%6==0:
                    listText[count] = listText[count].replace(" ","")
                if count%6==1:
                    listText[count] = listText[count].lstrip(" ")
                if count%6==2:
                    listText[count] = listText[count].lstrip(" ")
                if count%6==3:
                    listText[count] = listText[count].replace(" ","")
                if count%6==4:
                    listText[count] = listText[count].replace(" ","")
                    if "/" in listText[count]:
                        a = Fraction(listText[count])
                        listText[count] = round(float(a),3)
                    listText[count] = round(float(listText[count]) * float(listboxVar),2)
                if count%6==5:
                    listText[count] = listText[count].lstrip(" ")

                nlist.append(listText[count])
                count+=1
            TempRefineList.append(nlist)
        treelistCount+=1    

    goodsList = []
    detectionList = []
    if RemoveReduplicationCheckbox_Var.get() == 0: #중복 허용
        treelistCount = 0
        for x in range(0,len(TempRefineList)):#불러온 임시 목록 개수
            TempRefineList.sort(key= lambda x:x[1])
            refineList.append(TempRefineList[treelistCount])
            treelistCount +=1

    elif RemoveReduplicationCheckbox_Var.get() == 1: #중복 제거
        for i in TempRefineList:
            goodsList.append(i[1]) #재료명만

        #tempList.sort(key= lambda x:x[1])
        for t in goodsList:
            res_list = [i for i, value in enumerate(TempRefineList) if value[1] == t]#중복되는 값의 위치를 찾음
            #print("New indices list : " + str(res_list)) 
            detectionList.append(res_list)

        goodsIndex=list(set([tuple(set(val))for val in detectionList])) #중복값 없는 재료 위치 인덱스
        tuple_to_list = [list(row) for row in goodsIndex] #위에 모든 튜플 적용값을 전부 리스트로 깔끔하게 변환

        if os.path.isfile('./res/banList/BanListFile.txt') == True:
            openbanfile = open('./res/banList/BanListFile.txt','r',encoding="utf-8")
            readbantext = openbanfile.read()
            readbansplit_text = readbantext.split("\n")
            openbanfile.close()
        else:
            tkinter.messagebox.showwarning("파일 없음", "banList폴더에 BanListFile.txt 파일이 없습니다. \n 생성해주세요.")
        readbansplit_text = list(filter(None, readbansplit_text))

        for val in tuple_to_list:

            if len(val) <  2 :#인덱스가 한개인값, 즉 중복 없는값은 바로 최종 리스트에 넣기
                subList = []
                subUnit = ''
                subNote = ''
                if StandardCheckbox_Var.get() == 1:
                    subUnit = ''
                else:
                    subUnit = TempRefineList[val[0]][2]

                if NoteCheckBox_Var.get() == 1:
                    subNote = ''
                else:
                    subNote = TempRefineList[val[0]][5]
                subList = ['0',TempRefineList[val[0]][1],subUnit,TempRefineList[val[0]][3],TempRefineList[val[0]][4],subNote]
                refineList.append(subList)

            else: #같은 이름의 값 처리

                temp_unitList = [] # 모든 단위 종류 
                resultVar = [] #refine에 올릴 변수

                for j in val:
                    temp_unitList.append(TempRefineList[j][3])

                set_pure_temp_unit = list(set(temp_unitList))
                
                print("val: " + str(val))
                for i in range(len(set_pure_temp_unit)):#단위 개수만큼 반복
                    mergeValue = 0
                    mergeStandard = ''
                    mergeNote = ''
                    for j in val:#j는 val에 인덱스 1개
                        if set_pure_temp_unit[i] == TempRefineList[j][3]: #단위에 따른 개수 - ㅇㅋ
                            mergeValue = round(mergeValue + float(TempRefineList[j][4]),2)
                            if StandardCheckbox_Var.get() == 0: #제거
                                if mergeStandard == '':
                                    mergeStandard = TempRefineList[j][2]
                                else:
                                    if mergeStandard == TempRefineList[j][2]:
                                        pass
                                    else  :
                                        if str(mergeStandard).replace(" ","") == str(TempRefineList[j][2]).replace(" ",""):
                                            pass
                                        else:
                                            mergeStandard += "\n" + TempRefineList[j][2]
                            else:
                                mergeStandard = ''

                            if NoteCheckBox_Var.get() == 0:    
                                if mergeNote == '':
                                    mergeNote = TempRefineList[j][5]
                                else:
                                    if mergeNote == TempRefineList[j][5]:
                                        pass
                                    else:
                                        if str(mergeNote).replace(" ","") == str(TempRefineList[j][5]).replace(" ",""):
                                            pass
                                        else:
                                            mergeNote += "\n" + TempRefineList[j][5]
                            else:
                                mergeNote = ''
                    resultVar = ['0', TempRefineList[val[0]][1],mergeStandard,set_pure_temp_unit[i],mergeValue,mergeNote]
                    refineList.append(resultVar)
                    print("result: "+ str(resultVar))
   
    tempDel_List = []
    for y in readbansplit_text:
        for x in refineList:
            if str(x[1]).replace(" ","") == str(y).replace(" ",""):
                tempDel_List.append(x)
                
    for y in tempDel_List:
        for x in refineList:
            if x == y:
                refineList.remove(x)

    global restoreNum
    restoreNum = 1            
    for x in refineList:
        x[0] = restoreNum
        restoreNum +=1
    # with open('./TempFileList/Result1.csv', 'w', encoding='utf-8') as fileT:
    #     writer = csv.writer(fileT)
    #     writer.writerow(TempRefineList)

#============================================================
#[프로그램]내보낼 데이터 정제
#============================================================
refineList = []
TempRefineList = []
def refine_table_Time():

    global refineList
    global TempRefineList
    global restoreNumList

    refineList = []
    restoreNumList = []
    TempRefineList =[]
    path=[]

    TimeKindList = [] #treeList 항목 에 시간대 - 전체 시간값 종류
    for x in range(1,treeNumCount):
        TimeKindList.append(listbox.tree.item(x)["values"][4])
    
    
    TimeKindList_set = [] #treeList 항목 에 전체 시간값 종류중 중복 제거된 시간값
    for x in TimeKindList:
        if x not in TimeKindList_set:
            TimeKindList_set.append(x)
    TimeKindList_set.sort(reverse=False)#중복제거된 값 - 정렬
    print("tree에 있는 중복 제거 값: "+str(TimeKindList_set))

    #사용자가 설정한 시간 값
    tempKindTime=TimeKind_Var.get().split(",")
    for x in range(0,len(tempKindTime)):
        tempKindTime[x]=tempKindTime[x].replace(" ","")
    print("사용자 시간 값: "+str(tempKindTime))
    #TODO: 사용자가 설정한 시간값에 개수를 맞춰야함
    #사용자 설정값 - tree 값
    #오전 오후 저녁 - 오전 저녁 = 오후
    #오전 오후 저녁 - 오전 오후 저녁 기타 = 기타

    CheckInTimeList = list(set(tempKindTime)-set(TimeKindList_set))
    print(CheckInTimeList) #사용자가 설정한 시간값 중에 없는거 찾기

    CheckOutTimeList = list(set(list(set(tempKindTime)^set(TimeKindList_set))) - set(CheckInTimeList))
    print(CheckOutTimeList) #사용자가 설정한 시간값 이후에 추가된값 판별용

    refinetempKindList=tempKindTime+CheckOutTimeList
    TimeKindList_count = [] #treeList에 각 시간값별 항목 개수
    for x in range(0,len(refinetempKindList)):
        TimeKindList_count.append(TimeKindList.count(refinetempKindList[x]))
        #ex) 오전-2,오후-5,저녁-3 [2,5,3]
    print(TimeKindList_count)

    TimeKindList_set_List = [] #오전, 오후, 저녁 순으로 정렬된 항목들
    for y in refinetempKindList:
        for x in range(1,treeNumCount):
            if listbox.tree.item(x)["values"][4] == y:
                TimeKindList_set_List.append(listbox.tree.item(x)["values"][2])
    print("=====")
    print(TimeKindList_set_List)

    kindListCount = 0
    for y in range(0, len(TimeKindList_count)):#시간대 개수
        for x in range(0,TimeKindList_count[y]):#시간에 해당하는 목록 개수[오전4,오후3,저녁,3]
            for infileName in os.listdir("./TempFileList"): #이 경로에 있는 파일 다 불러옴
                reinFileName=infileName.replace(" ","")
                if reinFileName[reinFileName.find(".")+1:-4] == TimeKindList_set_List[kindListCount]:
                    path.append("./TempFileList/"+infileName)
            kindListCount += 1
    print(path)

    kindListCount = 0
    for y in range(0, len(TimeKindList_count)):#시간대 개수
        TempRefineList = []
        for q in range(0,TimeKindList_count[y]):#시간에 해당하는 목록 개수[오전4,오후3,저녁,3]
            with open(path[kindListCount]) as file: 
                lines = file.readlines() #파일 안의 내용물
            lines = [line.rstrip('\n') for line in lines]#엔터 제거

            strText = str(lines)
            orignalText = strText[strText.find("비      고")+10:strText.find("※국가기술자격")]
            orignalText=orignalText.replace("'","")
            orignalText=orignalText.replace("\\u3000","")
            listText=orignalText.split(",") #정제된 텍스트 리스트
            
            #요리 내 재료 개수 카운트
            count = 0 #그냥 1씩 증가만 하는 카운트용 변수
            listCount = 0 #요리에 재료의 개수
            for x in listText:
                if count%6==0 :
                    listCount += 1
                    if x == "" or x == " ": #리스트에 1번째 값이 없으면 중지
                        break
                count+=1
            
            listboxVar = 0
            for x in range(1,treeNumCount):
                if TimeKindList_set_List[kindListCount]==listbox.tree.item(x)["values"][2]:
                    listboxVar = listbox.tree.item(x)["values"][1]
            count=0
            for i in range(listCount-1):
                nlist = []
                for j in range(6): #항목은 6개씩 끊어서
                    if count%6==0:
                        listText[count] = listText[count].replace(" ","")
                    if count%6==1:
                        listText[count] = listText[count].lstrip(" ")
                    if count%6==2:
                        listText[count] = listText[count].lstrip(" ")
                    if count%6==3:
                        listText[count] = listText[count].replace(" ","")
                    if count%6==4:
                        listText[count] = listText[count].replace(" ","")
                        if "/" in listText[count]:
                            a = Fraction(listText[count])
                            listText[count] = round(float(a),3)
                        listText[count] = round(float(listText[count]) * float(listboxVar),2)
                    if count%6==5:
                        listText[count] = listText[count].lstrip(" ")

                    nlist.append(listText[count])
                    count+=1
                TempRefineList.append(nlist)

            kindListCount += 1

        LocalTempList = [] #여기에서만 쓸 임시 리스트
        goodsList = [] #단순 항목 이름
        detectionList = [] #중복 제거될 항목 이름

        if RemoveReduplicationCheckbox_Var.get() == 0: #중복 허용
            treelistCount = 0
            for x in range(0,len(TempRefineList)):#불러온 임시 목록 개수
                TempRefineList.sort(key= lambda x:x[1])
                LocalTempList.append(TempRefineList[treelistCount])
                treelistCount +=1

        elif RemoveReduplicationCheckbox_Var.get() == 1: #중복 제거
            for i in TempRefineList:
                if i[1][-1] == " ":
                    goodsList.append(i[1][:-1])
                else:
                    goodsList.append(i[1]) 

            #tempList.sort(key= lambda x:x[1])
            for t in goodsList:
                res_list = [i for i, value in enumerate(TempRefineList) if value[1] == t]#중복되는 값의 위치를 찾음
                #print("New indices list : " + str(res_list)) 
                detectionList.append(res_list)
            
            goodsIndex=list(set([tuple(set(val))for val in detectionList])) #중복값 없는 재료 위치 인덱스
            tuple_to_list = [list(row) for row in goodsIndex] #위에 모든 튜플 적용값을 전부 리스트로 깔끔하게 변환
            
            for val in tuple_to_list:
                if len(val) <  2 :#인덱스가 한개인값, 즉 중복 없는값은 바로 최종 리스트에 넣기
                    subList = []
                    subUnit = ''
                    subNote = ''
                    if StandardCheckbox_Var.get() == 1:
                        subUnit = ''
                    else:
                        subUnit = TempRefineList[val[0]][2]
                    if NoteCheckBox_Var.get() == 1:
                        subNote = ''
                    else:
                        subNote = TempRefineList[val[0]][5]
                    subList = ['0',TempRefineList[val[0]][1],subUnit,TempRefineList[val[0]][3],TempRefineList[val[0]][4],subNote]
                    LocalTempList.append(subList)

                else: #같은 이름의 값 처리
                    temp_unitList = [] # 모든 단위 종류
                    resultVar = [] #refine에 올릴 변수
                    for j in val:
                        temp_unitList.append(TempRefineList[j][3])

                    set_pure_temp_unit = list(set(temp_unitList))
                    

                    for i in range(len(set_pure_temp_unit)):#단위 개수만큼 반복
                        mergeValue = 0
                        mergeStandard = ''
                        mergeNote = ''
                        for j in val:#j는 val에 인덱스 1개
                            if set_pure_temp_unit[i] == TempRefineList[j][3]: #단위에 따른 개수 - ㅇㅋ
                                mergeValue = round(mergeValue + float(TempRefineList[j][4]),2)
                                if StandardCheckbox_Var.get() == 0: #제거
                                    if mergeStandard == '':
                                        mergeStandard = TempRefineList[j][2]
                                    else:
                                        if mergeStandard == TempRefineList[j][2]:
                                            pass
                                        else  :
                                            if str(mergeStandard).replace(" ","") == str(TempRefineList[j][2]).replace(" ",""):
                                                pass
                                            else:
                                                mergeStandard += "\n" + TempRefineList[j][2]
                                else:
                                    mergeStandard = ''

                                if NoteCheckBox_Var.get() == 0:    
                                    if mergeNote == '':
                                        mergeNote = TempRefineList[j][5]
                                    else:
                                        if mergeNote == TempRefineList[j][5]:
                                            pass
                                        else:
                                            if str(mergeNote).replace(" ","") == str(TempRefineList[j][5]).replace(" ",""):
                                                pass
                                            else:
                                                mergeNote += "\n" + TempRefineList[j][5]
                                else:
                                    mergeNote = ''
                        resultVar = ['0', TempRefineList[val[0]][1],mergeStandard,set_pure_temp_unit[i],mergeValue,mergeNote]
                        LocalTempList.append(resultVar)        

        if os.path.isfile('./res/banList/BanListFile.txt') == True:
            openbanfile = open('./res/banList/BanListFile.txt','r',encoding="utf-8")
            readbantext = openbanfile.read()
            readbansplit_text = readbantext.split("\n")
            openbanfile.close()
        else:
            tkinter.messagebox.showwarning("파일 없음", "banList폴더에 BanListFile.txt 파일이 없습니다. \n 생성해주세요.")
        readbansplit_text = list(filter(None, readbansplit_text))
        
        #중복 제거
        tempDel_List = []
        for y in readbansplit_text:
            for x in LocalTempList:
                if str(x[1]).replace(" ","") == str(y).replace(" ",""):
                    tempDel_List.append(x)
                    
        for y in tempDel_List:
            for x in LocalTempList:
                if x == y:
                    LocalTempList.remove(x)
        
        #번호 재정렬
        restoreNum = 1            
        for x in LocalTempList:
            x[0] = restoreNum
            restoreNum +=1

        restoreNumList.append(restoreNum-1)
        refineList.append(LocalTempList)
    print("=====")
    print(refineList)
    print(restoreNumList)
    # with open('./TempFileList/Result1.csv', 'w', encoding='utf-8') as fileT:
    #     writer = csv.writer(fileT)
    #     writer.writerow(TempRefineList)


#if RemoveTimeCheckbox_Var.get() == 1: #시간값 제거 하겠다(기본값)
#============================================================
#[프로그램] hwp 폰트 세팅(hwp, 폰트 사이즈, 1:폰트 바꾸기, 2:폰트 사이즈만 바꾸기, 3: 폰트만 바꾸기)
#============================================================
def hwp_fontSetting(hwp,font,size,isType):
    if isType == 1:
        hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
        hwp.HParameterSet.HCharShape.FaceNameUser = font
        hwp.HParameterSet.HCharShape.FontTypeUser = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameSymbol = font
        hwp.HParameterSet.HCharShape.FontTypeSymbol = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameOther = font
        hwp.HParameterSet.HCharShape.FontTypeOther = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameJapanese = font
        hwp.HParameterSet.HCharShape.FontTypeJapanese = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameHanja = font
        hwp.HParameterSet.HCharShape.FontTypeHanja = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameLatin = font
        hwp.HParameterSet.HCharShape.FontTypeLatin = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameHangul = font
        hwp.HParameterSet.HCharShape.FontTypeHangul = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.Height = hwp.PointToHwpUnit(size)
        hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)
    elif isType == 2:
        hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
        hwp.HParameterSet.HCharShape.Height = hwp.PointToHwpUnit(size)
        hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)
    elif isType == 3:
        hwp.HAction.GetDefault("CharShape", hwp.HParameterSet.HCharShape.HSet)
        hwp.HParameterSet.HCharShape.FaceNameUser = font
        hwp.HParameterSet.HCharShape.FontTypeUser = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameSymbol = font
        hwp.HParameterSet.HCharShape.FontTypeSymbol = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameOther = font
        hwp.HParameterSet.HCharShape.FontTypeOther = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameJapanese = font
        hwp.HParameterSet.HCharShape.FontTypeJapanese = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameHanja = font
        hwp.HParameterSet.HCharShape.FontTypeHanja = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameLatin = font
        hwp.HParameterSet.HCharShape.FontTypeLatin = hwp.FontType("TTF")
        hwp.HParameterSet.HCharShape.FaceNameHangul = font
        hwp.HParameterSet.HCharShape.FontTypeHangul = hwp.FontType("TTF")
        hwp.HAction.Execute("CharShape", hwp.HParameterSet.HCharShape.HSet)

#============================================================
#[프로그램] hwp 표 만들기(hwp, col값,row값,col(가로) 크기 리스트, 세로 크기)
#============================================================
def hwp_createTable(hwp,representTable_col_var,representTable_row_var,representTable_col_list,representTable_row_list):
    hwp.HAction.GetDefault("TableCreate", hwp.HParameterSet.HTableCreation.HSet)
    hwp.HParameterSet.HTableCreation.Cols = representTable_col_var #가로
    hwp.HParameterSet.HTableCreation.Rows = representTable_row_var #세로
    hwp.HParameterSet.HTableCreation.WidthType = 2
    hwp.HParameterSet.HTableCreation.HeightType = 1

    hwp.HParameterSet.HTableCreation.CreateItemArray("ColWidth", representTable_col_var)
    for x in range(0,representTable_col_var):
        hwp.HParameterSet.HTableCreation.ColWidth.SetItem(x, hwp.MiliToHwpUnit(representTable_col_list[x]))  # 1열
    
    hwp.HParameterSet.HTableCreation.CreateItemArray("RowHeight", representTable_row_var)
    for x in range(0,representTable_row_var):
        hwp.HParameterSet.HTableCreation.RowHeight.SetItem(x, hwp.MiliToHwpUnit(representTable_row_list))
    hwp.HParameterSet.HTableCreation.TableProperties.TreatAsChar = 1  # 글자처럼 취급
    hwp.HParameterSet.HTableCreation.TableProperties.Width = hwp.MiliToHwpUnit(1)  # 표 너비
    hwp.HAction.Execute("TableCreate", hwp.HParameterSet.HTableCreation.HSet)

#
def BTN_Empty_Result():
    pLog.append_log("버튼 동작 실행","Empty_Result 실행")
    if Right_Result_Combo.get() == "한글":
        hwp = win32.gencache.EnsureDispatch('HWPFrame.HwpObject')  # 한/글 열기
        hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글')  # 해당 윈도우의 핸들값 찾기

        #win32gui.ShowWindow(hwnd,1)#창 백그라운드에서 실행
        hwp.RegisterModule('FilePathCheckDLL', 'FilePathCheckerModule')
        hwp.XHwpWindows.Item(0).Visible = True  # 숨김해제

        #여백 세팅
        hwp.HAction.GetDefault("ModifySection", hwp.HParameterSet.HSecDef.HSet)
        hwp.HParameterSet.HSecDef.PageDef.LeftMargin = hwp.MiliToHwpUnit(20.0)
        hwp.HParameterSet.HSecDef.PageDef.TopMargin = hwp.MiliToHwpUnit(15.0)
        hwp.HParameterSet.HSecDef.PageDef.RightMargin = hwp.MiliToHwpUnit(20.0)
        hwp.HParameterSet.HSecDef.PageDef.BottomMargin = hwp.MiliToHwpUnit(15.0)
        hwp.HParameterSet.HSecDef.PageDef.HeaderLen = hwp.MiliToHwpUnit(10.0)
        hwp.HParameterSet.HSecDef.PageDef.FooterLen = hwp.MiliToHwpUnit(10.0)
        hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyClass", 24)
        hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyTo", 2)
        hwp.HAction.Execute("ModifySection", hwp.HParameterSet.HSecDef.HSet)
        
        hwp.HAction.Run("ParagraphShapeAlignCenter")
        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
        hwp.HParameterSet.HInsertText.Text = "발주서"
        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
        hwp.HAction.Run("SelectAll")
        hwp_fontSetting(hwp,"HY헤드라인M",22,1)

        
        hwp.HAction.Run("MoveRight")
        hwp.HAction.Run("BreakPara")
        hwp.HAction.Run("MoveRight")
        hwp.HAction.Run("BreakPara")
        #대표 테이블
        hwp_createTable(hwp,2,5,[30,46],6)
        hwp.HAction.Run("TableCellBlock")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp_fontSetting(hwp,"HY헤드라인M",12,2)
        hwp.HAction.Run("TableLeftCell")
        hwp_fontSetting(hwp,"HY헤드라인M",13,1)

        hwp.MovePos(3)
        hwp.HAction.Run("DeleteBack")

        #담당자 테이블
        hwp.HAction.Run("MoveSelPrevWord")
        hwp.HAction.Run("Copy")
        hwp.MovePos(3)
        hwp.HAction.GetDefault("Paste", hwp.HParameterSet.HSelectionOpt.HSet)
        hwp.HAction.Execute("Paste", hwp.HParameterSet.HSelectionOpt.HSet)
        
        hwp.MovePos(3)

        #특이사항
        hwp_createTable(hwp,2,1,[30,132],10)
        
        hwp.HAction.Run("TableCellBlock")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp_fontSetting(hwp,"HY헤드라인M",12,2)
        hwp.HAction.Run("TableLeftCell")
        hwp_fontSetting(hwp,"HY헤드라인M",13,1)

        hwp.MovePos(3)
        hwp.HAction.Run("MoveRight")


        hwp_createTable(hwp,2,1,[30,132],10)
        hwp.HAction.Run("TableCellBlock")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp_fontSetting(hwp,"HY헤드라인M",12,2)
        hwp.HAction.Run("TableLeftCell")
        hwp_fontSetting(hwp,"HY헤드라인M",13,1)

        hwp.MovePos(3)
        hwp.HAction.Run("MoveRight")

        #재료 테이블
        # #표4:번호10, 재료명50, 규격25, 단위15, 수량15, 비고35
        # #listTable_size_var = [24 for _ in range(6)]
        linesize = 20

        hwp_createTable(hwp,6, linesize, [10,50,25,15,15,33],6)
        hwp.HAction.Run("TableCellBlock")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp.HAction.Run("TableCellBlockExtend")
        hwp_fontSetting(hwp,"HY헤드라인M",12,2)

        for x in range(0,linesize):
            hwp.HAction.Run("TableUpperCell")
        hwp_fontSetting(hwp,"HY헤드라인M",13,1)
        
        hwp.MovePos(3)
        hwp.HAction.Run("DeleteBack")
    elif Right_Result_Combo.get() == "엑셀":
        filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=Program_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        pLog.append_log("fileName: ", filename)
        if not filename:
            return
        write_wb = openpyxl.Workbook()
        write_ws = write_wb.active

        write_ws['A1'].font = Font(size=15,bold=True)
        write_ws['A1'] = '발주서'
        write_ws['A1'].border = Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))

        write_ws.append([""])
        write_ws.append(["업체명","","","발주 일자",""])
        write_ws.append(["업체 주소","","","납기 일자",""])
        write_ws.append(["대표명","","","납품 장소",""])
        write_ws.append(["대표 전화번호","","","구매 담당자",""])
        write_ws.append(["대표 메일","","","담당자 전화",""])
        
        write_ws.append([""])
        write_ws.append(["특이 사항"])
        write_ws.append([""])
        
        write_ws.append(["○재료 목록"])
        write_ws.append(["번호","재료명","규격","단위","수량","비고"])

        write_ws.merge_cells("A1:F2")
        write_ws.merge_cells("B9:F9")
        write_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        create_lineBox(write_ws,65,3,71,7)
        create_lineBox(write_ws,65,9,71,10)
        create_lineBox(write_ws,65,12,71,13)
        write_ws.column_dimensions['A'].width = 12
        write_ws.column_dimensions['B'].width = 11
        write_ws.column_dimensions['D'].width = 11
        
        write_wb.save(filename)
        os.startfile(filename)
    elif Right_Result_Combo.get() == "텍스트":
        filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=Program_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".txt", filetypes=[("HWP files", "*.txt")])
        pLog.append_log("fileName: ", filename)
        if not filename:
            return
        data = open(filename, 'w', encoding="UTF8")

        print("발주서\n",file = data)
        print("업체명 : " ,file = data)
        print("업체 주소 : " ,file = data)
        print("대표명 : " ,file = data)
        print("대표 전화번호 : ",file = data)
        print("대표 메일 : " ,file = data)
        print("\n",file = data)
        if RemoveDateCheckbox_Var.get() == 1:
            print("발주 일자 : " + str(Right_InputDate1_Entry_Date.get() +". "+ Right_InputDate1_Entry_Week.get()),file = data)
            print("납기 일자 : " + str(Right_InputDate2_Entry_Date.get() +". "+ Right_InputDate2_Entry_Week.get()),file = data)
        else:
            print("발주 일자 : " + "",file = data)
            print("납기 일자 : " + "",file = data)
        print("납품 장소 : " ,file = data)
        print("구매 담당자 : ",file = data)
        print("담당자 전화 : ",file = data)
        print("\n",file = data)

        print("특이사항 :", file = data)
        print("\n",file = data)

        print("○재료 목록", file = data)
        print("번호, 재료명, 규격, 단위, 수량, 비고\n", file = data)
        os.startfile(filename)

def BTN_Result():
    if treeNumCount >= 2:
        if Right_Result_Combo.get() == "한글":
            global refineListItemName
            BTN_InputDate_Check()#날짜 검증용
            if RemoveTimeCheckbox_Var.get() == 1: #시간값 제거하겠다(old ver)
                filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=Program_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".hwp", filetypes=[("HWP files", "*.hwp")])
                pLog.append_log("fileName: ", filename)
                refine_table()
                hwp = win32.gencache.EnsureDispatch('HWPFrame.HwpObject')  # 한/글 열기
                hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글')  # 해당 윈도우의 핸들값 찾기

                #win32gui.ShowWindow(hwnd,1)#창 백그라운드에서 실행
                hwp.RegisterModule('FilePathCheckDLL', 'FilePathCheckerModule')
                hwp.XHwpWindows.Item(0).Visible = True  # 숨김해제

                #여백 세팅
                hwp.HAction.GetDefault("ModifySection", hwp.HParameterSet.HSecDef.HSet)
                hwp.HParameterSet.HSecDef.PageDef.LeftMargin = hwp.MiliToHwpUnit(20.0)
                hwp.HParameterSet.HSecDef.PageDef.TopMargin = hwp.MiliToHwpUnit(15.0)
                hwp.HParameterSet.HSecDef.PageDef.RightMargin = hwp.MiliToHwpUnit(20.0)
                hwp.HParameterSet.HSecDef.PageDef.BottomMargin = hwp.MiliToHwpUnit(15.0)
                hwp.HParameterSet.HSecDef.PageDef.HeaderLen = hwp.MiliToHwpUnit(10.0)
                hwp.HParameterSet.HSecDef.PageDef.FooterLen = hwp.MiliToHwpUnit(10.0)
                hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyClass", 24)
                hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyTo", 2)
                hwp.HAction.Execute("ModifySection", hwp.HParameterSet.HSecDef.HSet)
                
                hwp.HAction.Run("ParagraphShapeAlignCenter")
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = "발주서"
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("SelectAll")
                hwp_fontSetting(hwp,"HY헤드라인M",22,1)

                
                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("BreakPara")
                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("BreakPara")
                #대표 테이블
                hwp_createTable(hwp,2,5,[30,46],6)
                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp_fontSetting(hwp,"HY헤드라인M",12,2)
                hwp.HAction.Run("TableLeftCell")
                hwp_fontSetting(hwp,"HY헤드라인M",13,1)

                hwp.MovePos(3)
                hwp.HAction.Run("DeleteBack")

                #담당자 테이블
                hwp.HAction.Run("MoveSelPrevWord")
                hwp.HAction.Run("Copy")
                hwp.MovePos(3)
                hwp.HAction.GetDefault("Paste", hwp.HParameterSet.HSelectionOpt.HSet)
                hwp.HAction.Execute("Paste", hwp.HParameterSet.HSelectionOpt.HSet)
                
                hwp.MovePos(3)

                #특이사항
                hwp_createTable(hwp,2,1,[30,130],6)
                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp_fontSetting(hwp,"HY헤드라인M",12,2)
                hwp.HAction.Run("TableLeftCell")
                hwp_fontSetting(hwp,"HY헤드라인M",13,1)

                hwp.MovePos(3)
                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("BreakPara")
                #재료 테이블
                # #표4:번호10, 재료명50, 규격25, 단위15, 수량15, 비고35
                # #listTable_size_var = [24 for _ in range(6)]

                global restoreNum
                if restoreNum < 21 :
                    linesize = 20
                else:
                    linesize = restoreNum

                hwp_createTable(hwp,6, linesize, [10,50,25,15,15,35],6)
                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp_fontSetting(hwp,"HY헤드라인M",12,2)
                for x in range(0,linesize):
                    hwp.HAction.Run("TableUpperCell")
                hwp_fontSetting(hwp,"HY헤드라인M",13,1)

                hwp.MovePos(3)
                hwp.HAction.Run("DeleteBack")

                # ctrl = hwp.HeadCtrl # 첫번째 컨트롤(HaedCtrl)부터 탐색 시작.
                # count = 0
                # while ctrl != None:
                #     nextctrl = ctrl.Next
                #     print(ctrl.CtrlID)
                #     if ctrl.CtrlID == "tbl":
                #         count += 1

                #     ctrl = nextctrl
                # print(count)

                #hwpactionid 기반
                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")
                hwp_tblcount = 0
                while True:
                    NowLocation = hwp.KeyIndicator()
                    #print(NowLocation[-1])
                    if NowLocation[-1] == "표":
                        hwp_tblcount += 1
                    hwp.Run("SelectCtrlFront")
                    LaterLocation = hwp.KeyIndicator()
                    if LaterLocation == NowLocation:
                        #print(hwp_tblcount)
                        break

                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("Cancel")
                #좌측 정보 리스트: 기준명
                ComList = ["업체명","업체 주소","대표명","대표 전화번호","대표 메일"]
                for x in range(0,len(ComList)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ComList[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if x == 4 :
                        break
                    else:
                        hwp.HAction.Run("MoveDown")

                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("Cancel")

                #좌측 정보 리스트: 정보 내용
                ComList1 = [CompanyNameEntry_Var.get(),CompanyAddressEntry_Var.get(),CompanyCEONameEntry_Var.get(),CompanyCEOTelEntry_Var.get(),CompanyMailEntry_Var.get()]
                for x in range(0,len(ComList1)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ComList1[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if x == 4 :
                        break
                    else:
                        hwp.HAction.Run("MoveDown")

                hwp.HAction.Run("MoveRight")
                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("Cancel")

                #우측 정보 리스트: 기준명
                ManagerList = ["발주 일자","납기 일자","납품 장소","구매 담당자","담당자 전화"]
                for x in range(0,len(ManagerList)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ManagerList[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if x == 4 :
                        break
                    else:
                        hwp.HAction.Run("MoveDown")

                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("Cancel")
                
                #우측 정보 리스트: 정보 내용
                if RemoveDateCheckbox_Var.get()==1:    
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = str(Right_InputDate1_Entry.get() +". "+ Right_InputDate1_Entry_Week.get())
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveDown")
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = str(Right_InputDate2_Entry_Date.get() +". "+ Right_InputDate2_Entry_Week.get())
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveDown")
                else:
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ""
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveDown")
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ""
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveDown")
                
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = DeliveryAddressEntry_Var.get()
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("MoveDown")
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = PurchasingManagerNameEntry_Var.get()
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("MoveDown")
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = PurchasingManagerTelEntry_Var.get()
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("MoveRight")

                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("Cancel")

                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = "특이 사항"
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("Cancel")

                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = Right_Uniqueness_Entry.get()
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                hwp.HAction.Run("MoveRight")

                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = "○재료 목록"
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                for x in range(0,4):
                    hwp.HAction.Run("MoveSelLeft")
                hwp.HAction.Run("ParagraphShapeAlignLeft")
                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("DeleteBack")

                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("MoveRight")

                ItemList = ["번호","재료명","규격","단위","수량","비고"]
                for x in range(0,len(ItemList)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ItemList[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveRight")
                
                listcount = 0
                for x in refineList:
                    for y in x:
                        if listcount % 6 == 2:
                            if StandardCheckbox_Var == 1:
                                y = ""
                            else:
                                pass
                        if listcount % 6 == 5:
                            if NoteCheckBox_Var == 1:
                                y = ""
                            else:
                                pass
                        hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HParameterSet.HInsertText.Text = y
                        hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                        hwp.HAction.Run("MoveRight")
                        listcount +=1

                hwp.MovePos(3)
                hwp.HAction.Run("MoveSelPrevWord")
                hwp.HAction.Run("MoveLeft")
                hwp.Run("SelectCtrlFront")
                try:
                    hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
                    hwp.HParameterSet.HShapeObject.TreatAsChar = 0
                    hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
                except:
                    print("서버")
                hwp.MovePos(3)
                
                hwp.SaveAs(filename)
            else:#시간값 제거 안하겠다

                filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=Program_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".hwp", filetypes=[("HWP files", "*.hwp")])
                pLog.append_log("fileName: ", filename)
                refine_table_Time()
                hwp = win32.gencache.EnsureDispatch('HWPFrame.HwpObject')  # 한/글 열기
                hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글')  # 해당 윈도우의 핸들값 찾기

                #win32gui.ShowWindow(hwnd,1)#창 백그라운드에서 실행
                hwp.RegisterModule('FilePathCheckDLL', 'FilePathCheckerModule')
                hwp.XHwpWindows.Item(0).Visible = True  # 숨김해제

                #여백 세팅
                hwp.HAction.GetDefault("ModifySection", hwp.HParameterSet.HSecDef.HSet)
                hwp.HParameterSet.HSecDef.PageDef.LeftMargin = hwp.MiliToHwpUnit(20.0)
                hwp.HParameterSet.HSecDef.PageDef.TopMargin = hwp.MiliToHwpUnit(15.0)
                hwp.HParameterSet.HSecDef.PageDef.RightMargin = hwp.MiliToHwpUnit(20.0)
                hwp.HParameterSet.HSecDef.PageDef.BottomMargin = hwp.MiliToHwpUnit(15.0)
                hwp.HParameterSet.HSecDef.PageDef.HeaderLen = hwp.MiliToHwpUnit(10.0)
                hwp.HParameterSet.HSecDef.PageDef.FooterLen = hwp.MiliToHwpUnit(10.0)
                hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyClass", 24)
                hwp.HParameterSet.HSecDef.HSet.SetItem("ApplyTo", 2)
                hwp.HAction.Execute("ModifySection", hwp.HParameterSet.HSecDef.HSet)
                
                #발주서 제목
                hwp.HAction.Run("ParagraphShapeAlignCenter")
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = "발주서"
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("SelectAll")
                hwp_fontSetting(hwp,"HY헤드라인M",22,1)
                
                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("BreakPara")
                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("BreakPara")

                #대표 테이블
                hwp_createTable(hwp,2,5,[30,46],6)
                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp_fontSetting(hwp,"HY헤드라인M",12,2)
                hwp.HAction.Run("TableLeftCell")
                hwp_fontSetting(hwp,"HY헤드라인M",13,1)

                hwp.MovePos(3)
                hwp.HAction.Run("DeleteBack")

                #담당자 테이블
                hwp.HAction.Run("MoveSelPrevWord")
                hwp.HAction.Run("Copy")
                hwp.MovePos(3)
                hwp.HAction.GetDefault("Paste", hwp.HParameterSet.HSelectionOpt.HSet)
                hwp.HAction.Execute("Paste", hwp.HParameterSet.HSelectionOpt.HSet)
                hwp.MovePos(3)

                #특이사항 테이블
                hwp_createTable(hwp,2,1,[30,130],6)
                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp_fontSetting(hwp,"HY헤드라인M",12,2)
                hwp.HAction.Run("TableLeftCell")
                hwp_fontSetting(hwp,"HY헤드라인M",13,1)
                hwp.MovePos(3)
                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("BreakPara")

                #테이블 크기 정하기
                resNum = 0
                for x in restoreNumList:
                    if x == 0:
                        resNum=resNum+1
                    else:
                        resNum = resNum + x
                resNum += 1        

                if resNum < 21 :
                    linesize = 20
                else:
                    linesize = resNum
                
                #재료 목록 테이블
                hwp_createTable(hwp,6, linesize, [10,50,25,15,15,35],6)
                hwp.HAction.Run("TableCellBlock")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp_fontSetting(hwp,"HY헤드라인M",12,2)
                for x in range(0,linesize):
                    hwp.HAction.Run("TableUpperCell")
                hwp_fontSetting(hwp,"HY헤드라인M",13,1)

                hwp.MovePos(3)
                hwp.HAction.Run("DeleteBack")

                #hwpactionid 기반
                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")
                hwp_tblcount = 0
                while True:
                    NowLocation = hwp.KeyIndicator()
                    #print(NowLocation[-1])
                    if NowLocation[-1] == "표":
                        hwp_tblcount += 1
                    hwp.Run("SelectCtrlFront")
                    LaterLocation = hwp.KeyIndicator()
                    if LaterLocation == NowLocation:
                        #print(hwp_tblcount)
                        break

                hwp.Run("MoveDocBegin")
                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("Cancel")
                #좌측 정보 리스트: 기준명
                ComList = ["업체명","업체 주소","대표명","대표 전화번호","대표 메일"]
                for x in range(0,len(ComList)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ComList[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if x == 4 :
                        break
                    else:
                        hwp.HAction.Run("MoveDown")

                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("Cancel")

                #좌측 정보 리스트: 정보 내용
                ComList1 = [CompanyNameEntry_Var.get(),CompanyAddressEntry_Var.get(),CompanyCEONameEntry_Var.get(),CompanyCEOTelEntry_Var.get(),CompanyMailEntry_Var.get()]
                for x in range(0,len(ComList1)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ComList1[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if x == 4 :
                        break
                    else:
                        hwp.HAction.Run("MoveDown")

                hwp.HAction.Run("MoveRight")
                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("Cancel")

                #우측 정보 리스트: 기준명
                ManagerList = ["발주 일자","납기 일자","납품 장소","구매 담당자","담당자 전화"]
                for x in range(0,len(ManagerList)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ManagerList[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    if x == 4 :
                        break
                    else:
                        hwp.HAction.Run("MoveDown")

                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("Cancel")
                
                #우측 정보 리스트: 정보 내용
                if RemoveDateCheckbox_Var.get()==1:    
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = str(Right_InputDate1_Entry.get() +". "+ Right_InputDate1_Entry_Week.get())
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveDown")
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = str(Right_InputDate2_Entry_Date.get() +". "+ Right_InputDate2_Entry_Week.get())
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveDown")
                else:
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ""
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveDown")
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ""
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveDown")
                
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = DeliveryAddressEntry_Var.get()
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("MoveDown")
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = PurchasingManagerNameEntry_Var.get()
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("MoveDown")
                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = PurchasingManagerTelEntry_Var.get()
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HAction.Run("MoveRight")

                hwp.Run("SelectCtrlFront")
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("Cancel")

                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = "특이 사항"
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("Cancel")

                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = Right_Uniqueness_Entry.get()
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                hwp.HAction.Run("MoveRight")

                hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                hwp.HParameterSet.HInsertText.Text = "○재료 목록"
                hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)

                for x in range(0,4):
                    hwp.HAction.Run("MoveSelLeft")
                hwp.HAction.Run("ParagraphShapeAlignLeft")
                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("DeleteBack")

                hwp.HAction.Run("MoveRight")
                hwp.HAction.Run("MoveRight")

                ItemList = ["번호","재료명","규격","단위","수량","비고"]
                for x in range(0,len(ItemList)):
                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HParameterSet.HInsertText.Text = ItemList[x]
                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                    hwp.HAction.Run("MoveRight")
                

                #refine에 오전,오후,저녁 재료 항목 들어있음
                for z in range(0,len(refineList)):#refineList는 전체값
                    listcount = 0
                    if not refineList[z]:
                        for i in range(0,6):
                            hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                            hwp.HParameterSet.HInsertText.Text = ""
                            hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                            hwp.HAction.Run("MoveRight")
                    else:
                        for x in refineList[z]:#refineList[z]는 1개의 시간대
                                for y in x: #시간대 안에 항목
                                    
                                    if listcount % 6 == 2:
                                        if StandardCheckbox_Var == 1:
                                            y = ""
                                        else:
                                            pass
                                    if listcount % 6 == 5:
                                        if NoteCheckBox_Var == 1:
                                            y = ""
                                        else:
                                            pass
                                    hwp.HAction.GetDefault("InsertText", hwp.HParameterSet.HInsertText.HSet)
                                    hwp.HParameterSet.HInsertText.Text = y
                                    hwp.HAction.Execute("InsertText", hwp.HParameterSet.HInsertText.HSet)
                                    hwp.HAction.Run("MoveRight")
                                    listcount +=1
                
                #재료 테이블 제목 테두리 설정
                hwp.MovePos(3)
                hwp.HAction.Run("MoveSelPrevWord")
                hwp.HAction.Run("MoveLeft")
                hwp.Run("SelectCtrlFront")
                
                hwp.HAction.Run("ShapeObjTableSelCell")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.Run("TableRightCell")
                hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
                hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
                hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

                #실질적인 테이블 테두리 크기
                for x in range(0,len(restoreNumList)):#[14,12,21] x=14
                    if restoreNumList[x] == 0:
                        hwp.HAction.Run("TableLowerCell")
                        hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                        hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
                        hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
                        hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
                        hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
                        hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                    else:
                        for y in range(restoreNumList[x]):
                            hwp.HAction.Run("TableLowerCell")
                        hwp.HAction.GetDefault("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)
                        hwp.HParameterSet.HCellBorderFill.BorderWidthBottom = hwp.HwpLineWidth("0.7mm")
                        hwp.HParameterSet.HCellBorderFill.BorderWidthTop = hwp.HwpLineWidth("0.7mm")
                        hwp.HParameterSet.HCellBorderFill.BorderWidthRight = hwp.HwpLineWidth("0.7mm")
                        hwp.HParameterSet.HCellBorderFill.BorderWidthLeft = hwp.HwpLineWidth("0.7mm")
                        hwp.HAction.Execute("CellBorder", hwp.HParameterSet.HCellBorderFill.HSet)

                hwp.MovePos(3)
                hwp.HAction.Run("MoveSelPrevWord")
                hwp.HAction.Run("MoveLeft")
                hwp.Run("SelectCtrlFront")
                try:
                    hwp.HAction.GetDefault("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
                    hwp.HParameterSet.HShapeObject.TreatAsChar = 0#글자취급x
                    hwp.HAction.Execute("TablePropertyDialog", hwp.HParameterSet.HShapeObject.HSet)
                except:
                    print("서버")
                    #아마도 범위 초과 설정 해서 그런듯
                hwp.MovePos(3)
                hwp.SaveAs(filename)
        elif Right_Result_Combo.get() == "엑셀":
            filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=Program_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            pLog.append_log("fileName: ", filename)
            if not filename:
                return
            if RemoveTimeCheckbox_Var.get() == 1:
                
                refine_table()
                write_wb = openpyxl.Workbook()
                write_ws = write_wb.active
                write_ws['A1'].font = Font(size=15,bold=True)
                write_ws['A1'] = '발주서'
                write_ws['A1'].border = Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))
                
                write_ws.append([""])
                colList = [12, 12.5, 18, 10.5, 13, 25]
                colCount = 0
                for col in range(1,len(colList)+1):
                    write_ws.column_dimensions[get_column_letter(col)].width = colList[colCount]
                    colCount+=1

                if RemoveDateCheckbox_Var.get() == 1:
                    write_ws.append(["업체명",CompanyNameEntry_Var.get(),"","발주 일자",Right_InputDate1_Entry_Date.get()])
                    write_ws.append(["업체 주소",CompanyAddressEntry_Var.get(),"","납기 일자",Right_InputDate2_Entry_Date.get()])
                else:
                    write_ws.append(["업체명",CompanyNameEntry_Var.get(),"","발주 일자",""])
                    write_ws.append(["업체 주소",CompanyAddressEntry_Var.get(),"","납기 일자",""])

                write_ws.append(["대표명",CompanyCEONameEntry_Var.get(),"","납품 장소",DeliveryAddressEntry_Var.get()])
                write_ws.append(["대표 전화번호",CompanyCEOTelEntry_Var.get(),"","구매 담당자",PurchasingManagerNameEntry_Var.get()])
                write_ws.append(["대표 메일",CompanyMailEntry_Var.get(),"","담당자 전화",PurchasingManagerTelEntry_Var.get()])
                write_ws.append([""])
                write_ws.append(["특이 사항",Right_Uniqueness_Entry.get()])
                write_ws.append([""])

                write_ws.append(["○재료 목록"])
                write_ws.append(["번호","재료명","규격","단위","수량","비고"])
                
                listcount = 0
                for i in refineList:
                    for j in i:
                        if listcount % 6 == 2:
                            if StandardCheckbox_Var == 1:
                                i[j] = ""
                            else:
                                pass
                        if listcount % 6 == 5:
                            if NoteCheckBox_Var == 1:
                                i[j] = ""
                            else:
                                pass
                    write_ws.append(i)

                write_ws.merge_cells("A1:F2")
                write_ws.merge_cells("B9:F9")
                write_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                
                create_lineBox(write_ws,65,3,71,7)
                create_lineBox(write_ws,65,9,71,10)
                create_lineBox(write_ws,65,12,71,13)
                create_lineBox(write_ws,65,13,71,13+len(refineList)-1)
                write_ws.column_dimensions['A'].width = 12
                write_ws.column_dimensions['B'].width = 11
                write_ws.column_dimensions['D'].width = 11
                write_wb.save(filename)
            else:
                refine_table_Time()
                write_wb = openpyxl.Workbook()
                write_ws = write_wb.active
                write_ws['A1'].font = Font(size=15,bold=True)
                write_ws['A1'] = '발주서'
                write_ws['A1'].border = Border(left=Side(style="medium"),right=Side(style="medium"),top=Side(style="medium"),bottom=Side(style="medium"))
                
                write_ws.append([""])
                colList = [12, 12.5, 18, 10.5, 13, 25]
                colCount = 0
                for col in range(1,len(colList)+1):
                    write_ws.column_dimensions[get_column_letter(col)].width = colList[colCount]
                    colCount+=1

                if RemoveDateCheckbox_Var.get() == 1:
                    write_ws.append(["업체명",CompanyNameEntry_Var.get(),"","발주 일자",Right_InputDate1_Entry_Date.get()])
                    write_ws.append(["업체 주소",CompanyAddressEntry_Var.get(),"","납기 일자",Right_InputDate2_Entry_Date.get()])
                else:
                    write_ws.append(["업체명",CompanyNameEntry_Var.get(),"","발주 일자",""])
                    write_ws.append(["업체 주소",CompanyAddressEntry_Var.get(),"","납기 일자",""])

                write_ws.append(["대표명",CompanyCEONameEntry_Var.get(),"","납품 장소",DeliveryAddressEntry_Var.get()])
                write_ws.append(["대표 전화번호",CompanyCEOTelEntry_Var.get(),"","구매 담당자",PurchasingManagerNameEntry_Var.get()])
                write_ws.append(["대표 메일",CompanyMailEntry_Var.get(),"","담당자 전화",PurchasingManagerTelEntry_Var.get()])
                write_ws.append([""])
                write_ws.append(["특이 사항",Right_Uniqueness_Entry.get()])
                write_ws.append([""])

                write_ws.append(["○재료 목록"])
                write_ws.append(["번호","재료명","규격","단위","수량","비고"])

                        #refine에 오전,오후,저녁 재료 항목 들어있음
                for z in range(0,len(refineList)):
                    listcount = 0
                    if not refineList[z]:
                        write_ws.append([""])
                    else:
                        for i in refineList[z]:
                            for j in i:
                                if listcount % 6 == 2:
                                    if StandardCheckbox_Var == 1:
                                        i[j] = ""
                                    else:
                                        pass
                                if listcount % 6 == 5:
                                    if NoteCheckBox_Var == 1:
                                        i[j] = ""
                                    else:
                                        pass
                            write_ws.append(i)

                write_ws.merge_cells("A1:F2")
                write_ws.merge_cells("B9:F9")
                write_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                
                create_lineBox(write_ws,65,3,71,7)
                create_lineBox(write_ws,65,9,71,10)
                create_lineBox(write_ws,65,12,71,13)


                past_accumulate = 13
                for i in range(0,len(restoreNumList)):
                    if restoreNumList[i] == 0:
                        restoreNumList[i] = 1
                        create_lineBox_ForList(write_ws,65,past_accumulate,71,past_accumulate+restoreNumList[i])
                        past_accumulate += restoreNumList[i]
                    else:
                        create_lineBox_ForList(write_ws,65,past_accumulate,70,past_accumulate+restoreNumList[i])
                        past_accumulate += restoreNumList[i]

                write_ws.column_dimensions['A'].width = 12
                write_ws.column_dimensions['B'].width = 11
                write_ws.column_dimensions['D'].width = 11
                write_wb.save(filename)

            os.startfile(filename)
            
        elif Right_Result_Combo.get() == "텍스트":
            filename = filedialog.asksaveasfilename(initialfile=datetime.datetime.today().strftime("%Y_%m_%d"),initialdir=Program_Result_Save_Dir_Var.get(), title="Select file",defaultextension=".txt", filetypes=[("TXT files", "*.txt")])
            pLog.append_log("fileName: ", filename)
            if not filename:
                return
            if RemoveTimeCheckbox_Var.get() == 1:
                refine_table()
                data = open(filename, 'w', encoding="UTF8")

                print("발주서\n",file = data)
                print("업체명 : " + CompanyNameEntry_Var.get(),file = data)
                print("업체 주소 : " + CompanyAddressEntry_Var.get(),file = data)
                print("대표명 : " + CompanyCEONameEntry_Var.get(),file = data)
                print("대표 전화번호 : " + CompanyCEOTelEntry_Var.get(),file = data)
                print("대표 메일 : " + CompanyMailEntry_Var.get(),file = data)
                print("\n",file = data)
                if RemoveDateCheckbox_Var.get() == 1:
                    print("발주 일자 : " + str(Right_InputDate1_Entry_Date.get() +". "+ Right_InputDate1_Entry_Week.get()),file = data)
                    print("납기 일자 : " + str(Right_InputDate2_Entry_Date.get() +". "+ Right_InputDate2_Entry_Week.get()),file = data)
                else:
                    print("발주 일자 : " + "",file = data)
                    print("납기 일자 : " + "",file = data)
                print("납품 장소 : " + CompanyCEONameEntry_Var.get(),file = data)
                print("구매 담당자 : " + CompanyCEOTelEntry_Var.get(),file = data)
                print("담당자 전화 : " + CompanyMailEntry_Var.get(),file = data)
                print("\n",file = data)

                print("특이사항 :" + Right_Uniqueness_Entry.get(), file = data)
                print("\n",file = data)

                print("○재료 목록", file = data)
                print("번호, 재료명, 규격, 단위, 수량, 비고\n", file = data)

                listcount = 0
                for i in refineList:
                    for j in i:
                        if listcount % 6 == 2:
                            if StandardCheckbox_Var == 1:
                                i[j] = ""
                            else:
                                pass
                        if listcount % 6 == 5:
                            if NoteCheckBox_Var == 1:
                                i[j] = ""
                            else:
                                pass
                    print(i,file = data)
                
                data.close()

            else:
                refine_table_Time()
                data = open(filename, 'w', encoding="UTF8")

                print("발주서\n",file = data)
                print("업체명 : " + CompanyNameEntry_Var.get(),file = data)
                print("업체 주소 : " + CompanyAddressEntry_Var.get(),file = data)
                print("대표명 : " + CompanyCEONameEntry_Var.get(),file = data)
                print("대표 전화번호 : " + CompanyCEOTelEntry_Var.get(),file = data)
                print("대표 메일 : " + CompanyMailEntry_Var.get(),file = data)
                print("\n",file = data)
                if RemoveDateCheckbox_Var.get() == 1:
                    print("발주 일자 : " + str(Right_InputDate1_Entry_Date.get() +". "+ Right_InputDate1_Entry_Week.get()),file = data)
                    print("납기 일자 : " + str(Right_InputDate2_Entry_Date.get() +". "+ Right_InputDate2_Entry_Week.get()),file = data)
                else:
                    print("발주 일자 : " + "",file = data)
                    print("납기 일자 : " + "",file = data)
                print("납품 장소 : " + CompanyCEONameEntry_Var.get(),file = data)
                print("구매 담당자 : " + CompanyCEOTelEntry_Var.get(),file = data)
                print("담당자 전화 : " + CompanyMailEntry_Var.get(),file = data)
                print("\n",file = data)

                print("특이사항 :" + Right_Uniqueness_Entry.get(), file = data)
                print("\n",file = data)

                print("○재료 목록", file = data)
                print("번호, 재료명, 규격, 단위, 수량, 비고\n", file = data)

                for z in range(0,len(refineList)):
                    listcount = 0
                    if not refineList[z]:
                        print("[,,,,,]",file = data)
                    else:
                        for i in refineList[z]:
                            for j in i:
                                if listcount % 6 == 2:
                                    if StandardCheckbox_Var == 1:
                                        i[j] = ""
                                    else:
                                        pass
                                if listcount % 6 == 5:
                                    if NoteCheckBox_Var == 1:
                                        i[j] = ""
                                    else:
                                        pass

                            print(i,file = data)

                data.close()
            fd=open(filename,"r",encoding="UTF8")
            d=fd.read()
            fd.close()
            m=d.split("\n")
            s="\n".join(m[:-1])
            fd=open(filename,"w+",encoding="UTF8")
            for i in range(len(s)):
                fd.write(s[i])
            fd.close() 
            os.startfile(filename)
#============================================================
#[메인 - 우측패널 - 버튼] 결과 내보내기 HWP_빈문서
#============================================================
#제목: 발주서
#표1:업체명 ,업체 주소 ,대표명, 대표 전화번호, 대표 메일
#표2:납품 장소, 구매 담당자, 담당자 전화
#표3: 특이사항
#표4:번호, 재료명, 규격, 단위, 수량, 비고



#이건 표의 테두리를 굵게 해주는것임
def create_lineBox(ws,_c1,_n1,_c2,_n2): #c1:문자(좌), n1:숫자(좌), c2:문자(우), n2:숫자(우)

    #만약 높이가 한칸일때 
    if _n2 -_n1 == 1:
        for i in range(_c1+1,_c2-1):
            ws[chr(i)+str(_n1)].border = Border(top=Side(style="medium"),bottom=Side(style="medium"))
        ws[chr(_c1)+str(_n1)].border = Border(top=Side(style="medium"),bottom=Side(style="medium"),left=Side(style="medium"))
        ws[chr(_c2-1)+str(_n2-1)].border = Border(top=Side(style="medium"),bottom=Side(style="medium"),right=Side(style="medium"))
    else:
        #상단 라인
        for i in range(_c1,_c2):
            ws[chr(i)+str(_n1)].border = Border(top=Side(style="medium"))
        
        #우측 라인
        for i in range(_n1,(_n2+1)):
            ws[chr(_c2-1)+str(i)].border = Border(right=Side(style="medium"))
        
        #좌측 라인
        for i in range(_n1,(_n2+1)):
            ws[chr(_c1)+str(i)].border = Border(left=Side(style="medium"))

        #하단 라인
        for i in range(_c1,_c2):
            if i == _c1:
                ws[chr(i)+str(_n2)].border = Border(bottom=Side(style="medium"),left=Side(style="medium"))
            elif i == _c2-1:
                ws[chr(i)+str(_n2)].border = Border(bottom=Side(style="medium"),right=Side(style="medium"))
            else:
                ws[chr(i)+str(_n2)].border = Border(bottom=Side(style="medium"))

def create_lineBox_ForList(ws,_c1,_n1,_c2,_n2): #c1:문자(좌), n1:숫자(좌), c2:문자(우), n2:숫자(우)

    #만약 높이가 한칸일때 
    if _n2 -_n1 == 1:
        for i in range(_c1+1,_c2-1):
            ws[chr(i)+str(_n1)].border = Border(top=Side(style="medium"),bottom=Side(style="medium"))
        ws[chr(_c1)+str(_n1)].border = Border(top=Side(style="medium"),bottom=Side(style="medium"),left=Side(style="medium"))
        ws[chr(_c2-1)+str(_n2-1)].border = Border(top=Side(style="medium"),bottom=Side(style="medium"),right=Side(style="medium"))
    else:
        
        #우측 라인
        for i in range(_n1,_n2):
            if i == _n2-1:
                ws[chr(_c2)+str(i)].border = Border(right=Side(style="medium"),bottom=Side(style="medium"))
            else:
                ws[chr(_c2)+str(i)].border = Border(right=Side(style="medium"))
        
        #좌측 라인
        for i in range(_n1,_n2-1):
            ws[chr(_c1)+str(i)].border = Border(left=Side(style="medium"))
        
        #하단 라인
        for i in range(_c1,_c2):
            if i == _c1:
                ws[chr(i)+str(_n2-1)].border = Border(bottom=Side(style="medium"),left=Side(style="medium"))
            else:
                ws[chr(i)+str(_n2-1)].border = Border(bottom=Side(style="medium"))
#============================================================
#[프로그램] 엑셀 테두리 
#============================================================
def set_border(ws, cell_range, inputStyle):
    rows = ws[cell_range]
    for row in rows:
        if row == rows[0][0] or row == rows[0][-1] or row == rows[-1][0] or row == rows[-1][-1]:
            pass
        else:
            row[0].border = Border(left=Side(style=inputStyle))
            row[-1].border = Border(right=Side(style=inputStyle))
        for c in rows[0]:
            c.border = Border(top=Side(style=inputStyle))
        for c in rows[-1]:
            c.border = Border(bottom=Side(style=inputStyle))
    rows[0][0].border = Border(left=Side(style=inputStyle), top=Side(style=inputStyle))
    rows[0][-1].border = Border(right=Side(style=inputStyle), top=Side(style=inputStyle))
    rows[-1][0].border = Border(left=Side(style=inputStyle), bottom=Side(style=inputStyle))
    rows[-1][-1].border = Border(right=Side(style=inputStyle), bottom=Side(style=inputStyle))

class MultiListBox(object):
    def __init__(self):
        self.tree = None

        self._setup_widgets()
        self._build_tree()
            
    def _setup_widgets(self):
        self.tree =ttk.Treeview(columns=treeColumn_header, height=27, show="headings")
        
        vsb = Scrollbar(orient="vertical", command=self.tree.yview)
        hsb = Scrollbar(orient="horizontal", command=self.tree.xview)
        
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(column=0, row=0, sticky='nsew', in_=L_frame)
        
        vsb.grid(column=1, row=0, sticky='ns', in_=L_frame)
        hsb.grid(column=0, row=1, sticky='ew', in_=L_frame)
        
        L_frame.grid_columnconfigure(0, weight=1)
        L_frame.grid_rowconfigure(0, weight=1)


    def _build_tree(self):
        Phase_Header_Size = [50,60,230,80,60]
        colCount = 0
        for col in treeColumn_header:
            self.tree.heading(col, text=col.title(), command=lambda c=col: sortby(self.tree, c, 0))
            # adjust the column's width to the header string
            #self.tree.column(col, width=tkFont.Font().measure(col.title()))
            self.tree.column(col, width=Phase_Header_Size[colCount])
            colCount += 1
        itemNum = 0
        for item in treeItem_List:
            #self.tree.insert('', 'end', values=item)
            #self.tree.insert('', index=itemNum, iid=itemNum, values=item)
            #print(itemNum)
            # adjust column's width if necessary to fit each value
            # testcode: self.treeview.insert('', 'end', text="Item_"+str(self.i),
            #             values=(self.dose_entry.get() + " mg",
            #                     self.modified_entry.get()))
            # Increment counter
            #self.i = self.i + 1 :testcode
            itemNum+=1
            for ix, val in enumerate(item):
                #col_w = tkFont.Font().measure(val)
                col_w = 10
                if self.tree.column(treeColumn_header[ix],width=None)<col_w:
                    self.tree.column(treeColumn_header[ix], width=col_w)

#============================================================
#[프로그램] 날짜 업데이트
#============================================================
def update_date(event):
    Right_InputDate1_Entry_Week.set(days[datetime.date(int(Right_InputDate1_Entry_Date.get().split(".")[0]),int(Right_InputDate1_Entry_Date.get().split(".")[1]),int(Right_InputDate1_Entry_Date.get().split(".")[2])).weekday()])
    Right_InputDate2_Entry_Week.set(days[datetime.date(int(Right_InputDate2_Entry_Date.get().split(".")[0]),int(Right_InputDate2_Entry_Date.get().split(".")[1]),int(Right_InputDate2_Entry_Date.get().split(".")[2])).weekday()])
    
#============================================================
#[프로그램] 트리값 정렬
#============================================================
def sortby(tree, col, sortVal):
    data = [(tree.set(child, col), child) for child in tree.get_children('')]

    if data[0][0].isdigit() == True:
        data.sort(reverse=sortVal,key= lambda x : int(x[0]))
    else:
        data.sort(reverse=sortVal)
    for ix, item in enumerate(data):
        tree.move(item[1], '', ix)
    tree.heading(col, command=lambda col=col: sortby(tree, col, int(not sortVal)))

#============================================================
#[프로그램] 사용자 이름 이벤트
#============================================================
def EnterNameInfoLabel(event):
    Top_Name_info_Label_w.config(bg="lightyellow")
    Top_Name_info_Label.config(bg="lightyellow")

def LeaveNameInfoLabel(event):
    Top_Name_info_Label_w.config(bg="lightgray")
    Top_Name_info_Label.config(bg="lightgray")

#TODO:무 개수 정의 파싱
#============================================================
#[프로그램] MainLoop
#============================================================
if __name__ == "__main__":
    root = TkinterDnD.Tk()
    root.title("발주 종합")
    
    root.geometry("800x660+550+200")
    root.option_add("*Font","맑은고딕 12")
    root.resizable(False, False) #창 사이즈 변경 불가능
    root.iconbitmap(os.getcwd() + '\\res\\icon.ico')
    root.attributes('-topmost', True)
    root.attributes('-topmost', False)
    pLog = GUIT()
    # 1 시작시 보이는 메세지 'append_log'
    pLog.append_log("",'프로그램을 시작했습니다.')
    pLog.append_log("작업 파일",os.getcwd())
    SearchRadioValue = IntVar(None,1) #RIGHT_Frame 라디오 버튼 선택중인 값
    days = ['월요일','화요일','수요일','목요일','금요일','토요일','일요일']
    
    todayvalue = datetime.datetime.today().strftime("%Y.%m.%d") 
    todayDays = datetime.datetime.today().weekday()
    #C_days = todayvalue +"."+ days[todayDays]

    NowSettingFile = StringVar() #현재 사용자

    CompanyNameEntry_Var = StringVar()
    CompanyAddressEntry_Var = StringVar()
    CompanyCEONameEntry_Var = StringVar()
    CompanyCEOTelEntry_Var = StringVar()
    CompanyMailEntry_Var = StringVar()
    DeliveryAddressEntry_Var = StringVar()
    PurchasingManagerNameEntry_Var = StringVar()
    PurchasingManagerTelEntry_Var = StringVar()

    RemoveReduplicationCheckbox_Var=IntVar()#중복 제거
    
    TimeKind_Var = StringVar()
    RemoveTimeCheckbox_Var = IntVar()#시간값 제거

    RemoveDateCheckbox_Var=IntVar()#날짜 반영 여부
    ToolTipCheckbox_Var = IntVar()

    StandardCheckbox_Var = IntVar()#규격
    NoteCheckBox_Var = IntVar()#비고

    ListofUserInfo = []
    DARKMODE_VAR = IntVar()
    USERNAMEDIR_VAR = ""
    
    restoreNum = 0
    DATE_TOTAL_IMPORT_COMBO_INIT = 0
    DATE_TOTAL_EXPORT_COMBO_INIT = 0
    RESULT_COMBO_VAR = 0

    Right_InputDate1_Entry_Date = StringVar()
    Right_InputDate1_Entry_Date.set(todayvalue)
    Right_InputDate2_Entry_Date = StringVar()
    Right_InputDate2_Entry_Date.set(todayvalue)

    Program_Save_Dir_Var = StringVar()
    Program_Result_Save_Dir_Var = StringVar()
    TotalDate_Result_Save_Dir_Var = StringVar()
    Program_UseFile_Dir_Var = StringVar()
    #root.file = filedialog.askopenfile(initialdir='path', title='select file', filetypes=(('jpeg files', '*.jgp'), ('all files', '*.*')))
    #root_dir = "C:/Users/nsn04/OneDrive/바탕 화면/integrative/FolderList"

    #============================================================
    #[프로그램]System init
    #============================================================
    #Appdata - gen_py - delete
    path = os.getenv('APPDATA')
    appdataPath = path[:path.rfind("\\")]
    genPath = "\\Local\\Temp\\gen_py"
    app_genPath = appdataPath + genPath
    if os.path.isdir(app_genPath) == True:
        shutil.rmtree(app_genPath)
        pLog.append_log("실행init [Appdata-gen_py제거]:", path)
    else:
        pLog.append_log("실행init [Appdata-gen_py제거]:",'폴더 없음')
        pass
    #==============================
    #Folder init
    #==============================
    #res - init
    if os.path.isdir("./res"):
        pLog.append_log("실행init [res_init]:", "res폴더 존재")
    else:
        pLog.append_log("실행init [res_init]:", "res 없음-폴더 생성")
        os.mkdir("./res")
    #res - sys - init
    if os.path.isdir("./res/sys"):
        pLog.append_log("실행init [res_sys_init]:", "res_sys 폴더 존재")
    else:
        pLog.append_log("실행init [res_sys_init]:", "res_sys 없음-폴더 생성")
        os.mkdir("./res/sys")
    #res - user - init
    if os.path.isdir("./res/user"):
        pLog.append_log("실행init [res_user_init]:", "res_user 폴더 존재")
    else:
        pLog.append_log("실행init [res_user_init]:", "res_user 없음-폴더 생성")
        os.mkdir("./res/user")
    #res - sys - file[sys_date] - init
    if os.path.isfile("./res/sys/sys_date.txt"):
        pLog.append_log("실행init [sys_date.txt]:", "sys_date.txt 파일 존재")
    else:
        pLog.append_log("실행init [sys_date.txt]:", "sys_date.txt 파일 없음-생성")
        f = open("./res/sys/sys_date.txt", "w")
        f.close()
    #res - sys - file[systemp] - init
    #readsplit_text[0] : 0:다크모드 비활성, 1:다크모드 활성
    #readsplit_text[1] : 등록된 사용자 경로
    #readsplit_text[2] : 규격 체크
    #readsplit_text[3] : 비고 체크
    #readsplit_text[4] : 시간값 제거 체크
    #readsplit_text[5] : 날짜 반영 체크
    #readsplit_text[6] : 중복 제거 체크
    #readsplit_text[7] : 시간값 종류
    #readsplit_text[8] : 툴팁 사용
    #readsplit_text[9] : 요일별 종합 - 불러올 값 콤보박스[0,1,2]
    #readsplit_text[10] : 요일별 종합 - 내보낼 값 콤보박스[0,1,2]
    #readsplit_text[11] : 메인 결과 내보내기 경로 설정
    #readsplit_text[12] : 프로그램 작업 목록 저장 경로 설정
    #readsplit_text[13] : 요일별 종합 결과 확인 저장 경로 설정
    #readsplit_text[14] : 메인 - 결과 콤보박스 값
    #readsplit_text[15] : 프로그램에서 사용할 파일이 있는 폴더 경로 설정
    if os.path.isfile('./res/sys/systemp.txt') == True: #시스템 파일이 존재하면
        openfile = open('./res/sys/systemp.txt','r',encoding="utf-8")
        readtext = openfile.read()
        readsplit_text = readtext.split("\n")
        openfile.close()

        DARKMODE_VAR.set(readsplit_text[0])
        
        StandardCheckbox_Var.set(readsplit_text[2])
        NoteCheckBox_Var.set(readsplit_text[3])
        RemoveTimeCheckbox_Var.set(readsplit_text[4])
        RemoveDateCheckbox_Var.set(readsplit_text[5])
        RemoveReduplicationCheckbox_Var.set(readsplit_text[6])
        TimeKind_Var.set(readsplit_text[7])
        ToolTipCheckbox_Var.set(readsplit_text[8])
        DATE_TOTAL_IMPORT_COMBO_INIT = readsplit_text[9]
        DATE_TOTAL_EXPORT_COMBO_INIT = readsplit_text[10]

        #TODO:파일 경로가 있는지 확인할것
        if os.path.isdir(readsplit_text[11]) == True:
            Program_Save_Dir_Var.set(readsplit_text[11])
        else:
            Program_Save_Dir_Var.set(os.getcwd()+"\\Result\\작업 목록 데이터")

        if os.path.isdir(readsplit_text[12]) == True:
            Program_Result_Save_Dir_Var.set(readsplit_text[12])
        else:
            Program_Result_Save_Dir_Var.set(os.getcwd()+"\\Result\\최종 결과")

        if os.path.isdir(readsplit_text[13]) == True:
            TotalDate_Result_Save_Dir_Var.set(readsplit_text[13])
        else:
            TotalDate_Result_Save_Dir_Var.set(os.getcwd()+"\\Result\\요일 종합")

        RESULT_COMBO_VAR = readsplit_text[14]

        if os.path.isdir(readsplit_text[15]) == True:
            Program_UseFile_Dir_Var.set(readsplit_text[15])
        else:
            Program_UseFile_Dir_Var.set(os.getcwd()+"\\FolderList")

        if readsplit_text[1] != "":
            USERNAMEDIR_VAR = readsplit_text[1]

            strx=str(USERNAMEDIR_VAR)
            fileName = os.path.basename(strx)
            fileName1 = fileName[:-4]
            C_fileName = re.sub(r"[^a-zA-Z|가-힣|/※~()]","",fileName1) 
            C_fileName = C_fileName + " 님"
            NowSettingFile.set(C_fileName)
            
            user_filePath = "./res/user/"+fileName
            user_file = open(user_filePath,'r',encoding='utf-8')
            rdr = csv.reader(user_file)

            lstr = []
            for line in rdr:
                lstr.append("".join(line)) 
                
            CompanyNameEntry_Var.set(lstr[0])
            CompanyAddressEntry_Var.set(lstr[1])
            CompanyCEONameEntry_Var.set(lstr[2])
            CompanyCEOTelEntry_Var.set(lstr[3])
            CompanyMailEntry_Var.set(lstr[4])
            DeliveryAddressEntry_Var.set(lstr[5])
            PurchasingManagerNameEntry_Var.set(lstr[6])
            PurchasingManagerTelEntry_Var.set(lstr[7])
            user_file.close()
            for x in range(0,len(readsplit_text)):
                pLog.append_log("systemp readsplit_text: {}".format(x), readsplit_text[x])
        else:
            tkinter.messagebox.showwarning("사용자 등록", "사용자 등록이 되어있지 않습니다.\n등록해주세요.")
    else:
        pLog.append_log("실행init [systemp.txt]:", "systemp.txt 없음 - 생성")
        tkinter.messagebox.showwarning("사용자 등록", "사용자 등록이 되어있지 않습니다.\n등록해주세요.")
        inituserfile = open("./res/sys/systemp.txt", 'w')
        print(0,file=inituserfile)#0
        print("",file=inituserfile)#1
        print(0,file=inituserfile)#2
        print(0,file=inituserfile)#3
        print(0,file=inituserfile)#4
        print(0,file=inituserfile)#5
        print(0,file=inituserfile)#6
        print("오전,오후,저녁",file=inituserfile)#7
        print(0,file=inituserfile)#8
        print(0,file=inituserfile)#9
        print(0,file=inituserfile)#10
        print(os.getcwd()+"\\Result\\작업 목록 데이터",file=inituserfile)#11
        print(os.getcwd()+"\\Result\\최종 결과",file=inituserfile)#12
        print(os.getcwd()+"\\Result\\요일 종합",file=inituserfile)#13
        print(0,file=inituserfile)#14
        print(os.getcwd()+"\\FolderList",file=inituserfile)#15
        inituserfile.close()
    #==============================
    #res - banList - init
    if os.path.isdir("./res/banList"):
        pass
    else:
        os.mkdir("./res/banList")
    #res - banList - file[BanListFile] - init
    if os.path.isfile("./res/banList/BanListFile.txt"):
        pass
    else:
        f = open("./res/banList/BanListFile.txt", 'w')
        f.close()
    #==============================
    #FolderList - init
    if os.path.isdir("./FolderList"):
        pass
    else:
        os.mkdir("./FolderList")
    #==============================
    #Result - init
    if os.path.isdir("./Result"):
        pass
    else:
        os.mkdir("./Result")
    if os.path.isdir("./Result/요일 종합"):
        pass
    else:
        os.mkdir("./Result/요일 종합")
    if os.path.isdir("./Result/작업 목록 데이터"):
        pass
    else:
        os.mkdir("./Result/작업 목록 데이터")
    if os.path.isdir("./Result/최종 결과"):
        pass
    else:
        os.mkdir("./Result/최종 결과")
    #==============================
    #TempFileList - init
    if os.path.isdir("./TempFileList"):
        pass
    else:
        os.mkdir("./TempFileList")

    

    #============================================================
    #[메인 프레임] init
    #============================================================
    topFrame= Frame(root,width=800,height=45, relief="solid", bd=1, background="darkgray")
    topFrame.pack()
    
    L_frame=Frame(root, width=500,height=595, relief="solid", bd=1)
    L_frame.pack()
    L_frame.place(x=0,y=45)

    R_frame=tkinter.Frame(root,width=300,height=595,background="lightblue")
    R_frame.pack() 
    R_frame.place(x=500,y=45)

    # Test_Button=tkinter.Button(topFrame, text='test', command=testCode1, overrelief="solid", width=8, repeatdelay=1000, repeatinterval=100)
    # Test_Button.place(x=215,y=00)
    
    #============================================================
    #[탑 패널] init
    #============================================================
    Top_Name_info_Frame = Frame(topFrame,width=180,height=30, relief="solid", bd=1, bg="lightgray")
    Top_Name_info_Frame.pack()
    Top_Name_info_Frame.place(x=5,y=8)
    Top_Name_info_Frame_ttp = CreateToolTip(Top_Name_info_Frame, "현재 사용자를 보여줍니다.\n클릭시 사용자 등록창이 실행됩니다.")
    
    Top_Name_info_Label_w = Label(Top_Name_info_Frame,text="현재 사용자:",bg="lightgray")
    Top_Name_info_Label_w.bind('<Enter>', EnterNameInfoLabel)
    Top_Name_info_Label_w.bind('<Leave>', LeaveNameInfoLabel)
    Top_Name_info_Label_w.bind('<Button-1>',BTN_Regist_click)
    Top_Name_info_Label_w.pack()
    Top_Name_info_Label_w.place(x=1,y=3)

    Top_Name_info_Label = Label(Top_Name_info_Frame,textvariable=NowSettingFile,bg="lightgray")
    Top_Name_info_Label.bind('<Enter>', EnterNameInfoLabel)
    Top_Name_info_Label.bind('<Leave>', LeaveNameInfoLabel)
    Top_Name_info_Label.bind('<Button-1>',BTN_Regist_click)
    Top_Name_info_Label.pack()
    Top_Name_info_Label.place(x=96,y=3)
    
    Top_Combo = ttk.Combobox(topFrame,width=6, state="readonly")
    Top_Combo['values']=("번호", "수량", "이름")
    Top_Combo.current(2)
    Top_Combo.place(x=460,y=10,height=26)
    Top_Combo_ttp = CreateToolTip(Top_Combo, "검색시 기준값 입니다.")

    Top_Search_Entry = Entry(topFrame, width=24)
    Top_Search_Entry.place(x=530,y=10,height=25)
    #tip.bind_widget(Right_Search_Entry, balloonmsg = "검색할 단어를 입력해주세요.")
    Top_Search_Entry.bind('<Return>',BTN_SearchItem_Return)
    Top_Search_Entry_ttp = CreateToolTip(Top_Search_Entry, "검색창 입니다.")

    Top_Search_Button = tkinter.Button(topFrame, text='검색', command=BTN_SearchItem, overrelief="solid", width=6)
    Top_Search_Button.place(x=730,y=10)
    

    #============================================================
    #[좌측 패널] init
    #============================================================
    listbox = MultiListBox()
    listbox.tree.drop_target_register(DND_FILES)
    listbox.tree.dnd_bind("<<Drop>>",drag_n_drop_AddItem)
    
    #============================================================
    #[우측 패널] init
    #============================================================
    #============================================================
    #항목 패널
    #============================================================
    Right_List_Text = Label(R_frame, text="【항목】", background="lightblue")
    Right_List_Text.pack()
    Right_List_Text.place(x=5,y=15)
    
    Right_AddItem_Button=tkinter.Button(R_frame, text='항목 추가', command=BTN_AddItem, overrelief="solid", width=8, repeatdelay=1000, repeatinterval=100)
    Right_AddItem_Button.place(x=15,y=40)
    Right_AddItem_Button_ttp = CreateToolTip(Right_AddItem_Button, "좌측 항목 추가 버튼 입니다.")
    #insert(몇번째 자리에, 넣을 값)
    #pop(x번쨰 인덱스값 삭제)
    #self.tree.insert('', 'end', values=item)
    
    #listbox.tree.insert(parent='', index=15, iid=15, text='', values=('5','Manjeet','Echo','test',5))
    #listbox.tree.insert(parent='', index=16, iid=16, text='', values=('6','ssss','Echo','test',9))
    #listbox.tree.insert('',index=15, iid=15, values=('21','sttte','cees','gehe','s'))

    Right_Remove_Button = tkinter.Button(R_frame, text='항목 삭제', command=BTN_RemoveItem, overrelief="solid", width=8)
    Right_Remove_Button.place(x=110,y=40)
    Right_Remove_Button_ttp = CreateToolTip(Right_Remove_Button, "좌측 항목 삭제 버튼 입니다.")

    Right_ALLRemove_Button = tkinter.Button(R_frame, text='전체 삭제', command= BTN_ALL_RemoveItem, overrelief="solid", width=8)
    Right_ALLRemove_Button.place(x=205,y=40)
    Right_ALLRemove_Button_ttp = CreateToolTip(Right_ALLRemove_Button, "좌측 항목 전체 삭제 버튼 입니다.")

    Right_Open_Dir_Button = tkinter.Button(R_frame, text='폴더 열기', command= BTN_Start_AddItem, overrelief="solid", width=8)
    Right_Open_Dir_Button.place(x=205,y=10)
    Right_Open_Dir_Button_ttp = CreateToolTip(Right_Open_Dir_Button, "좌측 항목 전체 삭제 버튼 입니다.")

    Right_Frame_Sep1 = ttk.Separator(R_frame, orient="horizontal")	
    Right_Frame_Sep1.place(relx=0, rely=0.13, relwidth=1, relheight=0)

    #============================================================
    #수량 입력 패널
    #============================================================
    Right_InputNum_Text = Label(R_frame, text="【수량 입력】", background="lightblue")
    Right_InputNum_Text.place(x=5,y=90)
    
    Right_BanList_Button = tkinter.Button(R_frame, text='제외할 재료', command=MenuBTN_OpenBan, overrelief="solid", width=10)
    Right_BanList_Button.place(x=190,y=90)
    Right_BanList_Button_ttp = CreateToolTip(Right_BanList_Button, "제외할 재료를 설정 할수 있는 창이 실행됩니다.")
     
    Right_InputNum_Entry = Entry(R_frame, width=25)
    Right_InputNum_Entry.place(x=15,y=120)
    Right_InputNum_Entry.bind('<FocusIn>',focus_InputNum)
    Right_InputNum_Entry.bind('<Return>',BTN_InputNum_Return)

    Right_InputNum_Button = tkinter.Button(R_frame, text='확인', command=BTN_InputNum, overrelief="solid", width=6)
    Right_InputNum_Button.place(x=225,y=120)
    Right_InputNum_Button_ttp = CreateToolTip(Right_InputNum_Button, "버튼을 누를시 해당 항목에 대해 수량이 설정됩니다.")

    Right_Frame_Sep2 = ttk.Separator(R_frame, orient="horizontal")	
    Right_Frame_Sep2.place(relx=0, rely=0.26, relwidth=1, relheight=0)
    #============================================================
    #날짜 패널
    #============================================================
    Right_InputDate_Text = Label(R_frame, text="【날짜 입력】", background="lightblue")
    Right_InputDate_Text.place(x=5,y=170)

    Right_Date_CheckBox=Checkbutton(R_frame,text="날짜 반영",variable=RemoveDateCheckbox_Var,background="lightblue")
    Right_Date_CheckBox.place(x=195,y=170)
    Right_Date_CheckBox_ttp = CreateToolTip(Right_Date_CheckBox, "출력시 날짜를 반영합니다.")

    Right_InputDate1_Entry = Entry(R_frame, width=16, textvariable=Right_InputDate1_Entry_Date)
    Right_InputDate1_Entry.place(x=15,y=200)
    Right_InputDate1_Entry_ttp = CreateToolTip(Right_InputDate1_Entry, "발주 일자 입니다.")

    Right_InputDate1_Entry_Week = StringVar()
    Right_InputDate1_Entry_Week.set(days[datetime.date(int(Right_InputDate1_Entry_Date.get().split(".")[0]),int(Right_InputDate1_Entry_Date.get().split(".")[1]),int(Right_InputDate1_Entry_Date.get().split(".")[2])).weekday()])
    
    Right_InputDate1_Entry1 = Label(R_frame, width=7, textvariable=Right_InputDate1_Entry_Week, background="white", relief="groove")
    Right_InputDate1_Entry1.place(x=150,y=200)
    Right_InputDate1_Entry.bind('<Return>',update_date)
    Right_InputDate1_Entry.bind('<FocusOut>',update_date)
    
    Connect_Text = Label(R_frame, text="~", background="lightblue")
    Connect_Text.place(x=110,y=220)
    
    Right_InputDate2_Entry = Entry(R_frame, width=16, textvariable=Right_InputDate2_Entry_Date)
    Right_InputDate2_Entry.place(x=15,y=240)
    Right_InputDate2_Entry_ttp = CreateToolTip(Right_InputDate2_Entry, "납기 일자 입니다.")

    Right_InputDate2_Entry_Week = StringVar()
    Right_InputDate2_Entry_Week.set(days[datetime.date(int(Right_InputDate2_Entry_Date.get().split(".")[0]),int(Right_InputDate2_Entry_Date.get().split(".")[1]),int(Right_InputDate2_Entry_Date.get().split(".")[2])).weekday()])

    Right_InputDate2_Entry1 = Label(R_frame, width=7, textvariable=Right_InputDate2_Entry_Week, background="white", relief="groove")
    Right_InputDate2_Entry1.place(x=150,y=240)
    Right_InputDate2_Entry.bind('<Return>',update_date)
    Right_InputDate2_Entry.bind('<FocusOut>',update_date)

    Right_InputDate_Button = tkinter.Button(R_frame, text='입력\n확인', command=BTN_InputDate, overrelief="solid", width=6, height= 3)
    Right_InputDate_Button.place(x=225,y=200)

    Right_InputDate_infoText = Label(R_frame,font=("Arial",8), text="※ 엔터 혹은 버튼을 클릭하면 요일이 정확히 반영 됩니다" ,background="lightblue",foreground="gray")
    Right_InputDate_infoText.place(x=10,y=265)

    Right_Frame_Sep3 = ttk.Separator(R_frame, orient="horizontal")	
    Right_Frame_Sep3.place(relx=0, rely=0.49, relwidth=1, relheight=0)

    #============================================================
    #시간대 입력 패널
    #============================================================
    Right_InputTime_Text = Label(R_frame, text="【시간대 입력】", background="lightblue")
    Right_InputTime_Text.place(x=5,y=305)

    tempKindTime=TimeKind_Var.get().split(",")
    for x in range(0,len(tempKindTime)):
        tempKindTime[x]=tempKindTime[x].replace(" ","")
    Right_Time_Combo = ttk.Combobox(R_frame,width=6, state="readonly")
    Right_Time_Combo['values']=tempKindTime
    Right_Time_Combo.current(0)
    Right_Time_Combo.place(x=150,y=305,height=26)
    Right_Time_Combo_ttp = CreateToolTip(Right_Time_Combo, "기본 시간대 이외의 시간 입력이 가능합니다.")

    Right_InputTimeNum_Button = tkinter.Button(R_frame, text='확인', command=BTN_InputTimeNum, overrelief="solid", width=6)
    Right_InputTimeNum_Button.place(x=225,y=305)

    Right_Time_CheckBox=Checkbutton(R_frame,text="시간값 제거",variable=RemoveTimeCheckbox_Var,background="lightblue")
    Right_Time_CheckBox.place(x=15,y=330)
    Right_Time_CheckBox_ttp = CreateToolTip(Right_Time_CheckBox, "시간값에 따른 출력 명령어 처리를 합니다.")

    Right_Frame_Sep4 = ttk.Separator(R_frame, orient="horizontal")
    Right_Frame_Sep4.place(relx=0, rely=0.66, relwidth=1, relheight=0)

    #============================================================
    #특이사항 패널
    #============================================================
    Right_Uniqueness_Text = Label(R_frame, text="【특이 사항】", background="lightblue")
    Right_Uniqueness_Text.pack()
    Right_Uniqueness_Text.place(x=5,y=405)

    Right_Uniqueness_Entry = Entry(R_frame, width=34)
    Right_Uniqueness_Entry.place(x=15,y=440)
    #tip.bind_widget(Right_Uniqueness_Entry, balloonmsg = "발주서에 입력될 특이사항이 있으면 입력해주세요.")
      
    Right_Uniqueness_Button = tkinter.Button(R_frame, text='입력 확인', command=BTN_UniquenessText, overrelief="solid", width=9)
    Right_Uniqueness_Button.place(x=200,y=405)
    Right_Uniqueness_Button_ttp = CreateToolTip(Right_Uniqueness_Button, "입력 확인용 입니다.")

    Right_Frame_Sep5 = ttk.Separator(R_frame, orient="horizontal")	
    Right_Frame_Sep5.place(relx=0, rely=0.79, relwidth=1, relheight=0)

    #============================================================
    #결과 내보내기 패널
    #============================================================
    Right_Export_Text = Label(R_frame, text="【결과 내보내기】", background="lightblue")
    Right_Export_Text.place(x=5,y=480)

    Right_CheckBox=Checkbutton(R_frame,text="중복 제거",variable=RemoveReduplicationCheckbox_Var,background="lightblue")
    Right_CheckBox.place(x=195,y=480)
    Right_CheckBox_ttp = CreateToolTip(Right_CheckBox, "재료 이름 중복을 제거합니다.(제거된 값은 합쳐 계산됩니다.)")

    Right_CheckBox1=Checkbutton(R_frame,text="규격 제거",variable=StandardCheckbox_Var,background="lightblue")
    Right_CheckBox1.place(x=15,y=510)
    Right_CheckBox1_ttp = CreateToolTip(Right_CheckBox1, "규격 값을 제거합니다.")

    Right_CheckBox2=Checkbutton(R_frame,text="비고 제거",variable=NoteCheckBox_Var,background="lightblue")
    Right_CheckBox2.place(x=115,y=510)
    Right_CheckBox2_ttp = CreateToolTip(Right_CheckBox2, "비고 값을 제거합니다.")

    Right_Empty_Export_Button = tkinter.Button(R_frame, text='E', command=BTN_Empty_Result, overrelief="solid")
    Right_Empty_Export_Button.place(x=15,y=560)
    Right_Empty_Export_Button_ttp = CreateToolTip(Right_Empty_Export_Button, "빈 양식을 생성합니다.")

    Right_Result_Combo = ttk.Combobox(R_frame,width=6, state="readonly")
    Right_Result_Combo['values']=["한글","엑셀","텍스트"]
    Right_Result_Combo.current(RESULT_COMBO_VAR)
    Right_Result_Combo.place(x=40,y=560,height=26)

    Right_Result_Export_Button = tkinter.Button(R_frame, text='결과 확인', command=BTN_Result, overrelief="solid", width=8)
    Right_Result_Export_Button.place(x=115,y=560)
    Right_Result_Export_Button_ttp = CreateToolTip(Right_Result_Export_Button, "해당 양식으로 결과물을 출력합니다.")

    Right_TotalResult_Export_Button = tkinter.Button(R_frame, text='요일 종합', command=BTN_Total_days, overrelief="solid", width=8)
    Right_TotalResult_Export_Button.place(x=210,y=560)
    Right_TotalResult_Export_Button_ttp = CreateToolTip(Right_TotalResult_Export_Button, "요일별 종합 창을 엽니다.")

    #============================================================
    #프로그램- 메뉴바
    #============================================================
    menubar = Menu(root)
    menu1 = Menu(menubar,tearoff=0)
    menu1.add_command(label="새 파일(Ctrl+N)",command=BTN_ALL_RemoveItem_RootMenu)
    menu1.add_command(label="작업 목록 불러오기(Ctrl+O)",command=BTN_LoadList_txt)
    menu1.add_command(label="작업 목록 저장하기(Ctrl+S)",command=BTN_SaveList_txt)
    menu1.add_separator()
    menu1.add_command(label="파일 변환",command= BTN_ConvertHWP)
    menu1.add_command(label="요일별 종합",command= BTN_Total_days)
    menu1.add_separator()
    menu1.add_command(label="창 닫기", command=root.destroy)
    menubar.add_cascade(label="파일",menu=menu1)

    menu2 = Menu(menubar,tearoff=0)
    menu2.add_command(label="제외 시킬 재료 설정",command= MenuBTN_OpenBan)
    menu2.add_command(label="사용자 설정",command= BTN_Regist)
    menu2.add_separator()
    menu2.add_command(label="프로그램 설정",command= BTN_SettingProgram)
    menu2.add_command(label="프로그램 초기화",command= BTN_SettingReset)
    menubar.add_cascade(label="설정",menu=menu2)

    menu3 = Menu(menubar,tearoff=0)
    menu3.add_command(label="간단 사용법",command= BTN_HOW_TO_USE)
    menu3.add_command(label="상세 사용법(F1)",command= BTN_Net)
    menubar.add_cascade(label="도움말",menu=menu3)

    root.bind_all("<F1>", BTN_Net_Key)
    root.bind_all("<Control-s>", BTN_SaveList_txt_Key)
    root.bind_all("<Control-o>", BTN_LoadList_txt_Key)
    root.bind_all("<Control-n>", BTN_ALL_RemoveItem_RootMenu_Key)

    BTN_DarkMode()
    root.config(menu=menubar)
    root.mainloop()

    pLog.run()


#program 종료시
atexit.register(exit_Function)
